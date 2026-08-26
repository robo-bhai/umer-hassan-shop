from django.contrib import admin
from django.utils.timezone import now
import shutil
from django.conf import settings
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from import_export.admin import ImportExportModelAdmin
from .models import *
from django.db.models import Sum, F, FloatField, DecimalField, ExpressionWrapper
from io import BytesIO
import datetime 
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.urls import reverse
from django.utils.html import format_html
from datetime import datetime, timedelta
import openpyxl
from django.template.loader import render_to_string
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table as ExcelTable, TableStyleInfo
from django.db import models
from decimal import Decimal
from django.contrib.admin.models import LogEntry
from django.utils.html import format_html
from django.contrib.admin import SimpleListFilter
from django.contrib.auth.models import User, Group
from django.contrib.auth.admin import UserAdmin, GroupAdmin
import os
from django.urls import path
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
import subprocess
from django.template.response import TemplateResponse
from django.http import HttpResponseForbidden
from django.utils.html import format_html
import io
#from .models import (
    #PurchaseRetrn,
    #PurchaseRetrnItem, 
   # SaleRetrn,
 #   SaleRetrnItem,  )



# Import at top
from .models import SystemSetting

# ============================================
# SHAREHOLDER ADMIN - FIXED
# ============================================

class ShareInline(admin.TabularInline):
    model = Share
    fields = ('share_type', 'quantity', 'purchase_price', 'certificate_number', 'issue_date')
    extra = 0
    readonly_fields = ('issue_date',)


class DividendPaymentInline(admin.TabularInline):
    model = DividendPayment
    fields = ('dividend', 'shares_held', 'amount', 'status', 'payment_date')
    readonly_fields = ('dividend', 'shares_held', 'amount')
    extra = 0


class MeetingAttendanceInline(admin.TabularInline):
    """Inline for managing meeting attendance"""
    model = MeetingAttendance
    fields = ('shareholder', 'status', 'check_in_time', 'check_out_time', 'notes')
    extra = 1
    autocomplete_fields = ['shareholder']
    
    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "shareholder":
            kwargs["queryset"] = Shareholder.objects.filter(status='active')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


@admin.register(Shareholder, site=admin.site)
class ShareholderAdmin(admin.ModelAdmin):
    list_display = ('shareholder_code', 'name', 'shareholder_type', 'phone', 'email', 'total_shares_display', 'total_investment_display', 'status_badge')
    list_filter = ('shareholder_type', 'status', 'is_founder', 'is_board_member')
    search_fields = ('shareholder_code', 'name', 'email', 'phone', 'cnic')
    inlines = [ShareInline, DividendPaymentInline]
    readonly_fields = ('shareholder_code', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Personal Information', {
            'fields': ('shareholder_code', 'name', 'shareholder_type', 'status')
        }),
        ('Contact Details', {
            'fields': ('email', 'phone', 'address')
        }),
        ('Identification', {
            'fields': ('cnic', 'passport_no')
        }),
        ('Company Details', {
            'fields': ('company_name', 'registration_no'),
            'classes': ('collapse',)
        }),
        ('Banking Details', {
            'fields': ('bank_name', 'account_number', 'account_title', 'iban'),
            'classes': ('collapse',)
        }),
        ('Additional Info', {
            'fields': ('is_founder', 'is_board_member', 'notes')
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def total_shares_display(self, obj):
        return f"{obj.total_shares():,}"
    total_shares_display.short_description = "Total Shares"
    
    def total_investment_display(self, obj):
        return f"Rs. {obj.total_investment():,.2f}"
    total_investment_display.short_description = "Total Investment"
    
    def status_badge(self, obj):
        return obj.status_badge
    status_badge.short_description = "Status"
    status_badge.allow_tags = True
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    actions = ['generate_dividend_report', 'export_shareholders_excel']
    
    def generate_dividend_report(self, request, queryset):
        """Generate dividend report for selected shareholders"""
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        story.append(Paragraph("Shareholder Dividend Report", styles['Heading1']))
        story.append(Paragraph(f"Generated: {now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.2*inch))
        
        data = [['Code', 'Name', 'Total Shares', 'Total Investment', 'Total Dividends']]
        for shareholder in queryset:
            data.append([
                shareholder.shareholder_code,
                shareholder.name,
                f"{shareholder.total_shares():,}",
                f"Rs. {shareholder.total_investment():,.2f}",
                f"Rs. {shareholder.total_dividends():,.2f}"
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="shareholder_dividend_report.pdf"'
        return response
    generate_dividend_report.short_description = "📄 Generate Dividend Report (PDF)"
    
    def export_shareholders_excel(self, request, queryset):
        """Export shareholders to Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shareholders"
        
        headers = ['Code', 'Name', 'Type', 'Email', 'Phone', 'Status', 'Total Shares', 'Total Investment']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        for shareholder in queryset:
            ws.append([
                shareholder.shareholder_code,
                shareholder.name,
                shareholder.get_shareholder_type_display(),
                shareholder.email or '',
                shareholder.phone or '',
                shareholder.status,
                shareholder.total_shares(),
                float(shareholder.total_investment())
            ])
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="shareholders_export.xlsx"'
        wb.save(response)
        return response
    export_shareholders_excel.short_description = "📊 Export to Excel"


@admin.register(Share, site=admin.site)
class ShareAdmin(admin.ModelAdmin):
    list_display = ('id', 'shareholder', 'share_type', 'quantity', 'purchase_price', 'total_value_display', 'issue_date')
    list_filter = ('share_type', 'is_locked', 'issue_date')
    search_fields = ('shareholder__name', 'shareholder__shareholder_code', 'certificate_number')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('Share Information', {
            'fields': ('shareholder', 'share_type', 'quantity', 'purchase_price', 'certificate_number')
        }),
        ('Issue Details', {
            'fields': ('issue_date', 'is_locked')
        }),
        ('Transfer Tracking', {
            'fields': ('transferred_from', 'transfer_date', 'transfer_notes'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def total_value_display(self, obj):
        return f"Rs. {obj.total_value():,.2f}"
    total_value_display.short_description = "Total Value"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(SharePrice, site=admin.site)
class SharePriceAdmin(admin.ModelAdmin):
    list_display = ('price', 'date', 'is_active', 'created_at')
    list_filter = ('is_active', 'date')
    search_fields = ('notes',)
    list_editable = ('is_active',)
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


class DividendPaymentInline(admin.TabularInline):
    model = DividendPayment
    fields = ('shareholder', 'shares_held', 'amount', 'status', 'payment_date')
    readonly_fields = ('shareholder', 'shares_held', 'amount')
    extra = 0
    can_delete = False


@admin.register(Dividend, site=admin.site)
class DividendAdmin(admin.ModelAdmin):
    list_display = ('dividend_no', 'amount_per_share', 'total_amount', 'declaration_date', 'record_date', 'payment_date', 'status')
    list_filter = ('status', 'is_interim', 'declaration_date')
    search_fields = ('dividend_no', 'notes')
    inlines = [DividendPaymentInline]
    readonly_fields = ('dividend_no', 'total_amount', 'created_at')
    actions = ['generate_dividend_payments', 'mark_paid_bulk']
    
    fieldsets = (
        ('Dividend Information', {
            'fields': ('dividend_no', 'amount_per_share', 'total_amount', 'is_interim')
        }),
        ('Dates', {
            'fields': ('declaration_date', 'record_date', 'payment_date')
        }),
        ('Status', {
            'fields': ('status',)
        }),
        ('Additional', {
            'fields': ('notes',)
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def generate_dividend_payments(self, request, queryset):
        """Generate dividend payments for all shareholders"""
        count = 0
        for dividend in queryset.filter(status='declared'):
            shareholders = Shareholder.objects.filter(status='active')
            for shareholder in shareholders:
                shares = shareholder.total_shares()
                if shares > 0:
                    amount = shares * dividend.amount_per_share
                    DividendPayment.objects.get_or_create(
                        dividend=dividend,
                        shareholder=shareholder,
                        defaults={
                            'shares_held': shares,
                            'amount': amount,
                            'status': 'pending'
                        }
                    )
                    count += 1
            dividend.status = 'approved'
            dividend.save()
        
        self.message_user(request, f"✅ {count} dividend payments generated for {queryset.count()} dividends!")
    generate_dividend_payments.short_description = "💰 Generate Dividend Payments"
    
    def mark_paid_bulk(self, request, queryset):
        """Mark selected dividend payments as paid"""
        count = 0
        for dividend in queryset:
            payments = dividend.payments.filter(status='pending')
            for payment in payments:
                payment.mark_paid(payment_method='Bank Transfer', processed_by=request.user)
                count += 1
        self.message_user(request, f"✅ {count} payments marked as paid!")
    mark_paid_bulk.short_description = "✅ Mark Payments as Paid"


@admin.register(ShareTransfer, site=admin.site)
class ShareTransferAdmin(admin.ModelAdmin):
    list_display = ('transfer_no', 'from_shareholder', 'to_shareholder', 'quantity', 'transfer_price', 'total_value_display', 'status_badge', 'transfer_date')
    list_filter = ('status', 'transfer_date')
    search_fields = ('transfer_no', 'from_shareholder__name', 'to_shareholder__name')
    readonly_fields = ('transfer_no', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Transfer Information', {
            'fields': ('transfer_no', 'from_shareholder', 'to_shareholder')
        }),
        ('Share Details', {
            'fields': ('quantity', 'transfer_price')
        }),
        ('Shares Selection', {
            'fields': ('shares',),
            'description': 'Select the shares to transfer'
        }),
        ('Dates', {
            'fields': ('transfer_date',)
        }),
        ('Status', {
            'fields': ('status', 'approved_by', 'approved_at', 'completed_at')
        }),
        ('Notes', {
            'fields': ('reason', 'notes')
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    # ✅ Use filter_horizontal with ManyToManyField
    filter_horizontal = ('shares',)
    
    def total_value_display(self, obj):
        return f"Rs. {obj.total_value():,.2f}"
    total_value_display.short_description = "Total Value"
    
    def status_badge(self, obj):
        return obj.status_badge
    status_badge.short_description = "Status"
    status_badge.allow_tags = True
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    actions = ['approve_transfers', 'complete_transfers', 'reject_transfers']
    
    def approve_transfers(self, request, queryset):
        count = 0
        for transfer in queryset.filter(status='pending'):
            transfer.approve(request.user)
            count += 1
        self.message_user(request, f"✅ {count} transfers approved!")
    approve_transfers.short_description = "✅ Approve Selected Transfers"
    
    def complete_transfers(self, request, queryset):
        count = 0
        for transfer in queryset.filter(status='approved'):
            transfer.complete(request.user)
            count += 1
        self.message_user(request, f"✅ {count} transfers completed!")
    complete_transfers.short_description = "✅ Complete Selected Transfers"
    
    def reject_transfers(self, request, queryset):
        count = 0
        for transfer in queryset.filter(status='pending'):
            transfer.reject(request.user)
            count += 1
        self.message_user(request, f"❌ {count} transfers rejected!")
    reject_transfers.short_description = "❌ Reject Selected Transfers"


@admin.register(ShareholderMeeting, site=admin.site)
class ShareholderMeetingAdmin(admin.ModelAdmin):
    """Shareholder Meeting Admin - Fixed for custom through model"""
    list_display = ('meeting_no', 'title', 'meeting_type', 'date', 'status', 'total_attendees')
    list_filter = ('meeting_type', 'status', 'date')
    search_fields = ('meeting_no', 'title', 'agenda')
    readonly_fields = ('meeting_no', 'created_at', 'updated_at', 'total_attendees')
    
    fieldsets = (
        ('Meeting Information', {
            'fields': ('meeting_no', 'title', 'meeting_type', 'date', 'time', 'venue', 'status')
        }),
        ('Agenda & Minutes', {
            'fields': ('agenda', 'minutes')
        }),
        # ✅ attendees removed from fieldsets - using inline instead
        ('Resolutions', {
            'fields': ('resolutions', 'resolutions_passed'),
            'classes': ('collapse',)
        }),
        ('Additional', {
            'fields': ('notes',)
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    # ✅ Use inline for attendance management
    inlines = [MeetingAttendanceInline]
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(MeetingAttendance, site=admin.site)
class MeetingAttendanceAdmin(admin.ModelAdmin):
    list_display = ('meeting', 'shareholder', 'status', 'check_in_time', 'check_out_time')
    list_filter = ('status', 'meeting')
    search_fields = ('shareholder__name', 'shareholder__shareholder_code')
    readonly_fields = ('created_at',)
    
    def has_add_permission(self, request):
        return False  # Attendance created via meeting

@admin.register(CashTransaction)
class CashTransactionAdmin(admin.ModelAdmin):
    list_display = ('date', 'transaction_type', 'amount', 'description', 'created_by')
    list_filter = ('transaction_type', 'date')
    search_fields = ('description', 'reference_no')
    readonly_fields = ('created_at',)

# Register at bottom with other admin registrations
@admin.register(SystemSetting)
class SystemSettingAdmin(admin.ModelAdmin):
    list_display = ('setting_key', 'setting_value', 'description', 'updated_at')
    list_editable = ('setting_value',)
    list_filter = ('setting_key',)
    search_fields = ('setting_key', 'description')
    
    fieldsets = (
        ('Module Visibility Settings', {
            'fields': ('setting_key', 'setting_value', 'description'),
            'description': '📌 Set "true" to show module, "false" to hide module'
        }),
    )
    
    def has_add_permission(self, request):
        # Only superuser can add new settings
        return request.user.is_superuser
    
    def has_delete_permission(self, request, obj=None):
        # Only superuser can delete
        return request.user.is_superuser

# Create default settings after migration
def create_default_system_settings():
    defaults = [
        ('show_hr_module', 'true', 'Show/Hide HR Module (Employees, Attendance, Payroll, Leaves, Salary Slips)'),
        ('show_production_module', 'true', 'Show/Hide Production Module (Production Orders, BOM, Operations)'),
        ('show_installment_module', 'true', 'Show/Hide Installment Module (Installment Plans, EMI Payments)'),
        ('show_reports_module', 'true', 'Show/Hide Reports Module (All Reports)'),
        ('show_whatsapp_module', 'true', 'Show/Hide WhatsApp Module (Messages, Reminders)'),
        ('show_inventory_module', 'true', 'Show/Hide Inventory Module (Stock, Batches, Transfers)'),
        ('show_purchase_module', 'true', 'Show/Hide Purchase Module (Purchases, Purchase Orders, GRN)'),
        ('show_sales_module', 'true', 'Show/Hide Sales Module (Sales, Sale Orders, Challans)'),
        ('show_accounts_module', 'true', 'Show/Hide Accounts Module (Expenses, Savings, Debts)'),
        ('show_backup_module', 'true', 'Show/Hide Database Backup Module'),
    ]
    
    for key, value, desc in defaults:
        SystemSetting.objects.get_or_create(
            setting_key=key,
            defaults={'setting_value': value, 'description': desc}
        )

# Call this after migration (will run on server start)
try:
    if SystemSetting.objects.count() == 0:
        create_default_system_settings()
except:
    pass

BACKUP_DIR = "C:/Users/Store/P1/dbbackup/"

def generate_invoice_pdf(sale_obj):
    """
    Premium Professional Invoice PDF with Amount in Words
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch, mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io
    from datetime import datetime
    import os
    
    buffer = io.BytesIO()
    
    # Page Setup
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        topMargin=0.4*inch,
        bottomMargin=0.4*inch,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch
    )
    
    styles = getSampleStyleSheet()
    story = []
    
    # Company Info
    company = CompanyInfo.objects.first()
    company_name = company.name if company else "MY SUPER STORE"
    company_address = company.address if company else "Khanewal"
    company_phone = company.contact_number if company else "0300 0000000"
    company_email = company.email if company else "info@mysuperstore.com"
    company_tagline = company.tagline if company and company.tagline else "Quality Products, Best Service"
    
    # ============================================
    # CUSTOM STYLES
    # ============================================
    
    title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Heading1'],
        fontSize=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a1a2e'),
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#6c757d'),
        spaceAfter=3
    )
    
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#6c757d'),
        fontName='Helvetica'
    )
    
    value_style = ParagraphStyle(
        'Value',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#1a1a2e'),
        fontName='Helvetica-Bold'
    )
    
    # ============================================
    # HEADER SECTION
    # ============================================
    
    story.append(Paragraph(company_name, title_style))
    story.append(Paragraph(company_tagline, subtitle_style))
    story.append(Paragraph(f"{company_address} | Phone: {company_phone}", subtitle_style))
    story.append(Spacer(1, 0.1*inch))
    
    # Decorative Line
    line_table = Table([['']], colWidths=[6.5*inch])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#667eea')),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(line_table)
    story.append(Spacer(1, 0.1*inch))
    
    # ============================================
    # INVOICE TITLE
    # ============================================
    
    invoice_title_style = ParagraphStyle(
        'InvoiceTitle2',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    story.append(Paragraph("INVOICE", invoice_title_style))
    story.append(Spacer(1, 0.1*inch))
    
    # ============================================
    # INVOICE DETAILS (Two Column)
    # ============================================
    
    # Left Column
    left_data = [
        [Paragraph("<b>Invoice No:</b>", label_style), Paragraph(f"<b>{sale_obj.bill_no}</b>", value_style)],
        [Paragraph("<b>Invoice Date:</b>", label_style), Paragraph(sale_obj.sale_date.strftime('%d %B, %Y'), value_style)],
    ]
    
    # Right Column
    right_data = [
        [Paragraph("<b>Customer Name:</b>", label_style), Paragraph(sale_obj.customer.name, value_style)],
        [Paragraph("<b>Customer Code:</b>", label_style), Paragraph(sale_obj.customer.customer_code or 'N/A', value_style)],
    ]
    
    if sale_obj.customer.address:
        left_data.append([Paragraph("<b>Address:</b>", label_style), Paragraph(sale_obj.customer.address[:50], label_style)])
    
    left_table = Table(left_data, colWidths=[1.2*inch, 2.5*inch])
    right_table = Table(right_data, colWidths=[1.2*inch, 2.5*inch])
    
    left_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    right_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    info_table = Table([[left_table, right_table]], colWidths=[4*inch, 3.5*inch])
    info_table.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.15*inch))
    
    # ============================================
    # ITEMS TABLE
    # ============================================
    
    table_data = [['#', 'Product', 'Qty', 'Unit', 'Unit Price', 'Total']]
    
    total_before_discount = Decimal('0.0')
    
    for idx, item in enumerate(sale_obj.saleitem_set.all(), 1):
        item_total = item.total_amt
        table_data.append([
            str(idx),
            item.product.name,
            f"{item.qty:,.2f}",
            item.product.unit.name if item.product.unit else '-',
            f"Rs. {item.price:,.2f}",
            f"Rs. {item_total:,.2f}"
        ])
        total_before_discount += item_total
    
    discount_value = sale_obj.discount_value
    total_amount = sale_obj.total_amount()
    paid_amount = sale_obj.paid
    outstanding = sale_obj.outstanding_balance()
    
    table_data.append(['', '', '', '', 'Subtotal:', f"Rs. {total_before_discount:,.2f}"])
    if discount_value > 0:
        table_data.append(['', '', '', '', 'Discount:', f"Rs. {discount_value:,.2f}"])
    table_data.append(['', '', '', '', 'Grand Total:', f"Rs. {total_amount:,.2f}"])
    
    # Payment Methods
    payments = sale_obj.payments.all()
    if payments.exists():
        table_data.append(['', '', '', '', '', ''])
        table_data.append(['', '', '', '', 'Payment Details:', ''])
        for payment in payments:
            method_name = payment.method.get_name_display() if payment.method else str(payment.method)
            ref_text = f" (Ref: {payment.reference_no})" if payment.reference_no else ""
            table_data.append(['', '', '', '', f'{method_name}{ref_text}:', f"Rs. {payment.amount:,.2f}"])
        table_data.append(['', '', '', '', 'Total Paid:', f"Rs. {paid_amount:,.2f}"])
        if outstanding > 0:
            table_data.append(['', '', '', '', 'Outstanding:', f"Rs. {outstanding:,.2f}"])
        elif outstanding < 0:
            change_amount = abs(outstanding)
            table_data.append(['', '', '', '', 'Change Return:', f"Rs. {change_amount:,.2f}"])
    else:
        table_data.append(['', '', '', '', 'Paid:', f"Rs. {paid_amount:,.2f}"])
        table_data.append(['', '', '', '', 'Outstanding:', f"Rs. {outstanding:,.2f}"])
    
    col_widths = [0.4*inch, 2.2*inch, 0.6*inch, 0.6*inch, 0.9*inch, 1.2*inch]
    
    items_table = Table(table_data, repeatRows=1, colWidths=col_widths)
    
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        ('FONTSIZE', (0, 1), (-1, -5), 8),
        ('VALIGN', (0, 1), (-1, -5), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -5), [colors.white, colors.HexColor('#f8f9fc')]),
        
        ('GRID', (0, 0), (-1, -5), 0.5, colors.HexColor('#e0e3eb')),
        
        ('ALIGN', (2, 1), (5, -5), 'RIGHT'),
        ('ALIGN', (0, 1), (0, -5), 'CENTER'),
        
        ('LINEABOVE', (4, -5), (5, -5), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (4, -4), (5, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (4, -1), (5, -1), colors.HexColor('#e8ecf4')),
    ]
    
    items_table.setStyle(TableStyle(table_style))
    story.append(items_table)
    story.append(Spacer(1, 0.2*inch))
    
    # ============================================
    # ✅ AMOUNT IN WORDS (ADDED)
    # ============================================
    
    def number_to_words(amount):
        """Convert number to words"""
        try:
            from num2words import num2words
            return num2words(amount, lang='en').title()
        except ImportError:
            # Simple conversion if num2words not installed
            return f"{amount:,.2f}"
        except:
            return f"{amount:,.2f}"
    
    amount_in_words = number_to_words(total_amount)
    
    # Amount in Words Box
    words_style = ParagraphStyle(
        'WordsStyle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#28a745'),
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#e8f5e9'),
        borderPadding=5,
        borderWidth=1,
        borderColor=colors.HexColor('#28a745'),
        borderRadius=5
    )
    
    story.append(Paragraph(f"<b>Amount in Words:</b> {amount_in_words} Only.", words_style))
    story.append(Spacer(1, 0.2*inch))
    
    # ============================================
    # FOOTER SECTION
    # ============================================
    
    footer_line = Table([['']], colWidths=[6.5*inch])
    footer_line.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#667eea')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(footer_line)
    
    thank_style = ParagraphStyle(
        'ThankYou',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#667eea'),
        fontName='Helvetica-Bold',
        spaceAfter=5
    )
    story.append(Paragraph("Thank you for your business!", thank_style))
    
    footer_note = company.footer_note if company and company.footer_note else "This is a computer generated invoice. No signature required."
    story.append(Paragraph(footer_note, subtitle_style))
    story.append(Spacer(1, 0.1*inch))
    
    story.append(Paragraph("<b>Terms & Conditions:</b>", label_style))
    story.append(Paragraph("1. Goods once sold will not be taken back.", label_style))
    story.append(Paragraph("2. All disputes subject to local jurisdiction.", label_style))
    story.append(Paragraph("3. Please retain this invoice for warranty claims.", label_style))
    story.append(Spacer(1, 0.1*inch))
    
    generated_date = datetime.now().strftime('%d %B, %Y at %I:%M %p')
    story.append(Paragraph(f"<font size='7' color='#adb5bd'>Generated on: {generated_date}</font>", subtitle_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

class EmiPaymentInline(admin.TabularInline):
    model = EmiPayment
    fields = ('installment_number', 'due_date', 'amount_due', 'amount_paid', 'status', 'payment_date', 'reference_no')
    readonly_fields = ('installment_number', 'due_date', 'amount_due')
    extra = 0
    can_delete = False
    
    def get_readonly_fields(self, request, obj=None):
        if obj and obj.status == 'paid':
            return self.readonly_fields + ('amount_paid', 'status')
        return self.readonly_fields


@admin.register(InstallmentPlan, site=admin.site)
class InstallmentPlanAdmin(admin.ModelAdmin):
    list_display = ('name', 'duration_months', 'down_payment_percent', 'interest_rate', 'late_fee_per_day', 'is_active', 'created_at')
    list_editable = ('is_active',)
    list_filter = ('is_active', 'duration_months')
    search_fields = ('name',)
    ordering = ('duration_months',)
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'duration_months', 'is_active')
        }),
        ('Payment Terms', {
            'fields': ('down_payment_percent', 'interest_rate', 'late_fee_per_day'),
            'description': 'Down payment % of total amount | Interest rate % per annum | Late fee Rs. per day'
        }),
    )


@admin.register(SaleInstallment, site=admin.site)
class SaleInstallmentAdmin(admin.ModelAdmin):
    list_display = ('sale_bill_no', 'customer_name', 'plan', 'total_amount_display', 
                    'down_payment_status', 'total_paid_display', 'remaining_display', 
                    'status', 'next_due_date', 'payment_progress_bar')
    list_filter = ('status', 'plan', 'down_payment_paid', 'start_date')
    search_fields = ('sale__bill_no', 'sale__customer__name', 'sale__customer__customer_code')
    inlines = [EmiPaymentInline]
    readonly_fields = ('total_amount', 'loan_amount', 'emi_amount', 'total_interest', 
                       'total_payable', 'late_fee_accrued', 'created_at', 'updated_at')
    actions = ['generate_emi_schedule', 'send_payment_reminders', 'mark_as_defaulted', 'export_installment_report']
    change_list_template = "admin/button.html"
    
    fieldsets = (
        ('Sale Information', {
            'fields': ('sale', 'plan', 'status')
        }),
        ('Amount Breakdown', {
            'fields': ('total_amount', 'down_payment_amount', 'down_payment_paid', 
                      'loan_amount', 'emi_amount', 'total_interest', 'total_payable')
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date', 'next_due_date')
        }),
        ('Late Fees', {
            'fields': ('late_fee_accrued', 'late_fee_paid'),
            'classes': ('collapse',)
        }),
        ('Tracking', {
            'fields': ('created_at', 'updated_at', 'created_by'),
            'classes': ('collapse',)
        }),
    )
    
    def sale_bill_no(self, obj):
        url = reverse('admin:app_sale_change', args=[obj.sale.id])
        return format_html('<a href="{}">{}</a>', url, obj.sale.bill_no)
    sale_bill_no.short_description = "Bill No"
    
    def customer_name(self, obj):
        return obj.sale.customer.name
    customer_name.short_description = "Customer"
    
    def total_amount_display(self, obj):
        return f"Rs. {obj.total_amount:,.2f}"
    total_amount_display.short_description = "Total Amount"
    
    def down_payment_status(self, obj):
        if obj.down_payment_paid:
            return format_html('<span style="color: #28a745;">✅ Paid</span>')
        return format_html('<span style="color: #dc3545;">❌ Pending</span>')
    down_payment_status.short_description = "Down Payment"
    
    def total_paid_display(self, obj):
        return f"Rs. {obj.total_paid():,.2f}"
    total_paid_display.short_description = "Total Paid"
    
    def remaining_display(self, obj):
        remaining = obj.remaining_amount()
        if remaining <= 0:
            return format_html('<span style="color: #28a745;">Rs. 0.00</span>')
        elif remaining > 0:
            return format_html('<span style="color: #dc3545;">Rs. {:.2f}</span>', remaining)
        return f"Rs. {remaining:,.2f}"
    remaining_display.short_description = "Remaining"
    
    def payment_progress_bar(self, obj):
        """Show payment progress bar"""
        if obj.total_payable > 0:
            percent = (obj.total_paid() / obj.total_payable) * 100
            color = '#28a745' if percent >= 100 else '#ffc107' if percent >= 50 else '#dc3545'
            return format_html(
                '<div style="width:100px; background:#e0e0e0; border-radius:10px; overflow:hidden;">'
                '<div style="width:{}%; background:{}; height:8px; border-radius:10px;"></div>'
                '</div><small>{}%</small>',
                percent, color, round(percent, 1)
            )
        return "-"
    payment_progress_bar.short_description = "Progress"
    
    # ============================================
    # ACTIONS
    # ============================================
    
    def generate_emi_schedule(self, request, queryset):
        """Generate EMI schedule for selected installments"""
        count = 0
        for installment in queryset:
            if not installment.emi_payments.exists() and installment.plan:
                for i in range(1, installment.plan.duration_months + 1):
                    due_date = installment.start_date + timedelta(days=30 * i)
                    EmiPayment.objects.create(
                        installment=installment,
                        installment_number=i,
                        due_date=due_date,
                        amount_due=installment.emi_amount
                    )
                count += 1
        self.message_user(request, f"✅ EMI schedules generated for {count} installments!")
    generate_emi_schedule.short_description = "📅 Generate EMI Schedule"
    
    def send_payment_reminders(self, request, queryset):
        """Send WhatsApp payment reminders"""
        from .whatsapp_utils import WhatsAppSender
        
        count = 0
        for installment in queryset:
            pending_emi = installment.next_emi_due()
            if pending_emi and pending_emi.status == 'pending' and installment.sale.customer.contact_number:
                message = f"""📱 *EMI REMINDER*

Customer: {installment.sale.customer.name}
Bill No: {installment.sale.bill_no}
EMI #{pending_emi.installment_number} Due: Rs. {pending_emi.amount_due:,.2f}
Due Date: {pending_emi.due_date}
Remaining Balance: Rs. {installment.remaining_amount():,.2f}

⚠️ Late fee of Rs. {installment.plan.late_fee_per_day}/day will apply after due date.

Please clear your payment on time!"""
                
                result = WhatsAppSender.send_direct_message(
                    installment.sale.customer.contact_number,
                    message
                )
                if result.get('success'):
                    count += 1
        
        self.message_user(request, f"✅ {count} reminders sent!")
    send_payment_reminders.short_description = "📱 Send Payment Reminders"
    
    def mark_as_defaulted(self, request, queryset):
        """Mark selected installments as defaulted"""
        count = queryset.update(status='defaulted')
        self.message_user(request, f"⚠️ {count} installments marked as defaulted!")
    mark_as_defaulted.short_description = "⚠️ Mark as Defaulted"
    
    def export_installment_report(self, request, queryset):
        """Export installment report as Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Installment Report"
        
        # Headers
        headers = ['Bill No', 'Customer', 'Plan', 'Total Amount', 'Down Payment', 
                   'EMI Amount', 'Total Paid', 'Remaining', 'Status', 'Start Date', 'Next Due']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for inst in queryset:
            ws.append([
                inst.sale.bill_no,
                inst.sale.customer.name,
                inst.plan.name if inst.plan else '-',
                float(inst.total_amount),
                float(inst.down_payment_amount),
                float(inst.emi_amount),
                float(inst.total_paid()),
                float(inst.remaining_amount()),
                inst.get_status_display(),
                inst.start_date.strftime('%d-%m-%Y'),
                inst.next_due_date.strftime('%d-%m-%Y') if inst.next_due_date else '-',
            ])
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="installment_report.xlsx"'
        wb.save(response)
        return response
    export_installment_report.short_description = "📊 Export to Excel"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(EmiPayment, site=admin.site)
class EmiPaymentAdmin(admin.ModelAdmin):
    list_display = ('installment_sale_bill', 'installment_number', 'due_date', 'amount_due_display', 
                    'amount_paid_display', 'status', 'payment_date', 'is_overdue_display')
    list_filter = ('status', 'payment_date', 'due_date')
    search_fields = ('installment__sale__bill_no', 'installment__sale__customer__name')
    readonly_fields = ('installment', 'installment_number', 'due_date', 'amount_due')
    actions = ['mark_as_paid']
    
    fieldsets = (
        ('EMI Information', {
            'fields': ('installment', 'installment_number', 'due_date', 'amount_due')
        }),
        ('Payment Details', {
            'fields': ('amount_paid', 'payment_date', 'payment_method', 'reference_no', 'status', 'notes')
        }),
        ('Late Fee', {
            'fields': ('late_fee_charged',),
            'classes': ('collapse',)
        }),
    )
    
    def installment_sale_bill(self, obj):
        return obj.installment.sale.bill_no
    installment_sale_bill.short_description = "Bill No"
    
    def amount_due_display(self, obj):
        return f"Rs. {obj.amount_due:,.2f}"
    amount_due_display.short_description = "Amount Due"
    
    def amount_paid_display(self, obj):
        if obj.amount_paid >= obj.amount_due:
            return format_html('<span style="color: #28a745;">Rs. {:.2f}</span>', obj.amount_paid)
        elif obj.amount_paid > 0:
            return format_html('<span style="color: #ff9800;">Rs. {:.2f}</span>', obj.amount_paid)
        return f"Rs. {obj.amount_paid:,.2f}"
    amount_paid_display.short_description = "Amount Paid"
    
    def is_overdue_display(self, obj):
        if obj.is_overdue():
            days = (now().date() - obj.due_date).days
            return format_html('<span style="color: #dc3545;">⚠️ {} days overdue</span>', days)
        return format_html('<span style="color: #28a745;">✅ On time</span>')
    is_overdue_display.short_description = "Overdue"
    
    def mark_as_paid(self, request, queryset):
        """Mark selected EMIs as paid"""
        count = 0
        for emi in queryset.filter(status='pending'):
            emi.mark_paid(emi.amount_due)
            count += 1
        self.message_user(request, f"✅ {count} EMIs marked as paid!")
    mark_as_paid.short_description = "✅ Mark as Paid"
    
    def has_add_permission(self, request):
        return False  # EMIs are auto-generated from installment

@admin.register(DailyClosing, site=admin.site)
class DailyClosingAdmin(admin.ModelAdmin):
    list_display = ('closing_date', 'opening_cash', 'cash_sales', 'total_sales', 'closing_cash', 'cash_difference', 'is_closed')
    list_filter = ('is_closed', 'closing_date')
    readonly_fields = ('cash_sales', 'card_sales', 'jazzcash_sales', 'easypaisa_sales', 'bank_transfer_sales', 'cash_difference', 'total_sales_today')
    
    def total_sales(self, obj):
        return f"Rs. {obj.total_sales_today():,.2f}"
    total_sales.short_description = 'Total Sales'


@admin.register(DailyClosingExpense, site=admin.site)
class DailyClosingExpenseAdmin(admin.ModelAdmin):
    list_display = ('closing', 'description', 'amount', 'category')

@admin.register(PaymentMethod, site=admin.site)
class PaymentMethodAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_active')
    list_editable = ('is_active',)


@admin.register(SalePayment, site=admin.site)
class SalePaymentAdmin(admin.ModelAdmin):
    list_display = ('sale', 'method', 'amount', 'payment_date', 'created_by')
    list_filter = ('method', 'payment_date')
    readonly_fields = ('payment_date',)

class CustomAdminSite(admin.AdminSite):
    """Custom Admin Site for Adding Backup/Restore/Delete View"""

    index_template = "admin/custom_index.html"

    def get_default_backup_dir(self):
        """Get default backup directory based on platform"""
        import platform
        
        if platform.system() == 'Linux' and os.path.exists('/storage/emulated/0/'):
            return '/storage/emulated/0/Download/Backups/'
        elif platform.system() == 'Windows':
            return os.path.join(os.path.expanduser('~'), 'Documents', 'Backups')
        else:
            return os.path.join(os.path.expanduser('~'), 'Backups')

    def get_active_backup_dir(self, request):
        """Get current backup directory (session ya default)"""
        session_path = request.session.get('backup_dir', '')
        if session_path and os.path.exists(session_path):
            return session_path
        
        default = self.get_default_backup_dir()
        os.makedirs(default, exist_ok=True)
        return default

    # ============================================
    # URLS
    # ============================================
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            # Database Backup/Restore
            path('database-backup/', self.admin_view(self.database_backup_view), name='database-backup'),
            path('backup-db/', self.admin_view(self.backup_database), name='backup-db'),
            path('restore-db/', self.admin_view(self.restore_database), name='restore-db'),
            path('delete-backup/<str:filename>/', self.admin_view(self.delete_backup), name='delete-backup'),
            path('set-backup-dir/', self.admin_view(self.set_backup_dir), name='set-backup-dir'),
            
            # Invoice & Documents
            path('invoice/<int:sale_id>/', self.admin_view(self.invoice_view), name='generate_invoice'),
            path('challan-pdf/<int:challan_id>/', self.admin_view(self.challan_pdf_view), name='generate_challan_pdf'),
            path('grn-pdf/<int:grn_id>/', self.admin_view(self.grn_pdf_view), name='generate_grn_pdf'),
            path('audit-pdf/<int:audit_id>/', self.admin_view(self.audit_pdf_view), name='generate_audit_pdf'),
            
            # Barcode & Search
            path('find-product-by-barcode/', self.admin_view(self.find_product_by_barcode_view), name='find_product_by_barcode'),
            path('find-product-by-serial/', self.admin_view(self.find_product_by_serial_view), name='find_product_by_serial'),
            path('get-product-price/', self.admin_view(self.get_product_price_view), name='get_product_price'),
            path('get-batch-selling-price/', self.admin_view(self.get_batch_selling_price_view), name='get_batch_selling_price'),
            path('get-next-bill-no/', self.admin_view(self.get_next_bill_no), name='get-next-bill-no'),
            path('app/product/supplier-barcode/', self.admin_view(self.supplier_barcode_view), name='supplier_barcode'),
            path('set-supplier-barcode/', self.admin_view(self.set_supplier_barcode), name='set_supplier_barcode'),
            
            # Status Updates
            path('app/saleorder/update-status/', self.admin_view(self.update_order_status_view), name='update_order_status'),
            path('app/purchaseorder/update-status/', self.admin_view(self.update_po_status_view), name='update_po_status'),
            
            # Dashboard
            path('dashboard/', self.admin_view(self.dashboard_view), name='dashboard'),
            path('dashboard/sales-forecast/', self.admin_view(self.sales_forecast_view), name='sales_forecast'),
            
            # System Health
            path('system-health/', self.admin_view(self.system_health_view), name='system-health'),
            
            # Range Reports
            path('reports/range/', self.admin_view(self.range_report_view), name='range_report'),
            path('reports/range/pdf/', self.admin_view(self.range_report_pdf), name='range_report_pdf'),
            
            # WhatsApp
            path('whatsapp/send/', self.admin_view(self.whatsapp_view), name='whatsapp_send'),
            path('whatsapp/reminders/', self.admin_view(self.whatsapp_reminders_view), name='whatsapp_reminders'),
            path('whatsapp/daily-summary/', self.admin_view(self.whatsapp_daily_summary_view), name='whatsapp_daily'),
        ]
        return custom_urls + urls

    # ============================================
    # BACKUP METHODS
    # ============================================
    
    def get_backup_files(self, request):
        """Get list of backup files from active backup directory"""
        try:
            backup_path = self.get_active_backup_dir(request)
            if not os.path.exists(backup_path):
                return []
            files = os.listdir(backup_path)
            files = [f for f in files if f.endswith(('.dump', '.sql', '.json', '.gz', '.sqlite'))]
            files = sorted(files, key=lambda x: os.path.getmtime(os.path.join(backup_path, x)), reverse=True)
            return files
        except Exception:
            return []

    def set_backup_dir(self, request):
        """AJAX: User se folder path lo aur session mein save karo"""
        from django.http import JsonResponse
        import json
        
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                backup_path = data.get('path', '').strip()
            except:
                backup_path = request.POST.get('path', '').strip()
            
            if backup_path:
                try:
                    os.makedirs(backup_path, exist_ok=True)
                except OSError as e:
                    return JsonResponse({'success': False, 'message': f'❌ Cannot create folder: {e}'})
                
                request.session['backup_dir'] = backup_path
                request.session.modified = True
                
                return JsonResponse({
                    'success': True, 
                    'message': f'✅ Backup folder set: {backup_path}',
                    'path': backup_path
                })
            else:
                return JsonResponse({'success': False, 'message': '❌ Please enter a valid path'})
        
        return JsonResponse({'success': False, 'message': 'Invalid request'})

    def database_backup_view(self, request):
        """Backup page view"""
        if not request.user.is_superuser:
            return HttpResponseForbidden("You are not allowed to access this page.")
        
        backup_path = self.get_active_backup_dir(request)
        backup_files = self.get_backup_files(request)
        
        import platform
        is_android = platform.system() == 'Linux' and os.path.exists('/storage/emulated/0/')
        is_windows = platform.system() == 'Windows'
        
        context = {
            **self.each_context(request),
            "title": "Database Backup & Restore",
            "backup_files": backup_files,
            "backup_dir": backup_path,
            "is_android": is_android,
            "is_windows": is_windows,
        }
        return TemplateResponse(request, "admin/database_backup.html", context)

    def backup_database(self, request):
        """Create backup"""
        backup_path = self.get_active_backup_dir(request)
        os.makedirs(backup_path, exist_ok=True)
        
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = os.path.join(backup_path, f'db_backup_{timestamp}.json')
            
            result = subprocess.run(
                ['python', 'manage.py', 'dumpdata', '--exclude', 'auth.permission', 
                 '--exclude', 'contenttypes', '--indent', '2', '--output', backup_file],
                check=True, capture_output=True, text=True
            )
            
            success_msg = f"✅ Database backup created!\n📁 {backup_file}"
            
            if 'sqlite' in settings.DATABASES['default']['ENGINE']:
                db_path = settings.DATABASES['default']['NAME']
                if os.path.exists(db_path):
                    sqlite_backup = os.path.join(backup_path, f'db_backup_{timestamp}.sqlite')
                    shutil.copy2(db_path, sqlite_backup)
                    success_msg += f"\n📁 {sqlite_backup}"
            
            messages.success(request, success_msg)
        except subprocess.CalledProcessError as e:
            messages.error(request, f"❌ Error while creating backup: {e.stderr}")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")
        
        return redirect("admin:database-backup")

    def restore_database(self, request):
        """Restore database from backup"""
        if request.method == 'POST':
            backup_file = request.POST.get('backup_file', '')
            backup_path = self.get_active_backup_dir(request)
            full_path = os.path.join(backup_path, backup_file)
            
            if '..' in backup_file or '/' in backup_file or '\\' in backup_file:
                messages.error(request, "❌ Invalid filename")
                return redirect("admin:database-backup")
            
            if not backup_file or not os.path.exists(full_path):
                messages.error(request, f"❌ Backup file not found: {backup_file}")
                return redirect("admin:database-backup")
            
            try:
                if backup_file.endswith('.json'):
                    result = subprocess.run(
                        ['python', 'manage.py', 'loaddata', full_path],
                        check=True, capture_output=True, text=True
                    )
                    messages.success(request, f"✅ Database restored from {backup_file}")
                
                elif backup_file.endswith('.sqlite'):
                    if 'sqlite' in settings.DATABASES['default']['ENGINE']:
                        db_path = settings.DATABASES['default']['NAME']
                        safety_backup = db_path + '.before_restore'
                        shutil.copy2(db_path, safety_backup)
                        shutil.copy2(full_path, db_path)
                        messages.success(request, f"✅ SQLite database restored!\n⚠ Old DB saved as .before_restore")
                    else:
                        messages.error(request, "❌ Not a SQLite database")
                
                else:
                    messages.error(request, f"❌ Unsupported format: {backup_file}")
                    
            except subprocess.CalledProcessError as e:
                messages.error(request, f"❌ Error: {e.stderr}")
            except Exception as e:
                messages.error(request, f"❌ Error: {str(e)}")
        
        return redirect("admin:database-backup")

    def delete_backup(self, request, filename):
        """Delete a backup file"""
        backup_path = self.get_active_backup_dir(request)
        
        if '..' in filename or '/' in filename or '\\' in filename:
            messages.error(request, "❌ Invalid filename")
            return redirect("admin:database-backup")
        
        file_path = os.path.join(backup_path, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            messages.success(request, f"🗑️ Backup file '{filename}' deleted!")
        else:
            messages.error(request, f"❌ File not found: '{filename}'")
        return redirect("admin:database-backup")

    # ============================================
    # GET NEXT BILL NUMBER
    # ============================================
    def get_next_bill_no(self, request):
        from django.http import JsonResponse
        
        prefix = request.GET.get('prefix', '')
        today = datetime.now().strftime('%Y%m%d')
        
        if not prefix:
            prefix = f'INV-{today}'
        
        try:
            last_sale = Sale.objects.filter(
                bill_no__startswith=prefix
            ).order_by('-bill_no').first()
            
            if last_sale:
                try:
                    last_num = int(last_sale.bill_no.split('-')[-1])
                    new_num = str(last_num + 1).zfill(4)
                except (ValueError, IndexError):
                    new_num = '0001'
            else:
                new_num = '0001'
            
            bill_no = f'{prefix}-{new_num}'
            return JsonResponse({'success': True, 'bill_no': bill_no})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    # ============================================
    # DOCUMENT VIEWS
    # ============================================
    def invoice_view(self, request, sale_id):
        sale = get_object_or_404(Sale, pk=sale_id)
        pdf_buffer = generate_invoice_pdf(sale)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Invoice_{sale.bill_no}.pdf"'
        return response

    def challan_pdf_view(self, request, challan_id):
        challan = get_object_or_404(DeliveryChallan, pk=challan_id)
        admin_obj = DeliveryChallanAdmin(DeliveryChallan, self)
        return admin_obj.generate_challan_pdf(request, DeliveryChallan.objects.filter(pk=challan_id))

    def grn_pdf_view(self, request, grn_id):
        grn = get_object_or_404(GoodsReceivedNote, pk=grn_id)
        admin_obj = GoodsReceivedNoteAdmin(GoodsReceivedNote, self)
        return admin_obj.generate_grn_pdf(request, GoodsReceivedNote.objects.filter(pk=grn_id))

    def audit_pdf_view(self, request, audit_id):
        audit = get_object_or_404(StockAudit, pk=audit_id)
        admin_obj = StockAuditAdmin(StockAudit, self)
        return admin_obj.generate_audit_pdf(request, StockAudit.objects.filter(pk=audit_id))

    # ============================================
    # BARCODE & SERIAL SEARCH
    # ============================================
    def find_product_by_barcode_view(self, request):
        from django.http import JsonResponse
        import re
        
        barcode = request.GET.get('barcode', '').strip()
        barcode = re.sub(r'[^0-9]', '', barcode)
        
        if not barcode:
            return JsonResponse({'success': False, 'message': 'No barcode provided'})
        
        try:
            product = Product.objects.get(barcode=barcode)
            return JsonResponse({
                'success': True,
                'product': {
                    'id': product.id,
                    'name': product.name,
                    'barcode': product.barcode,
                    'serial_no': product.serial_no or '',
                    'price': str(product.price)
                }
            })
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found'})

    def find_product_by_serial_view(self, request):
        """AJAX view to find product by serial number"""
        from django.http import JsonResponse
        
        serial = request.GET.get('serial', '').strip()
        
        if not serial:
            return JsonResponse({'success': False, 'message': 'No serial number provided'})
        
        try:
            product = Product.objects.get(serial_no__iexact=serial)
            return JsonResponse({
                'success': True,
                'product': {
                    'id': product.id,
                    'name': product.name,
                    'serial_no': product.serial_no,
                    'barcode': product.barcode or '',
                    'price': str(product.price)
                }
            })
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found with this serial number'})

    def get_product_price_view(self, request):
        from django.http import JsonResponse
        
        product_id = request.GET.get('product_id')
        if product_id:
            try:
                product = Product.objects.get(pk=product_id)
                return JsonResponse({
                    'success': True,
                    'price': str(product.price),
                    'name': product.name
                })
            except Product.DoesNotExist:
                pass
        return JsonResponse({'success': False})

    # ============================================
    # ✅ BATCH SELLING PRICE API (NEW!)
    # ============================================
    def get_batch_selling_price_view(self, request):
        """Get batch selling price for a product"""
        from django.http import JsonResponse
        
        product_id = request.GET.get('product_id')
        warehouse_id = request.GET.get('warehouse_id')
        
        if not product_id:
            return JsonResponse({'success': False, 'message': 'Product ID required'})
        
        try:
            product = Product.objects.get(pk=product_id)
            
            # ✅ Latest batch dhundo jisme selling price set hai
            batch = None
            if warehouse_id:
                batch = StockBatch.objects.filter(
                    product=product,
                    warehouse_id=warehouse_id,
                    remaining_qty__gt=0,
                    selling_price__gt=0
                ).order_by('id').first()
            else:
                batch = StockBatch.objects.filter(
                    product=product,
                    remaining_qty__gt=0,
                    selling_price__gt=0
                ).order_by('id').first()
            
            default_price = str(product.price)
            
            if batch and batch.selling_price > 0:
                return JsonResponse({
                    'success': True,
                    'price': str(batch.selling_price),
                    'purchase_price': str(batch.price),
                    'batch_selling_price': str(batch.selling_price),
                    'name': product.name,
                    'has_batch_price': True
                })
            else:
                return JsonResponse({
                    'success': True,
                    'price': default_price,
                    'name': product.name,
                    'has_batch_price': False
                })
                
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found'})

    def supplier_barcode_view(self, request):
        ids_str = request.GET.get('ids', '')
        product_ids = [int(id) for id in ids_str.split(',') if id]
        products = Product.objects.filter(id__in=product_ids)
        return render(request, "admin/supplier_barcode.html", {
            **self.each_context(request),
            "title": "Set Supplier Barcode",
            "products": products
        })

    def set_supplier_barcode(self, request):
        from django.http import JsonResponse
        import re
        
        product_id = request.POST.get('product_id')
        barcode = request.POST.get('barcode', '').strip()
        
        if not product_id or not barcode:
            return JsonResponse({'success': False, 'message': 'Missing data'})
        
        barcode = re.sub(r'[^0-9]', '', barcode)
        if not barcode:
            return JsonResponse({'success': False, 'message': 'Invalid barcode'})
        
        try:
            product = Product.objects.get(pk=product_id)
            if Product.objects.filter(barcode=barcode).exclude(pk=product_id).exists():
                return JsonResponse({'success': False, 'message': '❌ Already used by another product'})
            
            product.barcode = barcode
            product.use_custom_barcode = True
            product.save()
            return JsonResponse({
                'success': True,
                'message': f'✅ Barcode set for {product.name}',
                'barcode': barcode
            })
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product not found'})

    # ============================================
    # STATUS UPDATE VIEWS
    # ============================================
    def update_order_status_view(self, request):
        ids_str = request.GET.get('ids', '')
        order_ids = [int(id) for id in ids_str.split(',') if id]
        orders = SaleOrder.objects.filter(id__in=order_ids)
        
        if request.method == 'POST':
            new_status = request.POST.get('status')
            if new_status:
                orders.update(status=new_status)
                messages.success(request, f"✅ Updated {orders.count()} orders")
                return redirect("admin:app_saleorder_changelist")
        
        return render(request, "admin/update_order_status.html", {
            **self.each_context(request),
            "title": "Update Order Status",
            "orders": orders,
            "status_choices": SaleOrder.ORDER_STATUS
        })

    def update_po_status_view(self, request):
        ids_str = request.GET.get('ids', '')
        po_ids = [int(id) for id in ids_str.split(',') if id]
        orders = PurchaseOrder.objects.filter(id__in=po_ids)
        
        if request.method == 'POST':
            new_status = request.POST.get('status')
            if new_status:
                orders.update(status=new_status)
                messages.success(request, f"✅ Updated {orders.count()} POs")
                return redirect("admin:app_purchaseorder_changelist")
        
        return render(request, "admin/update_po_status.html", {
            **self.each_context(request),
            "title": "Update PO Status",
            "orders": orders,
            "status_choices": PurchaseOrder.ORDER_STATUS
        })

    # ============================================
    # DASHBOARD
    # ============================================
    def dashboard_view(self, request):
        from datetime import timedelta, date
        import json
        
        today = date.today()
        month_ago = today - timedelta(days=30)
        
        daily_sales = []
        for i in range(30):
            d = today - timedelta(days=29-i)
            total = Sale.objects.filter(sale_date__date=d).aggregate(
                t=Sum('saleitem__total_amt')
            )['t'] or 0
            daily_sales.append({'date': d.strftime('%d %b'), 'amount': float(total)})
        
        top_products = SaleItem.objects.filter(
            sale__sale_date__gte=month_ago
        ).values('product__name').annotate(
            total_qty=Sum('qty'), total_sales=Sum('total_amt')
        ).order_by('-total_sales')[:10]
        
        slow_moving = []
        for inv in Inventory.objects.filter(stock__gt=0):
            sales_30d = SaleItem.objects.filter(
                product=inv.product, sale__sale_date__gte=month_ago
            ).aggregate(t=Sum('qty'))['t'] or 0
            if sales_30d == 0 or (inv.stock > 0 and (sales_30d / inv.stock) < 0.1):
                slow_moving.append({
                    'name': inv.product.name,
                    'stock': inv.stock,
                    'sales_30d': sales_30d,
                    'warehouse': inv.warehouse.name
                })
        
        context = {
            **self.each_context(request),
            "title": "Dashboard",
            "daily_sales": json.dumps(daily_sales),
            "top_products": list(top_products),
            "slow_moving": slow_moving[:10],
            "total_customers": Customer.objects.count(),
            "total_products": Product.objects.count(),
            "total_sales_today": Sale.objects.filter(sale_date__date=today).count(),
            "total_purchase_today": Purchase.objects.filter(pur_date__date=today).count(),
            "low_stock_count": Inventory.objects.filter(stock__lt=F('product__low_stock_threshold')).count(),
        }
        return TemplateResponse(request, "admin/dashboard.html", context)

    def sales_forecast_view(self, request):
        from datetime import date, timedelta
        
        today = date.today()
        daily_data = []
        for i in range(90):
            d = today - timedelta(days=89-i)
            total = Sale.objects.filter(sale_date__date=d).aggregate(
                t=Sum('saleitem__total_amt')
            )['t'] or 0
            daily_data.append(float(total))
        
        forecast = []
        if len(daily_data) >= 7:
            for i in range(7):
                avg = sum(daily_data[-7:]) / 7
                next_date = today + timedelta(days=i+1)
                forecast.append({'date': next_date.strftime('%d %b'), 'amount': round(avg, 2)})
                daily_data.append(avg)
        
        return TemplateResponse(request, "admin/sales_forecast.html", {
            **self.each_context(request),
            "title": "Sales Forecast",
            "forecast": forecast
        })

    # ============================================
    # SYSTEM HEALTH
    # ============================================
    def system_health_view(self, request):
        cpu_percent = 0
        cpu_count = 0
        try:
            with open('/proc/cpuinfo', 'r') as f:
                cpu_count = sum(1 for line in f if line.startswith('processor'))
            with open('/proc/stat', 'r') as f:
                cpu_line = f.readline()
                cpu_values = [int(x) for x in cpu_line.split()[1:8]]
                idle = cpu_values[3]
                total = sum(cpu_values)
                if total > 0:
                    cpu_percent = round(100 - (idle / total * 100), 1)
        except:
            pass
        
        ram_total_gb = 0
        ram_used_gb = 0
        ram_percent = 0
        try:
            mem_info = {}
            with open('/proc/meminfo', 'r') as f:
                for line in f:
                    parts = line.split(':')
                    if len(parts) == 2:
                        mem_info[parts[0].strip()] = int(parts[1].strip().split()[0])
            total_kb = mem_info.get('MemTotal', 0)
            free_kb = mem_info.get('MemFree', 0)
            buffers_kb = mem_info.get('Buffers', 0)
            cached_kb = mem_info.get('Cached', 0)
            used_kb = total_kb - free_kb - buffers_kb - cached_kb
            ram_total_gb = round(total_kb / (1024**2), 2)
            ram_used_gb = round(used_kb / (1024**2), 2)
            if total_kb > 0:
                ram_percent = round((used_kb / total_kb) * 100, 1)
        except:
            pass
        
        disk_total_gb = 0
        disk_used_gb = 0
        disk_percent = 0
        try:
            stat = os.statvfs('/storage/emulated/0/')
            total_bytes = stat.f_frsize * stat.f_blocks
            free_bytes = stat.f_frsize * stat.f_bfree
            used_bytes = total_bytes - free_bytes
            disk_total_gb = round(total_bytes / (1024**3), 1)
            disk_used_gb = round(used_bytes / (1024**3), 1)
            if total_bytes > 0:
                disk_percent = round((used_bytes / total_bytes) * 100, 1)
        except:
            pass
        
        db_size_mb = 0
        try:
            if 'sqlite' in settings.DATABASES['default']['ENGINE']:
                db_path = settings.DATABASES['default']['NAME']
                if os.path.exists(db_path):
                    db_size_mb = round(os.path.getsize(db_path) / (1024**2), 2)
        except:
            pass
        
        model_counts = {
            'Products': Product.objects.count(),
            'Customers': Customer.objects.count(),
            'Vendors': Vendor.objects.count(),
            'Sales': Sale.objects.count(),
            'Purchases': Purchase.objects.count(),
            'Sale Orders': SaleOrder.objects.count(),
            'Purchase Orders': PurchaseOrder.objects.count(),
            'Challans': DeliveryChallan.objects.count(),
            'GRNs': GoodsReceivedNote.objects.count(),
            'Inventory Items': Inventory.objects.count(),
            'Users': User.objects.count(),
        }
        total_records = sum(model_counts.values())
        
        yesterday = now() - timedelta(hours=24)
        recent_sales = Sale.objects.filter(sale_date__gte=yesterday).count()
        recent_purchases = Purchase.objects.filter(pur_date__gte=yesterday).count()
        recent_orders = SaleOrder.objects.filter(order_date__gte=yesterday).count()
        recent_pos = PurchaseOrder.objects.filter(order_date__gte=yesterday).count()
        recent_challans = DeliveryChallan.objects.filter(challan_date__gte=yesterday).count()
        recent_grns = GoodsReceivedNote.objects.filter(grn_date__gte=yesterday).count()
        recent_logs = LogEntry.objects.select_related('user').order_by('-action_time')[:10]
        
        last_backup = None
        try:
            backup_files = self.get_backup_files(request)
            if backup_files:
                last_backup = backup_files[0]
        except:
            pass
        
        license_valid = License.objects.filter(
            expiry_date__gte=now().date(), is_active=True
        ).exists()
        
        import django, sys
        context = {
            **self.each_context(request),
            "title": "System Health",
            "cpu_percent": cpu_percent,
            "cpu_count": cpu_count,
            "ram_total_gb": ram_total_gb,
            "ram_used_gb": ram_used_gb,
            "ram_percent": ram_percent,
            "disk_total_gb": disk_total_gb,
            "disk_used_gb": disk_used_gb,
            "disk_percent": disk_percent,
            "db_size_mb": db_size_mb,
            "total_records": total_records,
            "model_counts": model_counts,
            "recent_sales": recent_sales,
            "recent_purchases": recent_purchases,
            "recent_orders": recent_orders,
            "recent_pos": recent_pos,
            "recent_challans": recent_challans,
            "recent_grns": recent_grns,
            "recent_logs": recent_logs,
            "last_backup": last_backup,
            "license_valid": license_valid,
            "django_version": django.get_version(),
            "python_version": sys.version.split()[0],
            "uptime_str": "N/A",
        }
        return TemplateResponse(request, "admin/system_health.html", context)

    # ============================================
    # RANGE REPORTING VIEWS
    # ============================================
    def range_report_view(self, request):
        """Range report page with date filters"""
        from datetime import date, timedelta
        
        to_date = request.GET.get('to_date', date.today().strftime('%Y-%m-%d'))
        from_date = request.GET.get('from_date', (date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
        report_type = request.GET.get('report_type', 'sales')
        
        from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        context = {
            **self.each_context(request),
            "title": "📊 Range Reports",
            "from_date": from_date,
            "to_date": to_date,
            "report_type": report_type,
        }
        
        if report_type == 'sales':
            data = ReportManager.get_sales_by_date_range(from_date_obj, to_date_obj)
            context['sales_data'] = data
            context['top_products'] = ReportManager.get_top_products_by_date_range(from_date_obj, to_date_obj)
        elif report_type == 'purchases':
            data = ReportManager.get_purchases_by_date_range(from_date_obj, to_date_obj)
            context['purchase_data'] = data
        elif report_type == 'daily':
            context['daily_data'] = ReportManager.get_daily_summary(from_date_obj, to_date_obj)
        
        return TemplateResponse(request, "admin/range_report.html", context)

    def range_report_pdf(self, request):
        """Generate PDF for range report"""
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        report_type = request.GET.get('report_type', 'sales')
        
        from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "YOUR COMPANY"
        story.append(Paragraph(company_name, styles['Heading1']))
        story.append(Paragraph(f"Report: {from_date} to {to_date}", styles['Heading2']))
        
        if report_type == 'sales':
            data = ReportManager.get_sales_by_date_range(from_date_obj, to_date_obj)
            story.append(Paragraph(f"Total Sales: {data['count']}"))
            story.append(Paragraph(f"Total Amount: Rs. {data['total_amount']:,.2f}"))
            story.append(Paragraph(f"Total Profit: Rs. {data['total_profit']:,.2f}"))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    # ============================================
    # WHATSAPP VIEWS
    # ============================================
    def whatsapp_view(self, request):
        customers = Customer.objects.exclude(contact_number__isnull=True).exclude(contact_number='')
        return TemplateResponse(request, "admin/whatsapp_send.html", {
            **self.each_context(request), "title": "📱 WhatsApp Messages", "customers": customers,
        })

    def whatsapp_reminders_view(self, request):
        from .whatsapp_utils import WhatsAppBulkSender
        if request.method == 'POST':
            customer_id = request.POST.get('customer_id')
            if customer_id:
                from .whatsapp_utils import WhatsAppSender
                customer = Customer.objects.get(id=customer_id)
                WhatsAppSender.send_direct_message(customer.contact_number, request.POST.get('message', 'Payment reminder'))
                messages.success(request, "✅ Link generated!")
        return TemplateResponse(request, "admin/whatsapp_reminders.html", {
            **self.each_context(request), "title": "💰 Payment Reminders",
            "customers": WhatsAppBulkSender.generate_reminder_links(),
        })

    def whatsapp_daily_summary_view(self, request):
        from .whatsapp_utils import WhatsAppSender
        if request.method == 'POST':
            phone = request.POST.get('phone', '')
            if phone:
                result = WhatsAppSender.send_daily_summary(phone)
                if result['success']:
                    messages.success(request, "✅ Daily summary link generated!")
                    return TemplateResponse(request, "admin/whatsapp_daily_summary.html", {
                        **self.each_context(request), "title": "📊 Daily Summary", "result": result,
                    })
        return TemplateResponse(request, "admin/whatsapp_daily_summary.html", {
            **self.each_context(request), "title": "📊 Daily Summary on WhatsApp",
        })

# Default Admin Site کو CustomAdminSite سے Replace کریں
admin.site = CustomAdminSite()

admin.site.site_header = "Welcome"
admin.site.site_title = "Web App"
admin.site.index_title = "Admin Dashboard"


# SALE ORDER ITEM INLINE
# ============================================
class SaleOrderItemInline(admin.TabularInline):
    model = SaleOrderItem
    fields = ('product', 'qty', 'price', 'total_amt', 'reserved')
    readonly_fields = ('total_amt',)
    extra = 10


# ============================================
# SALE ORDER ADMIN
# ============================================
# ============================================
# SALE ORDER ADMIN
# ============================================
@admin.register(SaleOrder, site=admin.site)
class SaleOrderAdmin(admin.ModelAdmin):
    list_display = ('order_no', 'order_date', 'customer','customer_po_number','warehouse', 'status', 
                    'formatted_total', 'formatted_advance', 'formatted_outstanding', 
                    'delivery_status_display', 'converted_to_sale')
    list_filter = ('status', 'order_date', 'warehouse', 'customer__group')
    search_fields = ('order_no', 'customer__name', 'customer__contact_number')
    search_help_text = "Search by Order No, Customer Name, or Contact Number"
    inlines = [SaleOrderItemInline]
    change_list_template = "admin/button.html"
    actions = [
        'convert_to_sale', 
        'create_delivery_challan', 
        'update_status', 
        'generate_order_pdf', 
        'generate_order_html',
        'send_order_status_whatsapp',
    ]
    
    readonly_fields = ('order_no', 'created_at', 'updated_at', 'converted_to_sale', 'created_by')
    
    fieldsets = (
        ('Order Information', {
            'fields': ('order_no', 'customer','customer_po_number', 'warehouse', 'order_date', 'delivery_date', 'status')
        }),
        ('Payment Information', {
            'fields': ('discount_value', 'advance_payment')
        }),
        ('Additional Information', {
            'fields': ('notes', 'created_at', 'updated_at', 'converted_to_sale')
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        readonly = list(self.readonly_fields)
        if obj and obj.status == 'invoiced':
            readonly.extend(['customer', 'warehouse', 'order_date', 'discount_value', 'advance_payment'])
        return readonly
    
    def formatted_total(self, obj):
        return f"Rs. {obj.total_amount():,.2f}"
    formatted_total.short_description = "Total"
    
    def formatted_advance(self, obj):
        return f"Rs. {obj.advance_payment:,.2f}"
    formatted_advance.short_description = "Advance"
    
    def formatted_outstanding(self, obj):
        return f"Rs. {obj.outstanding_advance():,.2f}"
    formatted_outstanding.short_description = "Outstanding"
    
    def delivery_status_display(self, obj):
        """Show delivery progress"""
        if obj.status in ['pending', 'confirmed']:
            return "Not Started"
        elif obj.status == 'partially_delivered':
            total_items = obj.items.count()
            if total_items > 0:
                delivered_count = 0
                for item in obj.items.all():
                    if obj.delivered_qty(item.product) >= item.qty:
                        delivered_count += 1
                percent = round((delivered_count / total_items) * 100)
                return format_html(
                    '<span style="color: #ff9800;">🟡 {}% Delivered</span>', percent
                )
            return "Partially Delivered"
        elif obj.status == 'delivered':
            return format_html('<span style="color: #28a745;">✅ Fully Delivered</span>')
        elif obj.status == 'invoiced':
            return format_html('<span style="color: #007bff;">📋 Invoiced</span>')
        elif obj.status == 'cancelled':
            return format_html('<span style="color: #dc3545;">❌ Cancelled</span>')
        return obj.get_status_display()
    delivery_status_display.short_description = "Delivery"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    # ============================================
    # CONVERT TO SALE
    # ============================================
    def convert_to_sale(self, request, queryset):
        """Convert selected orders to Sale/Invoice"""
        from django.db import transaction
        
        success_count = 0
        error_messages = []
        
        for order in queryset.filter(status__in=['confirmed', 'ready', 'delivered', 'partially_delivered'], 
                                     converted_to_sale__isnull=True):
            try:
                with transaction.atomic():
                    sale = Sale.objects.create(
                        warehouse=order.warehouse,
                        bill_no=f"INV-{order.order_no}",
                        customer=order.customer,
                        sale_date=now(),
                        paid=order.advance_payment,
                        discount_value=order.discount_value,
                        created_by=request.user
                    )
                    
                    for item in order.items.all():
                        sale_item = SaleItem.objects.create(
                            sale=sale,
                            product=item.product,
                            qty=item.qty,
                            price=item.price
                        )
                        
                        if item.reserved:
                            try:
                                inventory = Inventory.objects.get(
                                    product=item.product, 
                                    warehouse=order.warehouse
                                )
                                inventory.reserved_stock = max(0, inventory.reserved_stock - item.qty)
                                inventory.save()
                            except Inventory.DoesNotExist:
                                pass
                    
                    order.status = 'invoiced'
                    order.converted_to_sale = sale
                    order.save()
                    
                    success_count += 1
                    
            except Exception as e:
                error_messages.append(f"Order {order.order_no}: {str(e)}")
        
        if success_count:
            self.message_user(request, f"✅ {success_count} orders converted to sale successfully.")
        if error_messages:
            self.message_user(request, f"❌ Errors: {'; '.join(error_messages)}", level=messages.ERROR)
    
    convert_to_sale.short_description = "🔄 Convert to Sale/Invoice"
    
    # ============================================
    # CREATE DELIVERY CHALLAN
    # ============================================
    def create_delivery_challan(self, request, queryset):
        """Create Delivery Challan with Partial Delivery Support"""
        from django.db import transaction
        from django.db.models import Sum
        
        count = 0
        for order in queryset.filter(status__in=['confirmed', 'processing', 'ready', 'partially_delivered']):
            try:
                with transaction.atomic():
                    challan = DeliveryChallan.objects.create(
                        order=order,
                        customer=order.customer,
                        warehouse=order.warehouse,
                        reason='sale',
                        created_by=request.user
                    )
                    
                    for item in order.items.all():
                        already_delivered = DeliveryChallanItem.objects.filter(
                            challan__order=order,
                            product=item.product
                        ).aggregate(total=Sum('qty'))['total'] or 0
                        
                        pending = item.qty - already_delivered
                        
                        if pending > 0:
                            DeliveryChallanItem.objects.create(
                                challan=challan,
                                product=item.product,
                                order_qty=item.qty,
                                qty=pending,
                                price=item.price,
                                notes=f"Order: {order.order_no}"
                            )
                    
                    order.update_delivery_status()
                    count += 1
                    
            except Exception as e:
                self.message_user(request, f"Error for {order.order_no}: {e}", level=messages.ERROR)
        
        if count:
            self.message_user(request, f"✅ {count} delivery challans created successfully.")
    
    create_delivery_challan.short_description = "📋 Create Delivery Challan"
    
    # ============================================
    # UPDATE STATUS
    # ============================================
    def update_status(self, request, queryset):
        """Bulk update order status"""
        selected_ids = queryset.values_list('id', flat=True)
        ids_str = ','.join(map(str, selected_ids))
        return redirect(f"/admin/app/saleorder/update-status/?ids={ids_str}")
    update_status.short_description = "📝 Update Status"
    
    # ============================================
    # GENERATE ORDER PDF
    # ============================================
    def generate_order_pdf(self, request, queryset):
        """Generate PDF for selected orders"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        
        for order in queryset:
            story.append(Paragraph(f"Sale Order - {order.order_no}", styles['Heading1']))
            story.append(Paragraph(f"Customer: {order.customer.name}", styles['Normal']))
            story.append(Paragraph(f"Order Date: {order.order_date.strftime('%d-%m-%Y')}", styles['Normal']))
            story.append(Paragraph(f"Status: {order.get_status_display()}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            data = [['Product', 'Order Qty', 'Delivered', 'Pending', 'Price', 'Total']]
            
            for item in order.items.all():
                delivered = order.delivered_qty(item.product)
                pending_qty = item.qty - delivered
                data.append([
                    item.product.name,
                    str(item.qty),
                    str(delivered),
                    str(pending_qty),
                    f"{item.price:,.2f}",
                    f"{item.total_amt:,.2f}"
                ])
            
            data.append(['', '', '', '', 'Total:', f"{order.total_amount():,.2f}"])
            data.append(['', '', '', '', 'Discount:', f"{order.discount_value:,.2f}"])
            data.append(['', '', '', '', 'Grand Total:', f"{order.total_after_discount():,.2f}"])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.5*inch))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sale_orders.pdf"'
        return response
    generate_order_pdf.short_description = "📄 Generate Order PDF"
    
    # ============================================
    # GENERATE ORDER HTML
    # ============================================
    def generate_order_html(self, request, queryset):
        """Generate HTML report"""
        context = {'orders': queryset, 'date': datetime.now()}
        html = render_to_string('admin/sale_order_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="sale_orders.html"'
        return response
    generate_order_html.short_description = "🌐 Generate HTML Report"
    
    # ============================================
    # WHATSAPP - SEND ORDER STATUS
    # ============================================
    def send_order_status_whatsapp(self, request, queryset):
        """Orders ka status WhatsApp pe bhejo"""
        from .whatsapp_utils import WhatsAppSender
        
        count = 0
        links = []
        for order in queryset:
            if order.customer.contact_number:
                result = WhatsAppSender.send_order_status(
                    order.customer.contact_number,
                    order.customer.name,
                    order.order_no,
                    order.status
                )
                if result.get('success'):
                    count += 1
                    links.append({
                        'order': order.order_no,
                        'customer': order.customer.name,
                        'status': order.get_status_display(),
                        'url': result['url']
                    })
        
        if count > 0:
            msg = f"✅ {count} order status links generated!<br><br>"
            for link in links:
                msg += f"📋 {link['order']} ({link['status']}) - {link['customer']}<br>"
            if links:
                msg += f"<br><a href='{links[0]['url']}' target='_blank' style='background:#25D366; color:white; padding:8px 15px; border-radius:5px; text-decoration:none;'>📱 Send First Update</a>"
            self.message_user(request, format_html(msg))
        else:
            self.message_user(request, "❌ Selected orders ke customers ke paas phone number nahi hai!", level=messages.WARNING)
    
    send_order_status_whatsapp.short_description = "📱 Send Order Status on WhatsApp"

class PurchaseRetrnItemInline(admin.TabularInline):
    model = PurchaseRetrnItem
    extra = 1


class SaleRetrnItemInline(admin.TabularInline):
    model = SaleRetrnItem
    extra = 1


class PurchaseItemInline(admin.TabularInline):
    model = PurchaseItem
    fields = ('product', 'qty', 'price', 'total')
    extra = 1
    readonly_fields = ('total',)

    def total(self, obj):
        return Decimal(obj.qty) * Decimal(obj.price) if obj.qty and obj.price else 0
    total.short_description = 'Total'


class SaleItemInline(admin.TabularInline):
    model = SaleItem
    fields = ('product', 'qty', 'price')
    extra = 1
    autocomplete_fields =['product']

# ============================================
# DELIVERY CHALLAN ITEM INLINE
# ============================================
class DeliveryChallanItemInline(admin.TabularInline):
    model = DeliveryChallanItem
    fields = ('product', 'order_qty', 'qty', 'pending_qty', 'unit', 'price', 'total_amt', 'notes')
    readonly_fields = ('pending_qty', 'total_amt')
    extra = 10


# ============================================
# DELIVERY CHALLAN ADMIN
# ============================================
@admin.register(DeliveryChallan, site=admin.site)
class DeliveryChallanAdmin(admin.ModelAdmin):
    list_display = ('challan_no', 'challan_date', 'customer', 'warehouse', 'order_link', 
                    'reason', 'total_qty_display', 'total_items_display', 'converted_to_sale', 'print_challan')
    list_filter = ('reason', 'challan_date', 'warehouse', 'converted_to_sale')
    search_fields = ('challan_no', 'customer__name', 'order__order_no')
    search_help_text = "Search by Challan No, Customer, or Order No"
    inlines = [DeliveryChallanItemInline]
    change_list_template = "admin/button.html"
    actions = ['generate_challan_pdf', 'convert_challan_to_sale']
    
    fieldsets = (
        ('Challan Information', {
            'fields': ('challan_no', 'order', 'customer', 'warehouse', 'challan_date', 'reason')
        }),
        ('Transport Details', {
            'fields': ('vehicle_no', 'transport_name', 'driver_name', 'driver_contact'),
            'classes': ('collapse',)
        }),
        ('Additional Info', {
            'fields': ('notes',)
        }),
    )
    
    readonly_fields = ('challan_no',)
    
    def order_link(self, obj):
        if obj.order:
            url = reverse('admin:app_saleorder_change', args=[obj.order.id])
            return format_html('<a href="{}">{}</a>', url, obj.order.order_no)
        return "-"
    order_link.short_description = "Order"
    
    def total_qty_display(self, obj):
        return f"{obj.total_qty():,.2f}"
    total_qty_display.short_description = "Total Qty"
    
    def total_items_display(self, obj):
        return obj.total_items()
    total_items_display.short_description = "Items"
    
    def print_challan(self, obj):
        url = reverse('admin:generate_challan_pdf', args=[obj.pk])
        return format_html('<a href="{}" target="_blank">📄 Print</a>', url)
    print_challan.short_description = "Print"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        if obj.order and not obj.customer_id:
            obj.customer = obj.order.customer
        if obj.order and not obj.warehouse_id:
            obj.warehouse = obj.order.warehouse
        super().save_model(request, obj, form, change)
    
    def generate_challan_pdf(self, request, queryset):
        """Generate PDF for selected challans with Order Qty + Deliver Qty + Pending"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.3*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        company_address = company.address if company else "Address"
        company_phone = company.contact_number if company else "Phone"
        
        for challan in queryset:
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=5)
            story.append(Paragraph(company_name, title_style))
            story.append(Paragraph(company_address, styles['Normal']))
            story.append(Paragraph(f"Phone: {company_phone}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            story.append(Paragraph("<b>DELIVERY CHALLAN</b>", title_style))
            story.append(Spacer(1, 0.15*inch))
            
            details_data = [
                [Paragraph(f"<b>Challan No:</b> {challan.challan_no}", styles['Normal']),
                 Paragraph(f"<b>Date:</b> {challan.challan_date.strftime('%d-%m-%Y')}", styles['Normal'])],
                [Paragraph(f"<b>Order No:</b> {challan.order.order_no if challan.order else 'N/A'}", styles['Normal']),
                 Paragraph(f"<b>Vehicle No:</b> {challan.vehicle_no or '_________'}", styles['Normal'])],
                [Paragraph(f"<b>Reason:</b> {challan.get_reason_display()}", styles['Normal']),
                 Paragraph(f"<b>Transport:</b> {challan.transport_name or '_________'}", styles['Normal'])],
            ]
            
            details_table = Table(details_data, colWidths=[3.5*inch, 3.5*inch])
            details_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(details_table)
            
            from_to_data = [
                [Paragraph("<b>From:</b>", styles['Normal']), Paragraph("<b>To:</b>", styles['Normal'])],
                [Paragraph(company_name + "<br/>" + company_address + "<br/>Phone: " + company_phone, styles['Normal']),
                 Paragraph(f"{challan.customer.name}<br/>{challan.customer.address or ''}<br/>Contact: {challan.customer.contact_number or ''}", styles['Normal'])],
            ]
            
            from_to_table = Table(from_to_data, colWidths=[3.5*inch, 3.5*inch])
            from_to_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LINEBELOW', (0, 1), (-1, 1), 0.5, colors.grey),
            ]))
            story.append(from_to_table)
            story.append(Spacer(1, 0.2*inch))
            
            # ✅ Updated Items Table with Order Qty + Deliver Qty + Pending
            table_data = [['S.No', 'Product', 'Order Qty', 'Deliver Qty', 'Pending', 'Unit', 'Remarks']]
            
            total_order = 0
            total_deliver = 0
            total_pending = 0
            
            for idx, item in enumerate(challan.items.all(), 1):
                table_data.append([
                    str(idx),
                    item.product.name,
                    f"{item.order_qty:,.2f}",
                    f"{item.qty:,.2f}",
                    f"{item.pending_qty:,.2f}",
                    item.unit or '-',
                    item.notes or ''
                ])
                total_order += item.order_qty
                total_deliver += item.qty
                total_pending += item.pending_qty
            
            table_data.append(['', 'TOTAL', f"{total_order:,.2f}", f"{total_deliver:,.2f}", f"{total_pending:,.2f}", '', f"{challan.total_items()} items"])
            
            col_widths = [0.4*inch, 1.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch, 1.2*inch]
            items_table = Table(table_data, repeatRows=1, colWidths=col_widths)
            
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -2), 'LEFT'),
                ('ALIGN', (2, 1), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ]))
            
            story.append(items_table)
            story.append(Spacer(1, 0.4*inch))
            
            sign_data = [
                [Paragraph("<b>Received by:</b> _________________", styles['Normal']),
                 Paragraph("<b>Dispatched by:</b> _________________", styles['Normal'])],
                [Paragraph("Signature: _________________", styles['Normal']),
                 Paragraph("Signature: _________________", styles['Normal'])],
                [Paragraph("Date: _________________", styles['Normal']),
                 Paragraph("Date: _________________", styles['Normal'])],
            ]
            
            sign_table = Table(sign_data, colWidths=[3.5*inch, 3.5*inch])
            sign_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(sign_table)
            
            story.append(Spacer(1, 0.3*inch))
            story.append(Paragraph("This is a computer generated challan - no signature required.", styles['Normal']))
            story.append(Paragraph("<br/><br/>", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="delivery_challan.pdf"'
        return response
    generate_challan_pdf.short_description = "📄 Generate Challan PDF"
    
    def convert_challan_to_sale(self, request, queryset):
        """Convert challan to Sale/Invoice"""
        from django.db import transaction
        
        count = 0
        for challan in queryset.filter(converted_to_sale=False):
            try:
                with transaction.atomic():
                    sale = Sale.objects.create(
                        warehouse=challan.warehouse,
                        bill_no=f"INV-{challan.challan_no}",
                        customer=challan.customer,
                        sale_date=now(),
                        created_by=request.user
                    )
                    
                    for item in challan.items.all():
                        SaleItem.objects.create(
                            sale=sale,
                            product=item.product,
                            qty=item.qty,
                            price=item.price
                        )
                    
                    challan.converted_to_sale = True
                    challan.save()
                    count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        
        if count:
            self.message_user(request, f"✅ {count} challans converted to sale.")
    convert_challan_to_sale.short_description = "🔄 Convert to Sale/Invoice"

@admin.register(Warehouse, site=admin.site)
class WarehouseAdmin(admin.ModelAdmin):
    list_display = ('name', 'location', 'description')
    search_fields = ('name',)
    list_filter = ('location',)
    change_list_template = "admin/button.html"


@admin.register(WarehouseTransfer, site=admin.site)
class WarehouseTransferAdmin(admin.ModelAdmin):
    list_display = ('transfer_date', 'product', 'from_warehouse', 'to_warehouse', 'qty', 'created_by')
    list_filter = ('from_warehouse', 'to_warehouse', 'transfer_date')
    search_fields = ('product__name',)
    change_list_template = "admin/button.html"
    actions = ['generate_transfer_report', 'generate_transfer_pdf']

    def generate_transfer_report(self, request, queryset):
        """Generate HTML report for selected transfers."""
        context = {
            'transfers': queryset,
            'date': datetime.now()
        }
        html = render_to_string('admin/warehouse_transfer_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="warehouse_transfer_report.html"'
        return response
    generate_transfer_report.short_description = "Generate Transfer Report (HTML)"
    
    def generate_transfer_pdf(self, request, queryset):
        """Generate PDF report for warehouse transfers."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Warehouse Transfer Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Product', 'From Warehouse', 'To Warehouse', 'Quantity', 'Transferred By']]
        
        for transfer in queryset:
            table_data.append([
                transfer.transfer_date.strftime('%d-%m-%Y'),
                transfer.product.name,
                transfer.from_warehouse.name,
                transfer.to_warehouse.name,
                f"{transfer.qty:,.2f}",
                transfer.created_by.username if transfer.created_by else '-'
            ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="warehouse_transfer_report.pdf"'
        return response
    generate_transfer_pdf.short_description = "Generate Transfer Report (PDF)"


from django.urls import reverse
from django.utils.html import format_html
from .barcode_utils import generate_barcode_image, generate_barcode_label, generate_multiple_labels
from django.http import HttpResponse
import io

@admin.register(Product, site=admin.site)

class ProductAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    search_fields = ('serial_no', 'name', 'barcode')
    search_help_text = "Search by Serial No, Product Name, Barcode"
    list_display = ('serial_no', 'name', 'barcode', 'barcode_preview', 'use_custom_barcode', 
                    'description', 'unit', 'location', 'price', 'used', 'brand', 'category', 
                    'types', 'low_stock_threshold')
    ordering = ('serial_no',)
    change_list_template = "admin/button.html"
    actions = ['generate_barcodes', 'print_barcode_labels', 'print_selected_labels', 'print_single_labels', 'scan_supplier_barcode']
    
    # Fieldsets - price hataya gaya kyunke editable=False hai
    fieldsets = (
        ('Basic Information', {
            'fields': ('serial_no', 'name', 'description', 'used')
        }),
        ('Barcode Information', {
            'fields': ('use_custom_barcode', 'barcode'),
            'description': 'Check "Use Custom/Supplier Barcode" to manually enter supplier barcode instead of auto-generating.'
        }),
        ('Categorization', {
            'fields': ('unit', 'brand', 'category', 'types', 'location')
        }),
        ('Stock Settings', {
            'fields': ('low_stock_threshold',),
            'description': 'Set minimum stock level for alerts.'
        }),
    )
    
    # Price ko readonly_fields mein add karein
    readonly_fields = ('price',)
    
    def get_readonly_fields(self, request, obj=None):
        """Make barcode readonly if use_custom_barcode is False"""
        readonly = list(self.readonly_fields)
        if obj and not obj.use_custom_barcode:
            readonly.append('barcode')
        return readonly
    
    def changelist_view(self, request, extra_context=None):
        """Add has_barcode_action flag to context"""
        extra_context = extra_context or {}
        extra_context['has_barcode_action'] = True
        return super().changelist_view(request, extra_context=extra_context)
    
    def barcode_preview(self, obj):
        """Show barcode image in admin list"""
        if obj.barcode:
            try:
                barcode_img = generate_barcode_image(obj.barcode)
                if barcode_img:
                    return format_html('<img src="{}" width="150" height="50">', barcode_img)
                return obj.barcode
            except Exception as e:
                return obj.barcode
        return "No Barcode"
    barcode_preview.short_description = "Barcode Preview"
    
    def generate_barcodes(self, request, queryset):
        """Generate barcodes for selected products that don't have one AND don't use custom barcode"""
        count = 0
        for product in queryset.filter(barcode__isnull=True, use_custom_barcode=False):
            product.save()
            count += 1
        self.message_user(request, f"Generated barcodes for {count} products.")
    generate_barcodes.short_description = "🔄 Generate Barcodes for Selected"
    
    def scan_supplier_barcode(self, request, queryset):
        """Action to open barcode scanner for supplier barcode"""
        selected_ids = queryset.values_list('id', flat=True)
        ids_str = ','.join(map(str, selected_ids))
        return redirect(f"/admin/app/product/supplier-barcode/?ids={ids_str}")
    scan_supplier_barcode.short_description = "🏷️ Scan/Set Supplier Barcode"
    
    def print_barcode_labels(self, request, queryset):
        """Generate PDF with barcode labels for selected products"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.3*inch, bottomMargin=0.3*inch)
        styles = getSampleStyleSheet()
        story = []
        
        story.append(Paragraph("Barcode Labels", styles['Heading1']))
        story.append(Spacer(1, 0.2*inch))
        
        table_data = []
        row = []
        
        for i, product in enumerate(queryset):
            if not product.barcode:
                product.save()
            
            barcode_img_data = generate_barcode_image(product.barcode)
            
            if not barcode_img_data:
                barcode_img_data = ""
            
            cell_content = Paragraph(
                f"<b>{product.name[:20]}</b><br/>"
                f"<img src='{barcode_img_data}' width='120' height='40'/><br/>"
                f"{product.barcode}<br/>"
                f"Rs. {product.price:,.2f}",
                styles['Normal']
            )
            
            row.append(cell_content)
            
            if len(row) == 3 or i == queryset.count() - 1:
                while len(row) < 3:
                    row.append("")
                table_data.append(row)
                row = []
        
        if table_data:
            table = Table(table_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
            table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ]))
            
            story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="barcode_labels.pdf"'
        return response
    print_barcode_labels.short_description = "📄 Print Barcode Labels (PDF)"
    
    def print_selected_labels(self, request, queryset):
        """Print labels as PNG image sheet"""
        products_data = []
        for product in queryset:
            if not product.barcode:
                product.save()
            products_data.append({
                'barcode': product.barcode,
                'name': product.name,
                'price': str(product.price)
            })
        
        buffer = generate_multiple_labels(products_data)
        
        response = HttpResponse(buffer, content_type='image/png')
        response['Content-Disposition'] = 'attachment; filename="barcode_labels_sheet.png"'
        return response
    print_selected_labels.short_description = "🖼️ Print Labels Sheet (PNG)"
    
    def print_single_labels(self, request, queryset):
        """Print individual labels one per page"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        for product in queryset:
            if not product.barcode:
                product.save()
            
            barcode_img = generate_barcode_image(product.barcode)
            
            story.append(Paragraph(f"<b>{product.name}</b>", styles['Heading2']))
            if barcode_img:
                story.append(Paragraph(f"<img src='{barcode_img}' width='200' height='60'/>", styles['Normal']))
            story.append(Paragraph(f"Barcode: {product.barcode}", styles['Normal']))
            story.append(Paragraph(f"Price: Rs. {product.price:,.2f}", styles['Normal']))
            story.append(Spacer(1, 0.5*inch))
            story.append(Paragraph("- - - - - - - - - - - - - - - - - - - -", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="single_barcode_labels.pdf"'
        return response
    print_single_labels.short_description = "🏷️ Print Single Labels (One per page)"


# Product Alias Admin
@admin.register(ProductAlias, site=admin.site)
class ProductAliasAdmin(admin.ModelAdmin):
    list_display = ('alias', 'product', 'created_by', 'created_at')
    list_filter = ('created_by',)
    search_fields = ('alias', 'product__name')
    autocomplete_fields = ['product', 'created_by']
    readonly_fields = ('created_at',)
    
    fieldsets = (
        ('Alias Information', {
            'fields': ('product', 'alias', 'created_by')
        }),
    )
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

# ============================================
# VENDOR GROUP ADMIN
# ============================================
@admin.register(VendorGroup, site=admin.site)
class VendorGroupAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    search_help_text = "Search by Vendor Group Name"
    list_display = ['name', 'description', 'total_outstanding_balance_display']
    change_list_template = "admin/button.html"
    actions = ['generate_html_report', 'generate_vendor_group_pdf']

    def total_outstanding_balance_display(self, obj):
        return f'{obj.total_outstanding_balance():,.2f}'
    total_outstanding_balance_display.short_description = 'Total Outstanding Balance'

    def generate_html_report(self, request, queryset):
        VendorGroups = queryset
        context = {'VendorGroups': VendorGroups, 'date': datetime.now()}
        html = render_to_string('admin/VendorGroup_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="VendorGroup_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_vendor_group_pdf(self, request, queryset):
        """Generate PDF report for Vendor Groups."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Vendor Group Report", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Group Name', 'Description', 'Outstanding Balance']]
        
        for group in queryset:
            table_data.append([
                group.name,
                group.description or '-',
                f"{group.total_outstanding_balance():,.2f}"
            ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="vendor_group_report.pdf"'
        return response
    generate_vendor_group_pdf.short_description = 'Generate PDF Report'


# ============================================
# VENDOR ADMIN
# ============================================
@admin.register(Vendor, site=admin.site)
class VendorAdmin(admin.ModelAdmin):
    search_fields = ('vendor_code', 'name')
    search_help_text = "Search by Vendor Code or Vendor Name"
    list_display = ('vendor_code', 'name', 'group', 'contact_number', 'address', 
                    'opening_balance', 'formatted_outstanding_balance')
    ordering = ('vendor_code',)
    change_list_template = "admin/button.html"
    actions = ['generate_html_report', 'generate_vendor_pdf']
    
    fieldsets = (
        ('Vendor Information', {
            'fields': ('vendor_code', 'name', 'group', 'address')
        }),
        ('Contact Details', {
            'fields': ('contact_number',)
        }),
        ('Financial Information', {
            'fields': ('opening_balance',),
            'description': '⚠️ Opening Balance: Previous outstanding amount before using this system (before software start)'
        }),
    )

    def formatted_outstanding_balance(self, obj):
        return f"{obj.outstanding_balance():,.2f} PKR"
    formatted_outstanding_balance.short_description = "Total Outstanding"

    def generate_html_report(self, request, queryset):
        vendors = queryset
        context = {'vendors': vendors, 'date': datetime.now()}
        html = render_to_string('admin/vendor_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="vendor_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_vendor_pdf(self, request, queryset):
        """Generate PDF report for Vendors."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, 
                                      alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Vendor Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Updated table header with Vendor Code + Opening Balance
        table_data = [['Code', 'Name', 'Group', 'Contact', 'Opening Balance', 'Outstanding']]
        
        total_opening = Decimal('0.0')
        total_outstanding = Decimal('0.0')
        
        for vendor in queryset:
            opening = vendor.opening_balance
            balance = vendor.outstanding_balance()
            table_data.append([
                vendor.vendor_code or '-',
                vendor.name,
                vendor.group.name if vendor.group else '-',
                vendor.contact_number or '-',
                f"{opening:,.2f}",
                f"{balance:,.2f}"
            ])
            total_opening += opening
            total_outstanding += balance
        
        table_data.append(['', '', '', '', 'TOTAL:', f"{total_outstanding:,.2f}"])
        
        table = Table(table_data, repeatRows=1, 
                      colWidths=[0.8*inch, 1.3*inch, 1*inch, 1*inch, 1*inch, 1.1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="vendor_report.pdf"'
        return response
    generate_vendor_pdf.short_description = 'Generate PDF Report'
    
# ============================================
# PURCHASE ORDER ITEM INLINE
# ============================================
class PurchaseOrderItemInline(admin.TabularInline):
    model = PurchaseOrderItem
    fields = ('product', 'qty', 'price', 'total_amt')
    readonly_fields = ('total_amt',)
    extra = 10


# ============================================
# PURCHASE ORDER ADMIN
# ============================================
@admin.register(PurchaseOrder, site=admin.site)
class PurchaseOrderAdmin(admin.ModelAdmin):
    list_display = ('order_no', 'order_date', 'vendor','vendor_so_number', 'warehouse', 'status',
                    'formatted_total', 'formatted_advance', 'formatted_outstanding',
                    'receive_status_display', 'converted_to_purchase')
    list_filter = ('status', 'order_date', 'warehouse', 'vendor__group')
    search_fields = ('order_no', 'vendor__name')
    search_help_text = "Search by PO Number or Vendor Name"
    inlines = [PurchaseOrderItemInline]
    change_list_template = "admin/button.html"
    actions = ['convert_to_purchase', 'create_grn', 'update_po_status', 'generate_po_pdf']
    
    readonly_fields = ('order_no', 'created_at', 'updated_at', 'converted_to_purchase', 'created_by')
    
    fieldsets = (
        ('Order Information', {
            'fields': ('order_no', 'vendor','vendor_so_number', 'warehouse', 'order_date', 'expected_date', 'status')
        }),
        ('Payment Information', {
            'fields': ('discount_value', 'advance_payment')
        }),
        ('Additional Information', {
            'fields': ('notes', 'created_at', 'updated_at', 'converted_to_purchase')
        }),
    )
    
    def formatted_total(self, obj):
        return f"Rs. {obj.total_amount():,.2f}"
    formatted_total.short_description = "Total"
    
    def formatted_advance(self, obj):
        return f"Rs. {obj.advance_payment:,.2f}"
    formatted_advance.short_description = "Advance"
    
    def formatted_outstanding(self, obj):
        return f"Rs. {obj.outstanding_advance():,.2f}"
    formatted_outstanding.short_description = "Outstanding"
    
    def receive_status_display(self, obj):
        if obj.status in ['pending', 'confirmed']:
            return "Not Received"
        elif obj.status == 'partially_received':
            total_items = obj.items.count()
            if total_items > 0:
                received_count = sum(1 for item in obj.items.all() if obj.received_qty(item.product) >= item.qty)
                percent = round((received_count / total_items) * 100)
                return format_html('<span style="color: #ff9800;">🟡 {}% Received</span>', percent)
            return "Partially Received"
        elif obj.status == 'received':
            return format_html('<span style="color: #28a745;">✅ Received</span>')
        elif obj.status == 'completed':
            return format_html('<span style="color: #007bff;">📋 Completed</span>')
        elif obj.status == 'cancelled':
            return format_html('<span style="color: #dc3545;">❌ Cancelled</span>')
        return obj.get_status_display()
    receive_status_display.short_description = "Receive Status"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def convert_to_purchase(self, request, queryset):
        """Convert selected POs to Purchase"""
        success_count = 0
        for order in queryset.filter(status__in=['received', 'partially_received'], converted_to_purchase__isnull=True):
            try:
                with transaction.atomic():
                    purchase = Purchase.objects.create(
                        warehouse=order.warehouse,
                        bill_no=f"BILL-{order.order_no}",
                        vendor=order.vendor,
                        pur_date=now(),
                        paid=order.advance_payment,
                        created_by=request.user
                    )
                    
                    for item in order.items.all():
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=item.product,
                            qty=item.qty,
                            price=item.price
                        )
                    
                    order.status = 'completed'
                    order.converted_to_purchase = purchase
                    order.save()
                    success_count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        
        if success_count:
            self.message_user(request, f"✅ {success_count} POs converted to purchase.")
    convert_to_purchase.short_description = "🔄 Convert to Purchase"
    
    def create_grn(self, request, queryset):
        """Create GRN from selected POs"""
        from django.db.models import Sum
        
        count = 0
        for order in queryset.filter(status__in=['confirmed', 'processing', 'shipped', 'partially_received']):
            try:
                with transaction.atomic():
                    grn = GoodsReceivedNote.objects.create(
                        purchase_order=order,
                        vendor=order.vendor,
                        warehouse=order.warehouse,
                        reason='purchase',
                        created_by=request.user
                    )
                    
                    for item in order.items.all():
                        already_received = GoodsReceivedItem.objects.filter(
                            grn__purchase_order=order,
                            product=item.product
                        ).aggregate(total=Sum('qty'))['total'] or 0
                        
                        pending = item.qty - already_received
                        
                        if pending > 0:
                            GoodsReceivedItem.objects.create(
                                grn=grn,
                                product=item.product,
                                order_qty=item.qty,
                                qty=pending,
                                price=item.price,
                                notes=f"PO: {order.order_no}"
                            )
                    
                    order.update_receive_status()
                    count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        
        if count:
            self.message_user(request, f"✅ {count} GRNs created successfully.")
    create_grn.short_description = "📋 Create GRN (Receive Goods)"
    
    def update_po_status(self, request, queryset):
        selected_ids = queryset.values_list('id', flat=True)
        ids_str = ','.join(map(str, selected_ids))
        return redirect(f"/admin/app/purchaseorder/update-status/?ids={ids_str}")
    update_po_status.short_description = "📝 Update Status"
    
    def generate_po_pdf(self, request, queryset):
        """Generate PDF for POs"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        for order in queryset:
            story.append(Paragraph(f"Purchase Order - {order.order_no}", styles['Heading1']))
            story.append(Paragraph(f"Vendor: {order.vendor.name}", styles['Normal']))
            story.append(Paragraph(f"Status: {order.get_status_display()}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            data = [['Product', 'Order Qty', 'Received', 'Pending', 'Price', 'Total']]
            for item in order.items.all():
                received = order.received_qty(item.product)
                data.append([item.product.name, str(item.qty), str(received), str(item.qty - received), f"{item.price:,.2f}", f"{item.total_amt:,.2f}"])
            
            data.append(['', '', '', '', 'Total:', f"{order.total_amount():,.2f}"])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.5*inch))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    generate_po_pdf.short_description = "📄 Generate PO PDF"


# ============================================
# GRN ITEM INLINE
# ============================================
class GoodsReceivedItemInline(admin.TabularInline):
    model = GoodsReceivedItem
    fields = ('product', 'order_qty', 'qty', 'pending_qty', 'unit', 'price', 'total_amt', 'notes')
    readonly_fields = ('pending_qty', 'total_amt')
    extra = 10


# ============================================
# GRN ADMIN
# ============================================
@admin.register(GoodsReceivedNote, site=admin.site)
class GoodsReceivedNoteAdmin(admin.ModelAdmin):
    list_display = ('grn_no', 'grn_date', 'vendor', 'warehouse', 'po_link',
                    'reason', 'total_qty_display', 'total_items_display', 'converted_to_purchase', 'print_grn')
    list_filter = ('reason', 'grn_date', 'warehouse')
    search_fields = ('grn_no', 'vendor__name', 'purchase_order__order_no')
    inlines = [GoodsReceivedItemInline]
    change_list_template = "admin/button.html"
    actions = ['generate_grn_pdf', 'convert_grn_to_purchase']
    
    fieldsets = (
        ('GRN Information', {
            'fields': ('grn_no', 'purchase_order', 'vendor', 'warehouse', 'grn_date', 'reason')
        }),
        ('Supplier Invoice', {
            'fields': ('invoice_no',)
        }),
        ('Transport', {
            'fields': ('vehicle_no', 'driver_name'),
            'classes': ('collapse',)
        }),
        ('Remarks', {
            'fields': ('notes',)
        }),
    )
    
    readonly_fields = ('grn_no',)
    
    def po_link(self, obj):
        if obj.purchase_order:
            url = reverse('admin:app_purchaseorder_change', args=[obj.purchase_order.id])
            return format_html('<a href="{}">{}</a>', url, obj.purchase_order.order_no)
        return "-"
    po_link.short_description = "PO"
    
    def total_qty_display(self, obj):
        return f"{obj.total_qty():,.2f}"
    total_qty_display.short_description = "Total Qty"
    
    def total_items_display(self, obj):
        return obj.total_items()
    total_items_display.short_description = "Items"
    
    def print_grn(self, obj):
        url = reverse('admin:generate_grn_pdf', args=[obj.pk])
        return format_html('<a href="{}" target="_blank">📄 Print</a>', url)
    print_grn.short_description = "Print"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        if obj.purchase_order and not obj.vendor_id:
            obj.vendor = obj.purchase_order.vendor
        if obj.purchase_order and not obj.warehouse_id:
            obj.warehouse = obj.purchase_order.warehouse
        super().save_model(request, obj, form, change)
    
    def generate_grn_pdf(self, request, queryset):
        """Generate PDF for GRNs"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.3*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        for grn in queryset:
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER)
            story.append(Paragraph(company_name, title_style))
            story.append(Paragraph("<b>GOODS RECEIVED NOTE (GRN)</b>", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            details = [
                [Paragraph(f"<b>GRN No:</b> {grn.grn_no}"), Paragraph(f"<b>Date:</b> {grn.grn_date.strftime('%d-%m-%Y')}")],
                [Paragraph(f"<b>PO No:</b> {grn.purchase_order.order_no if grn.purchase_order else 'N/A'}"), Paragraph(f"<b>Invoice:</b> {grn.invoice_no or 'N/A'}")],
                [Paragraph(f"<b>Vendor:</b> {grn.vendor.name}"), Paragraph(f"<b>Vehicle:</b> {grn.vehicle_no or 'N/A'}")],
            ]
            t = Table(details, colWidths=[3.5*inch, 3.5*inch])
            story.append(t)
            story.append(Spacer(1, 0.2*inch))
            
            table_data = [['S.No', 'Product', 'Order Qty', 'Received', 'Pending', 'Unit']]
            total_order = 0
            total_rec = 0
            total_pend = 0
            
            for idx, item in enumerate(grn.items.all(), 1):
                table_data.append([str(idx), item.product.name, f"{item.order_qty:,.2f}", f"{item.qty:,.2f}", f"{item.pending_qty:,.2f}", item.unit or '-'])
                total_order += item.order_qty
                total_rec += item.qty
                total_pend += item.pending_qty
            
            table_data.append(['', 'TOTAL', f"{total_order:,.2f}", f"{total_rec:,.2f}", f"{total_pend:,.2f}", ''])
            
            items_table = Table(table_data, repeatRows=1)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 0.4*inch))
            
            sign = [
                [Paragraph("<b>Received by:</b> _________"), Paragraph("<b>Checked by:</b> _________")],
            ]
            story.append(Table(sign, colWidths=[3.5*inch, 3.5*inch]))
            story.append(Paragraph("<br/><br/>", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    generate_grn_pdf.short_description = "📄 Generate GRN PDF"
    
    def convert_grn_to_purchase(self, request, queryset):
        count = 0
        for grn in queryset.filter(converted_to_purchase=False):
            try:
                with transaction.atomic():
                    purchase = Purchase.objects.create(
                        warehouse=grn.warehouse,
                        bill_no=grn.invoice_no or f"BILL-{grn.grn_no}",
                        vendor=grn.vendor,
                        pur_date=now(),
                        created_by=request.user
                    )
                    for item in grn.items.all():
                        PurchaseItem.objects.create(purchase=purchase, product=item.product, qty=item.qty, price=item.price)
                    grn.converted_to_purchase = True
                    grn.save()
                    count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        if count:
            self.message_user(request, f"✅ {count} GRNs converted to purchase.")
    convert_grn_to_purchase.short_description = "🔄 Convert to Purchase"


# ============================================
# PURCHASE ADMIN
# ============================================
class PurchaseAdmin(admin.ModelAdmin):
    list_display = [
        'formatted_pur_date', 'bill_status', 'created_by', 'vendor', 'warehouse',
        'get_vendor_group', 'formatted_previous_balance', 'formatted_total_amount',
        'formatted_paid', 'formatted_outstanding_balance', 'formatted_outstanding_with_previous',
    ]
    search_fields = ('vendor__name','bill_no')
    search_help_text = "Search by Bill No, Vendor Name"
    list_filter = ['pur_date','vendor__group','warehouse',]
    inlines = [PurchaseItemInline]
    list_per_page = 50
    change_list_template = "admin/button.html"
    actions = ['generate_html_report', 'generate_vendor_ledger', 'generate_aging_report', 'generate_purchase_pdf']

    def save_model(self, request, obj, form, change):
        if not obj.created_by:  
            obj.created_by = request.user  
        super().save_model(request, obj, form, change)
    
    def formatted_pur_date(self, obj):
        return obj.pur_date.strftime('%d-%m-%Y')
    formatted_pur_date.short_description = 'Purchase Date'

    def formatted_previous_balance(self, obj):
        return f"{obj.previous_balance:,.2f} PKR"
    formatted_previous_balance.short_description = 'Previous Balance'

    def formatted_total_amount(self, obj):
        return f"{obj.total_amount():,.2f} PKR"
    formatted_total_amount.short_description = 'Total Amount'

    def formatted_paid(self, obj):
        return f"{obj.paid:,.2f} PKR"
    formatted_paid.short_description = 'Paid'

    def formatted_outstanding_balance(self, obj):
        return f"{obj.outstanding_balance():,.2f} PKR"
    formatted_outstanding_balance.short_description = 'Outstanding Balance'

    def formatted_outstanding_with_previous(self, obj):
        return f"{obj.outstanding_with_previous():,.2f} PKR"
    formatted_outstanding_with_previous.short_description = 'Outstanding (With Previous)'

    def get_vendor_group(self, obj):
        return obj.vendor.group.name if obj.vendor and obj.vendor.group else None
    get_vendor_group.short_description = 'Vendor Group'

    def generate_aging_report(self, request, queryset):
        today = datetime.now().date()
        aging_data = {
            '30_days': queryset.filter(pur_date__gte=today - timedelta(days=30)).aggregate(Sum('purchaseitem__total_amt'))['purchaseitem__total_amt__sum'] or 0,
            '60_days': queryset.filter(pur_date__gte=today - timedelta(days=60), pur_date__lt=today - timedelta(days=30)).aggregate(Sum('purchaseitem__total_amt'))['purchaseitem__total_amt__sum'] or 0,
            '90_days': queryset.filter(pur_date__lt=today - timedelta(days=60)).aggregate(Sum('purchaseitem__total_amt'))['purchaseitem__total_amt__sum'] or 0,
        }
        context = {'aging_data': aging_data, 'date': datetime.now()}
        html = render_to_string('admin/aging_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="aging_report.html"'
        return response
    generate_aging_report.short_description = "Generate Aging Report"

    def generate_vendor_ledger(self, request, queryset):
        total_outstanding = queryset.aggregate(
            total_outstanding=ExpressionWrapper(
            Sum('purchaseitem__total_amt') - Sum('paid'),
            output_field=DecimalField()
            )
            )['total_outstanding'] or 0
        context = {
        'purchases': queryset,
        'date': datetime.now(),
        'total_outstanding': total_outstanding
        }
        html = render_to_string('admin/vendor_ledger_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="vendor_ledger_report.html"'
        return response
    generate_vendor_ledger.short_description = "Generate Vendor Ledger Report"

    def generate_html_report(self, request, queryset):
        purchases = queryset
        context = {'purchases': purchases, 'date': datetime.now()}
        html = render_to_string('admin/purchase_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="purchase_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_purchase_pdf(self, request, queryset):
        """ReportLab use karke Purchase ka PDF report generate karein"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Purchase Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Bill No', 'Vendor', 'Group', 'Total', 'Paid', 'Outstanding']]
        
        total_amount = Decimal('0.0')
        total_paid = Decimal('0.0')
        total_outstanding = Decimal('0.0')
        
        for purchase in queryset:
            pur_total = purchase.total_amount()
            pur_outstanding = purchase.outstanding_balance()
            table_data.append([
                purchase.pur_date.strftime('%d-%m-%Y'),
                purchase.bill_no or 'Pending',
                purchase.vendor.name,
                purchase.vendor.group.name if purchase.vendor.group else '-',
                f"{pur_total:,.2f}",
                f"{purchase.paid:,.2f}",
                f"{pur_outstanding:,.2f}"
            ])
            total_amount += pur_total
            total_paid += purchase.paid
            total_outstanding += pur_outstanding
        
        table_data.append(['', '', '', 'TOTAL:', f"{total_amount:,.2f}", f"{total_paid:,.2f}", f"{total_outstanding:,.2f}"])
        
        table = Table(table_data, repeatRows=1, colWidths=[1*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="purchase_report.pdf"'
        return response
    generate_purchase_pdf.short_description = 'Generate PDF Report'


# ============================================
# PURCHASE ITEM ADMIN
# ============================================
class PurchaseItemAdmin(admin.ModelAdmin):
    list_display = ('get_pur_date','get_bill_no','vendor','get_vendor_group', 'product','get_category', 'qty', 'formatted_price', 'formatted_total')
    list_per_page = 25
    change_list_template = "admin/button.html"
    search_fields = ('product__name', 'purchase__bill_no', 'purchase__vendor__name')
    search_help_text = "Search by Product Name, Bill Number, or Vendor Name"
    list_filter = ('purchase__pur_date',)
    actions = ['generate_html_report', 'generate_purchase_item_pdf']

    def has_add_permission(self, request):
        return False

    def get_pur_date(self, obj):
        return obj.purchase.pur_date if obj.purchase else 'N/A'
    get_pur_date.short_description = 'Date'

    def get_bill_no(self, obj):
        return obj.purchase.bill_no if obj.purchase and hasattr(obj.purchase, 'bill_no') else 'N/A'
    get_bill_no.short_description = 'Bill No'
    
    def vendor(self, obj):
        return obj.purchase.vendor
    vendor.short_description = 'Vendor'

    def get_vendor_group(self, obj):
        return obj.purchase.vendor.group.name if obj.purchase and obj.purchase.vendor and obj.purchase.vendor.group else None
    get_vendor_group.short_description = 'Vendor Group'

    def get_category(self, obj):
        return obj.product.category.name if obj.product and obj.product.category else 'N/A'
    get_category.short_description = 'Category'

    def formatted_price(self, obj):
        return f"{obj.price:,.2f} PKR"
    formatted_price.short_description = 'Price'

    def formatted_total(self, obj):
        total = Decimal(obj.qty) * obj.price
        return f"{total:,.2f} PKR"
    formatted_total.short_description = 'Total Amount'

    def generate_html_report(self, request, queryset):
        PurchaseItems = queryset
        context = {'PurchaseItems': PurchaseItems, 'date': datetime.now()}
        html = render_to_string('admin/PurchaseItem_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="PurchaseItem_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_purchase_item_pdf(self, request, queryset):
        """Generate PDF report for Purchase Items."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Purchase Items Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Bill No', 'Vendor', 'Product', 'Qty', 'Price', 'Total']]
        
        total_qty = Decimal('0.0')
        total_amount = Decimal('0.0')
        
        for item in queryset:
            qty = Decimal(str(item.qty))
            total = qty * item.price
            table_data.append([
                item.purchase.pur_date.strftime('%d-%m-%Y') if item.purchase else 'N/A',
                item.purchase.bill_no if item.purchase else 'N/A',
                item.purchase.vendor.name if item.purchase else 'N/A',
                item.product.name,
                f"{qty:,.2f}",
                f"{item.price:,.2f}",
                f"{total:,.2f}"
            ])
            total_qty += qty
            total_amount += total
        
        table_data.append(['', '', '', 'TOTAL:', f"{total_qty:,.2f}", '', f"{total_amount:,.2f}"])
        
        table = Table(table_data, repeatRows=1, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 0.8*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="purchase_items_report.pdf"'
        return response
    generate_purchase_item_pdf.short_description = 'Generate PDF Report'


# ============================================
# CUSTOMER GROUP ADMIN
# ============================================
@admin.register(CustomerGroup, site=admin.site)
class CustomerGroupAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    change_list_template = "admin/button.html"
    search_help_text = "Search by Customer Group Name"
    list_display = ['name', 'description', 'total_outstanding_balance_display']
    actions = ['generate_html_report', 'generate_customer_group_pdf']
    
    def total_outstanding_balance_display(self, obj):
        return f'{obj.total_outstanding_balance():,.2f}'
    total_outstanding_balance_display.short_description = 'Total Outstanding Balance'

    def generate_html_report(self, request, queryset):
        CustomerGroups = queryset
        context = {'CustomerGroups': CustomerGroups, 'date': datetime.now()}
        html = render_to_string('admin/CustomerGroup_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="CustomerGroup_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_customer_group_pdf(self, request, queryset):
        """Generate PDF report for Customer Groups."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Customer Group Report", title_style))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Group Name', 'Description', 'Outstanding Balance']]
        
        for group in queryset:
            table_data.append([
                group.name,
                group.description or '-',
                f"{group.total_outstanding_balance():,.2f}"
            ])
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customer_group_report.pdf"'
        return response
    generate_customer_group_pdf.short_description = 'Generate PDF Report'


# ============================================
# CUSTOMER ADMIN
# ============================================
class CustomerAdmin(admin.ModelAdmin):
    search_fields = ('customer_code', 'name', 'contact_number')
    search_help_text = "Search by Customer Code, Name, or Contact Number"
    list_display = ('customer_code', 'name', 'group', 'contact_number', 'address', 
                    'adjusted_outstanding_balance', 'profit_margin')
    list_editable = ('profit_margin',)
    ordering = ('customer_code',)
    change_list_template = "admin/button.html"
    actions = ['generate_html_report', 'generate_customer_pdf']
    readonly_fields = ('adjusted_outstanding_balance',)
    
    fieldsets = (
        ('Customer Information', {
            'fields': ('customer_code', 'name', 'group', 'address', 'profit_margin')
        }),
        ('Contact Details', {
            'fields': ('contact_number', 
                       ('ref_name_1', 'ref_contact_number_1'),
                       ('ref_name_2', 'ref_contact_number_2'))
        }),
    )

    def adjusted_outstanding_balance(self, obj):
        return f'{obj.adjusted_outstanding_balance():,.2f}'
    adjusted_outstanding_balance.short_description = 'Total Outstanding Balance'

    def generate_html_report(self, request, queryset):
        customers = queryset
        context = {'customers': customers, 'date': datetime.now()}
        html = render_to_string('admin/customer_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="customer_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_customer_pdf(self, request, queryset):
        """Generate PDF report for Customers."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, 
                                      alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Customer Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        # Updated table header with Customer Code
        table_data = [['Code', 'Name', 'Group', 'Address', 'Contact', 'Margin %', 'Outstanding']]
        
        total_outstanding = Decimal('0.0')
        for customer in queryset:
            balance = customer.adjusted_outstanding_balance()
            table_data.append([
                customer.customer_code or '-',
                customer.name,
                customer.group.name if customer.group else '-',
                (customer.address[:30] + '...') if customer.address and len(customer.address) > 30 
                    else (customer.address or '-'),
                customer.contact_number or '-',
                f"{customer.profit_margin:.2f}%",
                f"{balance:,.2f}"
            ])
            total_outstanding += balance
        
        table_data.append(['', '', '', '', '', 'TOTAL:', f"{total_outstanding:,.2f}"])
        
        table = Table(table_data, repeatRows=1, 
                      colWidths=[0.8*inch, 1.3*inch, 1*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (5, 1), (6, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customer_report.pdf"'
        return response
    generate_customer_pdf.short_description = 'Generate PDF Report'


# ============================================
# SALE ITEM ADMIN
# ============================================
class SaleItemAdmin(admin.ModelAdmin):
    list_display = ('get_sale_date', 'get_bill_no','customer', 'get_customer_group', 'product','get_unit', 'get_category', 'qty', 'formatted_price', 'discount_applied', 'formatted_total', 'discounted_total', 'formatted_profit', 'get_purchase_bill_no')
    list_per_page = 25
    change_list_template = "admin/button.html"
    search_fields = ('product__name', 'sale__bill_no', 'sale__customer__name')
    search_help_text = "Search by Product Name, Bill Number, or Customer Name"
    list_filter = ('sale__sale_date', 'product__category', 'sale__customer')
    actions = ['generate_html_report', 'generate_sale_item_pdf']

    def has_add_permission(self, request):
        return False

    def get_purchase_bill_no(self, obj):
        if not obj.batches_used.exists():
            return "N/A"
        bill_nos = set(
            batch.purchase_item.purchase.bill_no 
            for batch in obj.batches_used.all() 
            if batch.purchase_item and batch.purchase_item.purchase.bill_no
        )
        return ", ".join(bill_nos) if bill_nos else "Waiting for Bill"
    get_purchase_bill_no.short_description = "Purchase Bill No"

    def discount_applied(self, obj):
        total_sale_amount = obj.sale.total_without_discount()
        if total_sale_amount > 0:
            discount_ratio = obj.total_amt / total_sale_amount
            discount = discount_ratio * obj.sale.discount_value
            return round(discount, 2)
        return 0.00

    def discounted_total(self, obj):
        return round(obj.total_amt - self.discount_applied(obj), 2)

    discount_applied.short_description = "Discount Applied"
    discounted_total.short_description = "Discounted Total"

    def customer(self, obj):
        return obj.sale.customer
    customer.short_description = 'Customer'

    def formatted_price(self, obj):
        return f"{obj.price:,.2f} PKR"
    formatted_price.short_description = 'Price'

    def formatted_total(self, obj):
        total = Decimal(obj.qty) * obj.price
        return f"{total:,.2f} PKR"
    formatted_total.short_description = 'Total Amount'

    def formatted_profit(self, obj):
        adjusted_profit = obj.profit - self.discount_applied(obj)
        return f"{adjusted_profit:,.2f} PKR"
    formatted_profit.short_description = 'Profit (Adjusted)'

    def get_sale_date(self, obj):
        return obj.sale.sale_date if obj.sale else 'N/A'
    get_sale_date.short_description = 'Date'

    def get_bill_no(self, obj):
        return obj.sale.bill_no if obj.sale and hasattr(obj.sale, 'bill_no') else 'N/A'
    get_bill_no.short_description = 'Bill No'

    def get_category(self, obj):
        return obj.product.category.name if obj.product and obj.product.category else 'N/A'
    get_category.short_description = 'Category'

    def get_unit(self, obj):
        return obj.product.unit
    get_unit.short_description = 'Unit'

    def generate_html_report(self, request, queryset):
        SaleItems = queryset
        context = {'SaleItems': SaleItems, 'date': datetime.now()}
        html = render_to_string('admin/SaleItem_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="SaleItem_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'

    def view_invoice(self, obj):
        url = reverse('generate_invoice', args=[obj.sale.pk])
        return format_html(f'<a href="{url}" target="_blank">View Invoice</a>')
    view_invoice.short_description = 'Invoice'

    def get_customer_group(self, obj):
        return obj.sale.customer.group.name if obj.sale and obj.sale.customer and obj.sale.customer.group else None
    get_customer_group.short_description = 'Customer Group'
    
    def generate_sale_item_pdf(self, request, queryset):
        """Generate PDF report for Sale Items."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Sale Items Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Bill No', 'Customer', 'Product', 'Qty', 'Price', 'Total', 'Profit']]
        
        total_qty = Decimal('0.0')
        total_amount = Decimal('0.0')
        total_profit = Decimal('0.0')
        
        for item in queryset:
            qty = Decimal(str(item.qty))
            total = qty * item.price
            profit = item.profit - item.discount if hasattr(item, 'discount') else item.profit
            table_data.append([
                item.sale.sale_date.strftime('%d-%m-%Y') if item.sale else 'N/A',
                item.sale.bill_no if item.sale else 'N/A',
                item.sale.customer.name if item.sale else 'N/A',
                item.product.name,
                f"{qty:,.2f}",
                f"{item.price:,.2f}",
                f"{total:,.2f}",
                f"{profit:,.2f}"
            ])
            total_qty += qty
            total_amount += total
            total_profit += profit
        
        table_data.append(['', '', '', 'TOTAL:', f"{total_qty:,.2f}", '', f"{total_amount:,.2f}", f"{total_profit:,.2f}"])
        
        table = Table(table_data, repeatRows=1, colWidths=[1*inch, 0.9*inch, 1.3*inch, 1.8*inch, 0.7*inch, 0.9*inch, 0.9*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sale_items_report.pdf"'
        return response
    generate_sale_item_pdf.short_description = 'Generate PDF Report'


# ============================================
# SALE ADMIN

# ============================================
# SALE ADMIN
# ============================================
class SaleAdmin(admin.ModelAdmin):
    list_display = ('sale_date','customer', 'warehouse', 'get_customer_group','bill_no','created_by',
                    'formatted_previous_balance', 'formatted_total_without_discount', 'discount_value', 
                    'formatted_total_amount', 'formatted_paid', 'formatted_total_outstanding_with_previous',
                    'formatted_total_profit', 'view_invoice')
    search_fields = ('customer__name','bill_no')
    search_help_text = "Search by Bill No, Customer Name"
    list_filter = ('sale_date','customer__group','warehouse',)
    change_form_template = "admin/barcode_sale.html"
    change_list_template = "admin/button.html"
    inlines = [SaleItemInline]
    actions = [
        'generate_html_report', 
        'generate_customer_ledger', 
        'generate_sales_profit_analysis', 
        'generate_sale_pdf',
        'send_whatsapp_invoice',
        'send_pdf_invoice_whatsapp',
        'send_payment_reminders',
    ]
    list_per_page = 50

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['warehouse'].required = True
        if not Warehouse.objects.exists():
            form.base_fields['warehouse'].help_text = "کوئی ویئرہاؤس موجود نہیں۔ براہ کرم پہلے ایک بنائیں۔"
        return form

    def formatted_sale_date(self, obj):
        return obj.sale_date.strftime('%d-%m-%y')
    formatted_sale_date.short_description = 'Sale Date'

    def formatted_total_without_discount(self, obj):
        return f"{obj.total_without_discount():,.2f} PKR"
    formatted_total_without_discount.short_description = 'Total_without_Discount'

    def formatted_total_amount(self, obj):
        return f"{obj.total_amount():,.2f} PKR"
    formatted_total_amount.short_description = 'Total Amount'

    def formatted_total_profit(self, obj):
        return f"{obj.total_profit():,.2f} PKR"
    formatted_total_profit.short_description = 'Total Profit'

    def formatted_paid(self, obj):
        return f"{obj.paid:,.2f} PKR"
    formatted_paid.short_description = 'Paid'

    def formatted_previous_balance(self, obj):
        return f"{obj.previous_balance:,.2f} PKR"
    formatted_previous_balance.short_description = 'Previous Balance'

    def formatted_total_outstanding_with_previous(self, obj):
        return f"{obj.total_outstanding_with_previous:,.2f} PKR"
    formatted_total_outstanding_with_previous.short_description = 'Total_Outstanding_With_Previous'

    def generate_sales_profit_analysis(self, request, queryset):
        profit_data = queryset.aggregate(
            total_sales=Sum('saleitem__total_amt'),
            total_profit=Sum('saleitem__profit')
        )
        context = {'profit_data': profit_data, 'date': datetime.now()}
        html = render_to_string('admin/sales_profit_analysis.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="sales_profit_analysis.html"'
        return response
    generate_sales_profit_analysis.short_description = "Generate Sales & Profit Analysis"

    def generate_html_report(self, request, queryset):
        sales = queryset
        context = {'sales': sales, 'date': datetime.now()}
        html = render_to_string('admin/sale_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="sale_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'

    def generate_customer_ledger(self, request, queryset):
        total_outstanding = queryset.aggregate(
            total_outstanding=ExpressionWrapper(
            Sum('saleitem__total_amt') - Sum('paid'),
            output_field=DecimalField()
            )
            )['total_outstanding'] or 0
        context = {'sales': queryset, 'date': datetime.now(), 'total_outstanding': total_outstanding}
        html = render_to_string('admin/customer_ledger_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="customer_ledger_report.html"'
        return response
    generate_customer_ledger.short_description = "Generate Customer Ledger Report"

    def view_invoice(self, obj):
        url = reverse('admin:generate_invoice', args=[obj.pk])
        return format_html(f'<a href="{url}" target="_blank">View Invoice</a>')
    view_invoice.short_description = 'Invoice'

    def get_customer_group(self, obj):
        return obj.customer.group.name if obj.customer and obj.customer.group else None
    get_customer_group.short_description = 'Customer Group'

    def save_model(self, request, obj, form, change):
        if not obj.created_by:  
            obj.created_by = request.user 
        if not obj.warehouse:
            default_warehouse = Warehouse.objects.first()
            if default_warehouse:
                obj.warehouse = default_warehouse
            else:
                self.message_user(request, "کوئی ویئرہاؤس موجود نہیں۔ براہ کرم پہلے ایک بنائیں۔", level=messages.ERROR)
                return
        super().save_model(request, obj, form, change)
    
    def generate_sale_pdf(self, request, queryset):
        """Sale ka PDF report generate karein"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Sale Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Date', 'Bill No', 'Customer', 'Group', 'Total', 'Discount', 'Paid', 'Profit']]
        
        total_amount = Decimal('0.0')
        total_discount = Decimal('0.0')
        total_paid = Decimal('0.0')
        total_profit = Decimal('0.0')
        
        for sale in queryset:
            total_amt = sale.total_without_discount()
            table_data.append([
                sale.sale_date.strftime('%d-%m-%Y'),
                sale.bill_no,
                sale.customer.name,
                sale.customer.group.name if sale.customer.group else '-',
                f"{total_amt:,.2f}",
                f"{sale.discount_value:,.2f}",
                f"{sale.paid:,.2f}",
                f"{sale.total_profit():,.2f}"
            ])
            total_amount += total_amt
            total_discount += sale.discount_value
            total_paid += sale.paid
            total_profit += sale.total_profit()
        
        table_data.append(['', '', '', 'TOTAL:', f"{total_amount:,.2f}", f"{total_discount:,.2f}", f"{total_paid:,.2f}", f"{total_profit:,.2f}"])
        
        table = Table(table_data, repeatRows=1, colWidths=[1*inch, 1*inch, 1.5*inch, 1*inch, 1*inch, 0.9*inch, 0.9*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="sale_report.pdf"'
        return response
    generate_sale_pdf.short_description = 'Generate PDF Report'

    # ============================================
    # WHATSAPP - TEXT INVOICE
    # ============================================
    def send_whatsapp_invoice(self, request, queryset):
        """Selected sales ka TEXT invoice WhatsApp pe bhejo"""
        from .whatsapp_utils import WhatsAppSender
        
        count = 0
        links = []
        for sale in queryset:
            if sale.customer.contact_number:
                result = WhatsAppSender.send_invoice(
                    sale.customer.contact_number,
                    sale.bill_no,
                    sale.total_amount()
                )
                if result.get('success'):
                    count += 1
                    links.append({
                        'bill': sale.bill_no,
                        'customer': sale.customer.name,
                        'url': result['url']
                    })
        
        if count > 0:
            msg = f"✅ {count} WhatsApp links generated!<br><br>"
            for link in links[:5]:
                msg += f"📋 {link['bill']} - {link['customer']}<br>"
            if len(links) > 5:
                msg += f"... and {len(links)-5} more<br>"
            if links:
                msg += f"<br><a href='{links[0]['url']}' target='_blank' style='background:#25D366; color:white; padding:8px 15px; border-radius:5px; text-decoration:none;'>📱 Send First Invoice</a>"
            
            self.message_user(request, format_html(msg))
        else:
            self.message_user(request, "❌ Selected customers don't have phone numbers!", level=messages.WARNING)

    send_whatsapp_invoice.short_description = "📱 Send TEXT Invoice on WhatsApp"

    # ============================================
    # WHATSAPP - PDF INVOICE
    # ============================================
    def send_pdf_invoice_whatsapp(self, request, queryset):
        """Selected sales ka PDF invoice WhatsApp pe bhejo"""
        from .whatsapp_utils import WhatsAppSender
        
        count = 0
        links = []
        for sale in queryset:
            if sale.customer.contact_number:
                result = WhatsAppSender.send_pdf_invoice_link(sale)
                if result.get('success'):
                    count += 1
                    links.append({
                        'bill': sale.bill_no,
                        'customer': sale.customer.name,
                        'url': result['url'],
                        'filepath': result.get('filepath', '')
                    })
        
        if count > 0:
            msg = f"✅ {count} PDF invoices generated!<br><br>"
            for link in links[:5]:
                msg += f"📋 {link['bill']} - {link['customer']}<br>"
            if links:
                msg += f"<br><a href='{links[0]['url']}' target='_blank' style='background:#25D366; color:white; padding:8px 15px; border-radius:5px; text-decoration:none;'>📱 Send First PDF Invoice</a>"
            self.message_user(request, format_html(msg))
        else:
            self.message_user(request, "❌ Selected customers don't have phone numbers!", level=messages.WARNING)

    send_pdf_invoice_whatsapp.short_description = "📄 Send PDF Invoice on WhatsApp"

    # ============================================
    # WHATSAPP - PAYMENT REMINDERS
    # ============================================
    def send_payment_reminders(self, request, queryset):
        """Outstanding customers ko reminder bhejo"""
        from .whatsapp_utils import WhatsAppBulkSender
        
        links = WhatsAppBulkSender.generate_reminder_links()
        
        if links:
            msg = f"✅ {len(links)} payment reminder links generated!<br><br>"
            msg += f"<a href='/admin/whatsapp/reminders/' style='background:#25D366; color:white; padding:8px 15px; border-radius:5px; text-decoration:none;'>📱 View All Reminders & Send</a>"
            self.message_user(request, format_html(msg))
        else:
            self.message_user(request, "🎉 All payments clear! No reminders needed.")

    send_payment_reminders.short_description = "💰 Send Payment Reminders"

# ============================================
# INVENTORY ADMIN
# ============================================
class InventoryAdmin(admin.ModelAdmin):
    list_display = ('product', 'get_description', 'get_unit', 'get_location', 'get_used', 'get_brand', 
                    'get_category', 'get_types', 'warehouse', 'stock', 'stock_value_display', 'low_stock_alert')
    ordering = ('product',)
    change_list_template = "admin/button.html"
    actions = ['send_bulk_low_stock_alerts', 'generate_html_report', 'generate_inventory_pdf']
    search_fields = (
        'product__name', 'warehouse__name', 'product__unit__name', 'product__location__name',
        'product__used', 'product__brand__name', 'product__category__name', 'product__types__name',
        'product__description',
    )
    search_help_text = 'Search by Product Name, Warehouse, Unit, Location, Used, Brand, Category, Type'

    def has_add_permission(self, request):
        return False

    def send_bulk_low_stock_alerts(self, request, queryset):
        Inventory.send_bulk_low_stock_alert()
        self.message_user(request, "Low stock alert email sent successfully.")
    send_bulk_low_stock_alerts.short_description = "Send Low Stock Alert"
    
    def stock_value_display(self, obj):
        stock_value = obj.stock_value()
        return f"{stock_value:,.2f} PKR"
    stock_value_display.short_description = 'Stock Value'

    def low_stock_alert(self, obj):
        if obj.check_low_stock():
            return format_html('<span style="color: red;">Low Stock</span>')
        return "OK"
    low_stock_alert.short_description = "Stock Alert"

    def get_description(self, obj):
        return obj.product.description if obj.product.description else "N/A"
    get_description.short_description = "Description"
    
    def get_unit(self, obj):
        return obj.product.unit.name if obj.product.unit else "N/A"
    get_unit.short_description = "Unit"

    def get_location(self, obj):
        return obj.product.location.name if obj.product.location else "N/A"
    get_location.short_description = "Location"

    def get_used(self, obj):
        return obj.product.used if obj.product.used else "N/A"
    get_used.short_description = "Used"

    def get_brand(self, obj):
        return obj.product.brand.name if obj.product.brand else "N/A"
    get_brand.short_description = "Brand"

    def get_category(self, obj):
        return obj.product.category.name if obj.product.category else "N/A"
    get_category.short_description = "Category"

    def get_types(self, obj):
        return obj.product.types.name if obj.product.types else "N/A"
    get_types.short_description = "Type"

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(request, extra_context=extra_context)
        qs = self.get_queryset(request)
        total_stock_value = sum(item.stock_value() for item in qs) or Decimal('0.0')
        summary_data = {'total_stock_value': f"{total_stock_value:,.2f} PKR"}
        if hasattr(response, 'context_data'):
            response.context_data['summary_data'] = summary_data
        return response

    def generate_html_report(self, request, queryset):
        inventorys = queryset
        context = {'inventorys': inventorys, 'date': datetime.now()}
        html = render_to_string('admin/inventory_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_inventory_pdf(self, request, queryset):
        """Inventory stock ka PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
        story.append(Paragraph(company_name, title_style))
        story.append(Paragraph("Inventory Stock Report", title_style))
        story.append(Paragraph(f"Date: {datetime.now().strftime('%d-%m-%Y')}", styles['Normal']))
        story.append(Spacer(1, 0.3*inch))
        
        table_data = [['Product', 'Warehouse', 'Stock', 'Price', 'Stock Value', 'Status']]
        
        total_value = Decimal('0.0')
        
        for idx, inv in enumerate(queryset, start=1):
            stock_value = inv.stock_value()
            status = "LOW STOCK!" if inv.check_low_stock() else "OK"
            table_data.append([
                inv.product.name[:30],
                inv.warehouse.name,
                f"{inv.stock:,.2f}",
                f"{inv.product.price:,.2f}",
                f"{stock_value:,.2f}",
                status
            ])
            total_value += Decimal(str(stock_value))
        
        table_data.append(['', '', '', 'TOTAL VALUE:', f"{total_value:,.2f}", ''])
        
        table = Table(table_data, repeatRows=1, colWidths=[2*inch, 1.2*inch, 0.8*inch, 1*inch, 1.2*inch, 1*inch])
        
        # Pehle base style apply karein
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]
        
        # Low stock rows ko highlight karein
        row_idx = 1
        for inv in queryset:
            if inv.check_low_stock():
                table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightpink))
                table_style.append(('TEXTCOLOR', (0, row_idx), (-1, row_idx), colors.red))
            row_idx += 1
        
        table.setStyle(TableStyle(table_style))
        story.append(table)
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
        return response
    generate_inventory_pdf.short_description = 'Generate PDF Report'


# ============================================
# STOCK BATCH ADMIN
# ============================================
# ============================================
# STOCK BATCH ADMIN
# ============================================
class StockBatchAdmin(admin.ModelAdmin):
    list_display = ('product', 'warehouse', 'qty', 'price', 'selling_price', 
                    'profit_display', 'profit_percent_display', 'remaining_qty', 
                    'batch_value_display', 'selling_value_display',
                    'get_purchase_bill_no', 'get_sale_bill_no')
    list_editable = ('selling_price',)
    search_fields = (
        'product__name', 'purchase_item__purchase__bill_no', 
        'sale_items__sale__bill_no', 'warehouse__name',
    )
    search_help_text = "Search by Product Name, Warehouse, Purchase Bill No, or Sale Bill No"
    list_filter = ('warehouse',)
    change_list_template = "admin/button.html"
    
    fields = ('product', 'warehouse', 'qty', 'price', 'selling_price', 'remaining_qty', 'purchase_item')
    readonly_fields = ('product', 'warehouse', 'qty', 'price', 'remaining_qty', 'purchase_item')

    def has_add_permission(self, request):
        return False

    # ============================================
    # DISPLAY FIELDS
    # ============================================
    def batch_value_display(self, obj):
        """Purchase price pe total value"""
        return f"Rs. {Decimal(obj.remaining_qty) * obj.price:,.2f}"
    batch_value_display.short_description = '📦 Stock Value'

    def selling_value_display(self, obj):
        """Selling price pe total value"""
        val = obj.selling_value()
        if obj.selling_price > 0:
            return f"Rs. {val:,.2f}"
        return "-"
    selling_value_display.short_description = '💰 Selling Value'

    def profit_display(self, obj):
        """Profit per unit with color"""
        profit = obj.profit_per_unit()
        if profit > 0:
            return format_html(
                '<span style="color:#28a745;font-weight:bold;">+Rs. {}</span>', 
                round(profit, 2)
            )
        elif obj.selling_price > 0 and profit == 0:
            return format_html('<span style="color:#ff9800;">Rs. 0 (No Profit)</span>')
        return "-"
    profit_display.short_description = '💰 Profit/Unit'
    
    def profit_percent_display(self, obj):
        """Profit margin percentage"""
        percent = obj.profit_margin_percent()
        if percent > 0:
            return format_html(
                '<span style="color:#28a745;font-weight:bold;">{}% ↑</span>', 
                round(percent, 1)
            )
        elif obj.selling_price > 0 and percent == 0:
            return format_html('<span style="color:#ff9800;">0%</span>')
        return "-"
    profit_percent_display.short_description = '📈 Margin %'

    def get_purchase_bill_no(self, obj):
        if obj.purchase_item and obj.purchase_item.purchase:
            return obj.purchase_item.purchase.bill_no if obj.purchase_item.purchase.bill_no else "Waiting for Bill"
        return "N/A"
    get_purchase_bill_no.short_description = "Purchase Bill No"

    def get_sale_bill_no(self, obj):
        if not obj.sale_items.exists():
            return "Not Sold Yet"
        bill_nos = set(
            sale_item.sale.bill_no 
            for sale_item in obj.sale_items.all() 
            if sale_item.sale and sale_item.sale.bill_no
        )
        return ", ".join(bill_nos) if bill_nos else "N/A"
    get_sale_bill_no.short_description = "Sale Bill No"

# ============================================
# MONTHLY CLOSING ADMIN
# ============================================
class MonthlyClosingAdmin(admin.ModelAdmin):
    list_display = ['month', 'previous_balance_display','total_purchase_balance_display',
                    'total_sale_balance_display', 'closing_balance_display',
                    'total_sale_profit_display', 'total_purchase_return_display',
                    'total_sale_return_display', 'net_profit_margin_display', 
                    'return_on_investment_display', 'locked']
    actions = ['lock_month', 'unlock_month', 'generate_html_report', 'generate_pnl', 
               'generate_cash_flow', 'generate_saving', 'generate_closing_pdf']
    ordering = ('-month',)
    change_list_template = "admin/button.html"

    def lock_month(self, request, queryset):
        queryset.update(locked=True)
        self.message_user(request, "Selected months have been locked.")
    lock_month.short_description = "Lock selected months"

    def unlock_month(self, request, queryset):
        queryset.update(locked=False)
        self.message_user(request, "Selected months have been unlocked.")
    unlock_month.short_description = "Unlock selected months"

    def total_purchase_return_display(self, obj):
        return f'{obj.total_purchase_return():,.2f}'
    total_purchase_return_display.short_description = 'Purchase Returns'

    def total_sale_return_display(self, obj):
        return f'{obj.total_sale_return():,.2f}'
    total_sale_return_display.short_description = 'Sale Returns'

    def total_purchase_balance_display(self, obj):
        return f'{obj.total_purchase_balance():,.2f}'
    total_purchase_balance_display.short_description = 'Total Purchase'

    def total_sale_profit_display(self, obj):
        return f'{obj.total_sale_profit():,.2f}'
    total_sale_profit_display.short_description = 'Total Profit'

    def total_sale_balance_display(self, obj):
        return f'{obj.total_sale_balance():,.2f}'
    total_sale_balance_display.short_description = 'Total Sale (COGS)'

    def closing_balance_display(self, obj):
        return f'{obj.closing_balance():,.2f}'
    closing_balance_display.short_description = 'Closing Balance'
    
    def previous_balance_display(self, obj):
        return f"{obj.get_previous_month_closing():,.2f}"
    previous_balance_display.short_description = "Previous Balance"

    def net_profit_margin_display(self, obj):
        return f"{obj.net_profit_margin():.2f}%"
    net_profit_margin_display.short_description = "Net Profit Margin"

    def return_on_investment_display(self, obj):
        return f"{obj.return_on_investment():.2f}%"
    return_on_investment_display.short_description = "ROI"

    def generate_pnl(self, request, queryset):
        for closing in queryset:
            income = closing.total_sale_profit()
            expenses = (Expense.objects.filter(date__month=closing.month.month).aggregate(Sum('amount'))['amount__sum'] or 0)
            profit = income - expenses
            context = {'income': income, 'expenses': expenses, 'profit': profit, 'month': closing.month}
            html = render_to_string('admin/pnl_report.html', context)
            response = HttpResponse(html, content_type='text/html')
            response['Content-Disposition'] = f'attachment; filename="pnl_{closing.month}.html"'
            return response
    generate_pnl.short_description = "Generate Profit & Loss Statement"

    def generate_saving(self, request, queryset):
        for closing in queryset:
            savings = (Saving.objects.filter(date__month=closing.month.month).aggregate(Sum('amount'))['amount__sum'] or 0)
            context = {'savings': savings, 'month': closing.month}
            html = render_to_string('admin/saving_report.html', context)
            response = HttpResponse(html, content_type='text/html')
            response['Content-Disposition'] = f'attachment; filename="saving_{closing.month}.html"'
            return response
    generate_saving.short_description = "Generate Saving"

    def generate_cash_flow(self, request, queryset):
        for closing in queryset:
            cash_in = Sale.objects.filter(sale_date__month=closing.month.month).aggregate(Sum('paid'))['paid__sum'] or 0
            cash_out = Purchase.objects.filter(pur_date__month=closing.month.month).aggregate(Sum('paid'))['paid__sum'] or 0
            net_cash = cash_in - cash_out
            context = {'cash_in': cash_in, 'cash_out': cash_out, 'net_cash': net_cash, 'month': closing.month}
            html = render_to_string('admin/cash_flow_report.html', context)
            response = HttpResponse(html, content_type='text/html')
            response['Content-Disposition'] = f'attachment; filename="cash_flow_{closing.month}.html"'
            return response
    generate_cash_flow.short_description = "Generate Cash Flow Statement"

    def generate_html_report(self, request, queryset):
        MonthlyClosings = queryset
        context = {'MonthlyClosings': MonthlyClosings, 'date': datetime.now()}
        html = render_to_string('admin/MonthlyClosing_report.html', context)
        response = HttpResponse(html, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="MonthlyClosing_report.html"'
        return response
    generate_html_report.short_description = 'Generate HTML Report'
    
    def generate_closing_pdf(self, request, queryset):
        """Monthly closing ka PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=10)
        
        for closing in queryset:
            story.append(Paragraph(company_name, title_style))
            story.append(Paragraph(f"Monthly Closing Report - {closing.month.strftime('%B %Y')}", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            data = [
                ['Description', 'Amount (PKR)'],
                ['Previous Balance', f"{closing.previous_balance:,.2f}"],
                ['Total Purchases', f"{closing.total_purchase_balance():,.2f}"],
                ['Purchase Returns', f"{closing.total_purchase_return():,.2f}"],
                ['Total Sales (COGS)', f"{closing.total_sale_balance():,.2f}"],
                ['Sale Returns', f"{closing.total_sale_return():,.2f}"],
                ['Total Profit', f"{closing.total_sale_profit():,.2f}"],
                ['Closing Balance', f"{closing.closing_balance():,.2f}"],
                ['Net Profit Margin', f"{closing.net_profit_margin():.2f}%"],
                ['Return on Investment (ROI)', f"{closing.return_on_investment():.2f}%"],
            ]
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, -2), (0, -1), 'Helvetica-Bold'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.5*inch))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="monthly_closing.pdf"'
        return response
    generate_closing_pdf.short_description = 'Generate PDF Report'


# ============================================
# PURCHASE RETURN ADMIN
# ============================================
class PurchaseRetrnAdmin(admin.ModelAdmin):
    inlines = [PurchaseRetrnItemInline]
    list_display = ('purchase_return_date','bill_no', 'purchase', 'vendor', 'created_by', 'return_reason')
    change_list_template = "admin/button.html"

    def save_model(self, request, obj, form, change):
        if not obj.created_by:  
            obj.created_by = request.user  
        super().save_model(request, obj, form, change)


class PurchaseRetrnItemAdmin(admin.ModelAdmin):
    list_display = ('purchase_return', 'product', 'qty', 'price', 'total_amt')
    change_list_template = "admin/button.html"

    def has_add_permission(self, request):
        return False


# ============================================
# SALE RETURN ADMIN
# ============================================
class SaleRetrnAdmin(admin.ModelAdmin):
    inlines = [SaleRetrnItemInline]
    list_display = ('sale_return_date','bill_no', 'sale', 'customer', 'created_by', 'return_reason')
    change_list_template = "admin/button.html"

    def save_model(self, request, obj, form, change):
        if not obj.created_by:  
            obj.created_by = request.user  
        super().save_model(request, obj, form, change)


class SaleRetrnItemAdmin(admin.ModelAdmin):
    list_display = ('sale_return', 'product', 'qty', 'price', 'total_amt')
    change_list_template = "admin/button.html"

    def has_add_permission(self, request):
        return False


# ============================================
# STOCK ADJUSTMENT ADMIN
# ============================================
@admin.register(StockAdjustmentItem, site=admin.site)
class StockAdjustmentItemAdmin(admin.ModelAdmin):
    list_display = ('adjustment', 'product', 'qty', 'price', 'total_value')
    search_fields = ('product__name',)
    readonly_fields = ('total_value',)
    change_list_template = "admin/button.html"

    def has_add_permission(self, request):
        return False


class StockAdjustmentItemInline(admin.TabularInline):
    model = StockAdjustmentItem
    extra = 1
    fields = ('product', 'qty', 'price', 'total_value')
    readonly_fields = ('total_value',)


@admin.register(StockAdjustment, site=admin.site)
class StockAdjustmentAdmin(admin.ModelAdmin):
    list_display = ('adjustment_type', 'reason', 'adjustment_date', 'created_by')
    change_list_template = "admin/button.html"
    search_fields = ('reason', 'created_by__username')
    inlines = [StockAdjustmentItemInline]

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)


# ============================================
# LOG ENTRY ADMIN
# ============================================
class ActionFlagFilter(SimpleListFilter):
    title = "Action Type"
    parameter_name = "action_flag"

    def lookups(self, request, model_admin):
        return [(1, "Added"), (2, "Changed"), (3, "Deleted")]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(action_flag=self.value())


class ContentTypeFilter(SimpleListFilter):
    title = "Content Type"
    parameter_name = "content_type"

    def lookups(self, request, model_admin):
        content_types = set(model_admin.get_queryset(request).values_list("content_type__id", "content_type__model"))
        return [(ct_id, model.replace("_", " ").title()) for ct_id, model in content_types]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(content_type__id=self.value())


class LogEntryAdmin(admin.ModelAdmin):
    change_list_template = "admin/button.html"
    list_display = ['action_time', 'user', 'content_type_display', 'action']
    list_filter = ['action_time', 'user', ActionFlagFilter, ContentTypeFilter]

    @admin.display(description="Action", ordering="action_flag")
    def action(self, obj):
        if obj.action_flag == 1:
            return format_html('<span style="color: blue; font-weight: bold;">Added</span>')
        elif obj.action_flag == 2:
            return format_html('<span style="color: orange; font-weight: bold;">Changed</span>')
        elif obj.action_flag == 3:
            return format_html('<span style="color: red; font-weight: bold;">Deleted</span>')
        return "Unknown"

    @admin.display(description="Content Type")
    def content_type_display(self, obj):
        return obj.content_type.model_class()._meta.verbose_name.title()

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return True 


# ============================================
# LICENSE ADMIN
# ============================================
class LicenseAdmin(admin.ModelAdmin):
    list_display = ('key', 'expiry_date', 'created_at', 'is_valid')
    readonly_fields = ('created_at',)
    change_list_template = "admin/button.html"

    def is_valid(self, obj):
        return "✔ Active" if obj.is_valid() else "❌ Expired"
    is_valid.short_description = "Status"


# ============================================
# EXPENSE, SAVING, DEBT ADMIN



@admin.register(Saving, site=admin.site)
class SavingAdmin(admin.ModelAdmin):
    list_display = ('description', 'amount', 'date', 'category')
    list_filter = ('category', 'date')
    search_fields = ('description',)


@admin.register(Debt, site=admin.site)
class DebtAdmin(admin.ModelAdmin):
    list_display = ('lender', 'amount', 'interest_rate', 'start_date', 'due_date')
    list_filter = ('start_date', 'due_date')
    search_fields = ('lender',)


# ============================================
# TRAINING ADMIN
# ============================================
class TrainingStepInline(admin.TabularInline):
    model = TrainingStep
    extra = 1
    fields = ['step_number', 'description', 'screenshot_preview']
    readonly_fields = ['screenshot_preview']

    def screenshot_preview(self, obj):
        if obj.screenshot:
            return format_html('<img src="{}" width="100" />', obj.screenshot.url)
        return "No Image"
    screenshot_preview.short_description = "Screenshot"


class TrainingTopicAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'order']
    ordering = ['category', 'order']
    search_fields = ['title']
    inlines = [TrainingStepInline]

    def has_module_permission(self, request):
        return request.user.is_superuser

    def has_view_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_add_permission(self, request):
        return request.user.is_superuser

    def has_change_permission(self, request, obj=None):
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


class TrainingStepAdmin(admin.ModelAdmin):
    list_display = ['topic_name', 'step_number_display', 'short_description', 'has_screenshot']
    ordering = ['topic', 'step_number']
    list_display_links = None

    def topic_name(self, obj):
        return obj.topic.title
    topic_name.short_description = "Topic"
    topic_name.admin_order_field = "topic"

    def step_number_display(self, obj):
        return obj.step_number
    step_number_display.short_description = "Step #"
    step_number_display.admin_order_field = "step_number"

    def short_description(self, obj):
        return obj.description[:60] + "..." if len(obj.description) > 60 else obj.description
    short_description.short_description = "Description"

    def has_screenshot(self, obj):
        return "✅" if obj.screenshot else "❌"
    has_screenshot.short_description = "Screenshot"

    def has_add_permission(self, request):
        return False


# ============================================
# COMPANY INFO ADMIN
# ============================================
class CompanyInfoAdmin(admin.ModelAdmin):
    list_display = ['name', 'contact_number', 'email', 'website', 'updated_at']
    readonly_fields = ['updated_at']

    def has_add_permission(self, request):
        # Sirf tab add karo jab koi record na ho
        if CompanyInfo.objects.exists():
            return False
        return True

    def has_delete_permission(self, request, obj=None):
        # Delete bilkul allow na karo
        return False

    def has_change_permission(self, request, obj=None):
        return True
    
    def has_view_permission(self, request, obj=None):
        return True
        

class StockAuditItemInline(admin.TabularInline):
    model = StockAuditItem
    fields = ('product', 'system_qty', 'physical_qty', 'variance_display', 'variance_value_display', 'unit', 'adjusted', 'notes')
    readonly_fields = ('system_qty', 'variance_display', 'variance_value_display', 'unit')
    extra = 0
    
    def variance_display(self, obj):
        if obj and obj.pk:
            return obj.variance_display()
        return "-"
    variance_display.short_description = "Variance"
    
    def variance_value_display(self, obj):
        if obj and obj.pk:
            return obj.variance_value_display()
        return "-"
    variance_value_display.short_description = "Variance Value"
    
@admin.register(StockAudit, site=admin.site)
class StockAuditAdmin(admin.ModelAdmin):
    list_display = ('audit_no', 'audit_date', 'warehouse', 'status', 'total_items_display', 
                    'variance_items_count', 'total_variance_value_display', 'print_audit')
    list_filter = ('status', 'audit_date', 'warehouse')
    search_fields = ('audit_no', 'warehouse__name')
    inlines = [StockAuditItemInline]
    change_list_template = "admin/button.html"
    actions = ['generate_audit_sheet', 'mark_completed', 'adjust_stock', 'generate_audit_pdf']
    
    fieldsets = (
        ('Audit Information', {
            'fields': ('audit_no', 'warehouse', 'audit_date', 'status')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )
    
    readonly_fields = ('audit_no', 'total_variance_value')
    
    def total_items_display(self, obj):
        return obj.total_items()
    total_items_display.short_description = "Items"
    
    def variance_items_count(self, obj):
        count = obj.items.filter(physical_qty__isnull=False).filter(
            models.Q(variance__gt=0) | models.Q(variance__lt=0)
        ).count()
        if count > 0:
            return format_html('<span style="color: red;">⚠ {} variance(s)</span>', count)
        return "0"
    variance_items_count.short_description = "Variances"
    
    def total_variance_value_display(self, obj):
        val = obj.total_variance_value
        if val < 0:
            return format_html('<span style="color: red;">Rs. ({})</span>', str(round(abs(val), 2)))
        elif val > 0:
            return format_html('<span style="color: green;">Rs. {}</span>', str(round(val, 2)))
        return "Rs. 0.00"
    total_variance_value_display.short_description = "Net Variance"
    
    def print_audit(self, obj):
        url = reverse('admin:generate_audit_pdf', args=[obj.pk])
        return format_html('<a href="{}" target="_blank">📄 Print</a>', url)
    print_audit.short_description = "Print"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for instance in instances:
            if instance.pk is None:
                inventory = Inventory.objects.filter(
                    product=instance.product, 
                    warehouse=formset.instance.warehouse
                ).first()
                instance.system_qty = inventory.stock if inventory else 0
            instance.save()
        formset.save_m2m()
    
    def generate_audit_sheet(self, request, queryset):
        """Auto-populate audit with all warehouse products"""
        for audit in queryset.filter(status='draft'):
            inventories = Inventory.objects.filter(warehouse=audit.warehouse)
            for inv in inventories:
                if not StockAuditItem.objects.filter(audit=audit, product=inv.product).exists():
                    StockAuditItem.objects.create(
                        audit=audit,
                        product=inv.product,
                        system_qty=inv.stock,
                        unit=inv.product.unit.name if inv.product.unit else ''
                    )
            audit.status = 'in_progress'
            audit.save()
        self.message_user(request, f"✅ Audit sheets generated with products.")
    generate_audit_sheet.short_description = "📋 Generate Audit Sheet (Add All Products)"
    
    def mark_completed(self, request, queryset):
        """Mark audit as completed and calculate total variance"""
        for audit in queryset.filter(status='in_progress'):
            total_var = audit.items.aggregate(t=Sum('variance_value'))['t'] or Decimal('0.0')
            audit.total_variance_value = total_var
            audit.status = 'completed'
            audit.save()
        self.message_user(request, f"✅ Audits marked as completed.")
    mark_completed.short_description = "✅ Mark as Completed"
    
    def adjust_stock(self, request, queryset):
        """Adjust stock based on audit variance"""
        count = 0
        for audit in queryset.filter(status='completed'):
            with transaction.atomic():
                for item in audit.items.filter(physical_qty__isnull=False, adjusted=False):
                    if item.variance != 0:
                        adj_type = 'increase' if item.variance > 0 else 'decrease'
                        adjustment = StockAdjustment.objects.create(
                            adjustment_type=adj_type,
                            reason=f"Audit {audit.audit_no}: Variance for {item.product.name}",
                            created_by=request.user
                        )
                        StockAdjustmentItem.objects.create(
                            adjustment=adjustment,
                            product=item.product,
                            qty=abs(item.variance),
                            price=item.price
                        )
                        item.adjusted = True
                        item.save()
                        count += 1
        if count:
            self.message_user(request, f"✅ {count} variances adjusted. Stock updated!")
        else:
            self.message_user(request, "No variances to adjust.")
    adjust_stock.short_description = "🔄 Adjust Stock (Fix Variances)"
    
    def generate_audit_pdf(self, request, queryset):
        """Generate PDF for audit with brackets for negative values"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.3*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        for audit in queryset:
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, alignment=TA_CENTER)
            story.append(Paragraph(company_name, title_style))
            story.append(Paragraph("<b>STOCK AUDIT REPORT</b>", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            details = [
                [Paragraph(f"<b>Audit No:</b> {audit.audit_no}"), 
                 Paragraph(f"<b>Date:</b> {audit.audit_date.strftime('%d-%m-%Y')}")],
                [Paragraph(f"<b>Warehouse:</b> {audit.warehouse.name}"), 
                 Paragraph(f"<b>Status:</b> {audit.get_status_display()}")],
            ]
            story.append(Table(details, colWidths=[3.5*inch, 3.5*inch]))
            story.append(Spacer(1, 0.2*inch))
            
            table_data = [['Product', 'System Qty', 'Physical Qty', 'Variance', 'Unit', 'Value']]
            total_var_value = Decimal('0.0')
            
            for item in audit.items.all():
                # Physical qty
                if item.physical_qty is not None:
                    physical = f"{item.physical_qty:,.0f}"
                else:
                    physical = '?'
                
                # ✅ Variance quantity with brackets for negative
                if item.physical_qty is not None:
                    if item.variance < 0:
                        var_str = f"({abs(item.variance):,.0f})"
                    elif item.variance > 0:
                        var_str = f"{item.variance:+,.0f}"
                    else:
                        var_str = "0"
                else:
                    var_str = '-'
                
                # ✅ Variance value with brackets for negative
                if item.physical_qty is not None:
                    if item.variance_value < 0:
                        val_str = f"Rs. ({abs(item.variance_value):,.2f})"
                    elif item.variance_value > 0:
                        val_str = f"Rs. {item.variance_value:,.2f}"
                    else:
                        val_str = "Rs. 0.00"
                else:
                    val_str = '-'
                
                table_data.append([
                    item.product.name[:25],
                    f"{item.system_qty:,.0f}",
                    physical,
                    var_str,
                    item.unit or '-',
                    val_str
                ])
                total_var_value += item.variance_value
            
            # ✅ Total row with brackets for negative
            if total_var_value < 0:
                total_str = f"Rs. ({abs(total_var_value):,.2f})"
            elif total_var_value > 0:
                total_str = f"Rs. {total_var_value:,.2f}"
            else:
                total_str = "Rs. 0.00"
            
            table_data.append(['', '', '', '', 'NET:', total_str])
            
            items_table = Table(table_data, repeatRows=1)
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#333333')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            story.append(items_table)
            
            story.append(Spacer(1, 0.5*inch))
            sign = [
                [Paragraph("<b>Counted by:</b> _________"), Paragraph("<b>Verified by:</b> _________")],
                [Paragraph("Date: _________"), Paragraph("Date: _________")],
            ]
            story.append(Table(sign, colWidths=[3.5*inch, 3.5*inch]))
            story.append(Paragraph("<br/><br/>", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    generate_audit_pdf.short_description = "📄 Generate Audit PDF"


# ============================================
# SALE QUOTATION ITEM INLINE
# ============================================
class SaleQuotationItemInline(admin.TabularInline):
    model = SaleQuotationItem
    fields = ('product', 'qty', 'price', 'total_amt')
    readonly_fields = ('total_amt',)
    extra = 10


# ============================================
# SALE QUOTATION ADMIN
# ============================================
@admin.register(SaleQuotation, site=admin.site)
class SaleQuotationAdmin(admin.ModelAdmin):
    list_display = ('quotation_no', 'quotation_date', 'customer', 'customer_po_number',
                    'valid_until', 'status', 'formatted_total', 'converted_to_order')
    list_filter = ('status', 'quotation_date', 'customer__group')
    search_fields = ('quotation_no', 'customer__name')
    inlines = [SaleQuotationItemInline]
    change_list_template = "admin/button.html"
    actions = ['convert_to_sale_order', 'generate_quotation_pdf']
    
    readonly_fields = ('quotation_no', 'created_at', 'updated_at', 'converted_to_order', 'created_by')
    
    fieldsets = (
        ('Quotation Information', {
            'fields': ('quotation_no', 'customer', 'customer_po_number', 'warehouse', 
                       'quotation_date', 'valid_until', 'status')
        }),
        ('Payment & Terms', {
            'fields': ('discount_value', 'terms_conditions')
        }),
        ('Additional', {
            'fields': ('notes', 'converted_to_order', 'created_at')
        }),
    )
    
    def formatted_total(self, obj):
        return f"Rs. {obj.total_amount():,.2f}"
    formatted_total.short_description = "Total"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def convert_to_sale_order(self, request, queryset):
        """Convert accepted quotations to Sale Order"""
        count = 0
        for quote in queryset.filter(status__in=['accepted'], converted_to_order__isnull=True):
            try:
                with transaction.atomic():
                    order = SaleOrder.objects.create(
                        customer=quote.customer,
                        warehouse=quote.warehouse,
                        status='confirmed',
                        discount_value=quote.discount_value,
                        customer_po_number=quote.customer_po_number,
                        notes=f"Converted from {quote.quotation_no}",
                        created_by=request.user
                    )
                    for item in quote.items.all():
                        SaleOrderItem.objects.create(
                            order=order,
                            product=item.product,
                            qty=item.qty,
                            price=item.price
                        )
                    quote.status = 'converted'
                    quote.converted_to_order = order
                    quote.save()
                    count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        
        if count:
            self.message_user(request, f"✅ {count} quotations converted to Sale Orders!")
    convert_to_sale_order.short_description = "🔄 Convert to Sale Order"
    
    def generate_quotation_pdf(self, request, queryset):
        """Generate PDF for quotations"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        for quote in queryset:
            story.append(Paragraph(f"QUOTATION - {quote.quotation_no}", styles['Heading1']))
            story.append(Paragraph(f"Customer: {quote.customer.name}", styles['Normal']))
            story.append(Paragraph(f"Valid Until: {quote.valid_until}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            data = [['Product', 'Qty', 'Price', 'Total']]
            for item in quote.items.all():
                data.append([item.product.name, str(item.qty), f"{item.price:,.2f}", f"{item.total_amt:,.2f}"])
            
            data.append(['', '', 'Total:', f"{quote.total_amount():,.2f}"])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.5*inch))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    generate_quotation_pdf.short_description = "📄 Generate Quotation PDF"
    
# ============================================
# PURCHASE QUOTATION ITEM INLINE
# ============================================
class PurchaseQuotationItemInline(admin.TabularInline):
    model = PurchaseQuotationItem
    fields = ('product', 'qty', 'price', 'total_amt')
    readonly_fields = ('total_amt',)
    extra = 10


# ============================================
# PURCHASE QUOTATION ADMIN
# ============================================
@admin.register(PurchaseQuotation, site=admin.site)
class PurchaseQuotationAdmin(admin.ModelAdmin):
    list_display = ('quotation_no', 'quotation_date', 'vendor', 'vendor_so_number',
                    'valid_until', 'status', 'formatted_total', 'converted_to_order')
    list_filter = ('status', 'quotation_date', 'vendor__group')
    search_fields = ('quotation_no', 'vendor__name')
    inlines = [PurchaseQuotationItemInline]
    change_list_template = "admin/button.html"
    actions = ['convert_to_purchase_order', 'generate_rfq_pdf']
    
    readonly_fields = ('quotation_no', 'created_at', 'updated_at', 'converted_to_order', 'created_by')
    
    fieldsets = (
        ('Quotation Information', {
            'fields': ('quotation_no', 'vendor', 'vendor_so_number', 'warehouse',
                       'quotation_date', 'valid_until', 'status')
        }),
        ('Vendor Details', {
            'fields': ('vendor_reference',)
        }),
        ('Payment & Terms', {
            'fields': ('discount_value', 'terms_conditions')
        }),
        ('Additional', {
            'fields': ('notes', 'converted_to_order', 'created_at')
        }),
    )
    
    def formatted_total(self, obj):
        return f"Rs. {obj.total_amount():,.2f}"
    formatted_total.short_description = "Total"
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    def convert_to_purchase_order(self, request, queryset):
        """Convert accepted quotations to Purchase Order"""
        count = 0
        for quote in queryset.filter(status__in=['accepted'], converted_to_order__isnull=True):
            try:
                with transaction.atomic():
                    order = PurchaseOrder.objects.create(
                        vendor=quote.vendor,
                        warehouse=quote.warehouse,
                        status='confirmed',
                        discount_value=quote.discount_value,
                        vendor_so_number=quote.vendor_so_number,
                        notes=f"Converted from {quote.quotation_no}",
                        created_by=request.user
                    )
                    for item in quote.items.all():
                        PurchaseOrderItem.objects.create(
                            order=order,
                            product=item.product,
                            qty=item.qty,
                            price=item.price
                        )
                    quote.status = 'converted'
                    quote.converted_to_order = order
                    quote.save()
                    count += 1
            except Exception as e:
                self.message_user(request, f"Error: {e}", level=messages.ERROR)
        
        if count:
            self.message_user(request, f"✅ {count} RFQs converted to Purchase Orders!")
    convert_to_purchase_order.short_description = "🔄 Convert to Purchase Order"
    
    def generate_rfq_pdf(self, request, queryset):
        """Generate PDF for RFQs"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        company = CompanyInfo.objects.first()
        company_name = company.name if company else "Company Name"
        
        for quote in queryset:
            story.append(Paragraph(f"REQUEST FOR QUOTATION - {quote.quotation_no}", styles['Heading1']))
            story.append(Paragraph(f"Vendor: {quote.vendor.name}", styles['Normal']))
            story.append(Paragraph(f"Valid Until: {quote.valid_until}", styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
            
            data = [['Product', 'Qty', 'Price', 'Total']]
            for item in quote.items.all():
                data.append([item.product.name, str(item.qty), f"{item.price:,.2f}", f"{item.total_amt:,.2f}"])
            
            data.append(['', '', 'Total:', f"{quote.total_amount():,.2f}"])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.5*inch))
        
        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')
    generate_rfq_pdf.short_description = "📄 Generate RFQ PDF"
    
# admin.py mein yeh exact code copy karo

@admin.register(SalesTarget)
class SalesTargetAdmin(admin.ModelAdmin):
    list_display = ('id', 'target_type', 'target_amount', 'start_date', 'end_date', 
                    'current_progress_display', 'achieved_display', 'is_active', 'status_badge')
    list_filter = ('target_type', 'is_active', 'start_date')
    search_fields = ('salesman__username', 'product__name')
    list_editable = ('is_active',)  # ✅ Ab yeh sahi hai kyunke is_active list_display mein hai
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('target_type', 'target_amount', 'start_date', 'end_date')
        }),
        ('Target Specific', {
            'fields': ('salesman', 'product'),
            'classes': ('collapse',)
        }),
        ('Bonus Settings', {
            'fields': ('bonus_percentage', 'bonus_amount'),
            'description': 'Set bonus percentage or fixed amount for achieving target'
        }),
        ('Status', {
            'fields': ('is_active', 'created_by')
        }),
    )
    
    def current_progress_display(self, obj):
        progress = obj.current_progress()
        if progress >= 100:
            return format_html('<span style="color: #28a745; font-weight: bold;">✅ {}%</span>', round(progress, 1))
        elif progress >= 80:
            return format_html('<span style="color: #ff9800; font-weight: bold;">🟡 {}%</span>', round(progress, 1))
        else:
            return format_html('<span style="color: #dc3545; font-weight: bold;">🔴 {}%</span>', round(progress, 1))
    current_progress_display.short_description = 'Progress'
    
    def achieved_display(self, obj):
        achieved = obj.achieved_amount()
        target = obj.target_amount
        return format_html('<strong>Rs. {:,}</strong> / Rs. {:,}', achieved, target)
    achieved_display.short_description = 'Achieved / Target'
    
    def status_badge(self, obj):
        if obj.is_active:
            return format_html('<span style="background: #28a745; color: white; padding: 2px 8px; border-radius: 12px;">Active</span>')
        return format_html('<span style="background: #6c757d; color: white; padding: 2px 8px; border-radius: 12px;">Inactive</span>')
    status_badge.short_description = 'Status'
    
    def save_model(self, request, obj, form, change):
        if not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)
    
    actions = ['send_target_report']
    
    def send_target_report(self, request, queryset):
        """Send target report via WhatsApp"""
        from .whatsapp_utils import WhatsAppSender
        
        count = 0
        for target in queryset:
            if target.salesman and hasattr(target.salesman, 'profile') and target.salesman.profile.phone:
                message = f"🎯 *TARGET REPORT*\n\n"
                message += f"Target: Rs. {target.target_amount:,.0f}\n"
                message += f"Achieved: Rs. {target.achieved_amount():,.0f}\n"
                message += f"Progress: {target.current_progress():.0f}%\n"
                message += f"Days Left: {target.days_remaining()}\n"
                message += f"Daily Needed: Rs. {target.daily_needed():,.0f}"
                
                # WhatsAppSender.send_message(target.salesman.profile.phone, message)
                count += 1
        
        self.message_user(request, f"Report sent to {count} salesmen")
    send_target_report.short_description = "📱 Send Target Report via WhatsApp"

# ============================================
# REGISTER ALL MODELS
# ============================================
admin.site.register(CompanyInfo, CompanyInfoAdmin)
admin.site.register(Brand)
admin.site.register(Category)
admin.site.register(Types)
admin.site.register(Location)
admin.site.register(Unit)
admin.site.register(Purchase, PurchaseAdmin)
admin.site.register(Customer, CustomerAdmin)
admin.site.register(Sale, SaleAdmin)
admin.site.register(Inventory, InventoryAdmin)
admin.site.register(StockBatch, StockBatchAdmin)
admin.site.register(PurchaseItem, PurchaseItemAdmin)
admin.site.register(SaleItem, SaleItemAdmin)
admin.site.register(MonthlyClosing, MonthlyClosingAdmin)
admin.site.register(PurchaseRetrn, PurchaseRetrnAdmin)
admin.site.register(PurchaseRetrnItem, PurchaseRetrnItemAdmin)
admin.site.register(SaleRetrn, SaleRetrnAdmin)
admin.site.register(SaleRetrnItem, SaleRetrnItemAdmin)
admin.site.register(LogEntry, LogEntryAdmin)
admin.site.register(License, LicenseAdmin)
admin.site.register(User, UserAdmin)
admin.site.register(Group, GroupAdmin)
admin.site.register(TrainingStep, TrainingStepAdmin)

admin.site.register(InstallmentPlan, InstallmentPlanAdmin)

admin.site.register(SaleInstallment, SaleInstallmentAdmin)

admin.site.register(TrainingTopic, TrainingTopicAdmin)

admin.site.register(EmiPayment, EmiPaymentAdmin)


# ============================================
# RANGE REPORT ADMIN
# ============================================

class SaleRangeReportAdmin(admin.ModelAdmin):
    list_display = ('name', 'from_date', 'to_date', 'report_type', 'generated_at')
    list_filter = ('report_type',)
    search_fields = ('name',)
    readonly_fields = ('generated_at', 'generated_by')
    
    def save_model(self, request, obj, form, change):
        if not obj.generated_by:
            obj.generated_by = request.user
        super().save_model(request, obj, form, change)

admin.site.register(SaleRangeReport, SaleRangeReportAdmin)

@admin.register(Share, site=admin.site)
class ShareAdmin(admin.ModelAdmin):
    list_display = ('id', 'shareholder', 'share_type', 'quantity', 'purchase_price', 
                    'certificate_number', 'certificate_printed', 'issue_date')
    list_filter = ('share_type', 'is_locked', 'certificate_printed', 'issue_date')
    search_fields = ('shareholder__name', 'shareholder__shareholder_code', 'certificate_number')
    readonly_fields = ('created_at', 'updated_at', 'certificate_printed_at', 'certificate_printed_by')
    
    fieldsets = (
        ('Share Information', {
            'fields': ('shareholder', 'share_type', 'quantity', 'purchase_price')
        }),
        ('Certificate Information', {
            'fields': ('certificate_number', 'certificate_issue_date', 'certificate_printed', 
                       'certificate_printed_at', 'certificate_printed_by', 'certificate_template')
        }),
        ('Issue Details', {
            'fields': ('issue_date', 'is_locked')
        }),
        ('Transfer Tracking', {
            'fields': ('transferred_from', 'transfer_date', 'transfer_notes'),
            'classes': ('collapse',)
        }),
        ('Notes', {
            'fields': ('notes', 'certificate_notes')
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['generate_certificate', 'print_selected_certificates']
    
    def generate_certificate(self, request, queryset):
        """Generate certificate for selected shares"""
        count = 0
        for share in queryset:
            if not share.certificate_number:
                share.certificate_number = f"CERT-{share.id:06d}"
                share.certificate_issue_date = now().date()
                share.certificate_printed = True
                share.certificate_printed_at = now()
                share.certificate_printed_by = request.user
                share.save()
                count += 1
        self.message_user(request, f"✅ {count} certificates generated!")
    generate_certificate.short_description = "📄 Generate Certificates"
    
    def print_selected_certificates(self, request, queryset):
        """Print certificates for selected shares"""
        from .certificate_utils import generate_bulk_certificates
        share_ids = list(queryset.values_list('id', flat=True))
        zip_buffer = generate_bulk_certificates(share_ids)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="share_certificates.zip"'
        return response
    print_selected_certificates.short_description = "🖨️ Print Selected Certificates"

# ============================================
# DRIP ADMIN REGISTRATIONS
# ============================================

@admin.register(DividendReinvestmentPlan)
class DividendReinvestmentPlanAdmin(admin.ModelAdmin):
    list_display = ('dividend', 'plan_name', 'discount_type', 'discount_value', 'status', 'is_auto_enroll')
    list_filter = ('status', 'discount_type', 'is_auto_enroll')
    search_fields = ('plan_name', 'dividend__dividend_no')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('dividend', 'plan_name', 'status')
        }),
        ('Pricing Settings', {
            'fields': ('discount_type', 'discount_value', 'admin_fee')
        }),
        ('Share Limits', {
            'fields': ('min_shares', 'max_shares', 'fractional_shares_allowed', 'round_down_to_nearest')
        }),
        ('Dates', {
            'fields': ('start_date', 'end_date')
        }),
        ('Enrollment', {
            'fields': ('is_auto_enroll', 'is_default')
        }),
        ('System Fields', {
            'fields': ('created_by', 'created_at', 'updated_at', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['activate', 'deactivate', 'auto_enroll']
    
    def activate(self, request, queryset):
        queryset.update(status='active')
        self.message_user(request, f"✅ {queryset.count()} DRIP plans activated")
    activate.short_description = "Activate Selected Plans"
    
    def deactivate(self, request, queryset):
        queryset.update(status='inactive')
        self.message_user(request, f"⚠️ {queryset.count()} DRIP plans deactivated")
    deactivate.short_description = "Deactivate Selected Plans"
    
    def auto_enroll(self, request, queryset):
        count = 0
        for drip in queryset:
            if drip.is_auto_enroll:
                shareholders = Shareholder.objects.filter(status='active')
                for shareholder in shareholders:
                    ShareholderDRIPEnrollment.objects.get_or_create(
                        shareholder=shareholder,
                        drip=drip,
                        defaults={'status': 'active', 'created_by': request.user}
                    )
                    count += 1
        self.message_user(request, f"✅ {count} shareholders auto-enrolled")
    auto_enroll.short_description = "Auto-Enroll All Shareholders"


@admin.register(ShareholderDRIPEnrollment)
class ShareholderDRIPEnrollmentAdmin(admin.ModelAdmin):
    list_display = ('shareholder', 'drip', 'status', 'enrollment_date', 'auto_reinvest')
    list_filter = ('status', 'auto_reinvest', 'enrollment_date')
    search_fields = ('shareholder__name', 'drip__plan_name')
    readonly_fields = ('created_at', 'updated_at')
    
    actions = ['activate_enrollments', 'deactivate_enrollments']
    
    def activate_enrollments(self, request, queryset):
        queryset.update(status='active')
        self.message_user(request, f"✅ {queryset.count()} enrollments activated")
    activate_enrollments.short_description = "Activate Selected"
    
    def deactivate_enrollments(self, request, queryset):
        queryset.update(status='inactive')
        self.message_user(request, f"⚠️ {queryset.count()} enrollments deactivated")
    deactivate_enrollments.short_description = "Deactivate Selected"


@admin.register(DRIPTransaction)
class DRIPTransactionAdmin(admin.ModelAdmin):
    list_display = ('reference_no', 'dividend_payment', 'shares_purchased', 'purchase_price', 'total_cost', 'status')
    list_filter = ('status', 'processed_at')
    search_fields = ('reference_no', 'dividend_payment__shareholder__name')
    readonly_fields = ('reference_no', 'created_at')
    
# admin.py - Add these

@admin.register(ShareBuyback)
class ShareBuybackAdmin(admin.ModelAdmin):
    list_display = ('buyback_no', 'buyback_type', 'total_amount', 'shares_bought', 'progress_percent', 'status')
    list_filter = ('status', 'buyback_type')
    search_fields = ('buyback_no', 'description')
    readonly_fields = ('buyback_no', 'created_at', 'updated_at')
    
    actions = ['approve_buyback', 'complete_buyback']
    
    def approve_buyback(self, request, queryset):
        queryset.update(status='approved', approved_by=request.user, approved_at=now())
        self.message_user(request, f"✅ {queryset.count()} buybacks approved")
    approve_buyback.short_description = "Approve Selected Buybacks"
    
    def complete_buyback(self, request, queryset):
        queryset.update(status='completed', completed_at=now())
        self.message_user(request, f"✅ {queryset.count()} buybacks completed")
    complete_buyback.short_description = "Complete Selected Buybacks"


@admin.register(ShareholderDiscountProgram)
class ShareholderDiscountProgramAdmin(admin.ModelAdmin):
    list_display = ('name', 'discount_value', 'min_shares_required', 'is_active', 'priority')
    list_filter = ('is_active', 'discount_type')
    search_fields = ('name',)
    filter_horizontal = ('products', 'categories')


@admin.register(ShareholderLoan)
class ShareholderLoanAdmin(admin.ModelAdmin):
    list_display = ('loan_no', 'shareholder', 'principal', 'outstanding', 'status', 'get_ltv_percent')
    list_filter = ('status', 'loan_type')
    search_fields = ('loan_no', 'shareholder__name')
    readonly_fields = ('loan_no', 'created_at', 'updated_at')
    
    actions = ['approve_loans', 'disburse_loans']
    
    def approve_loans(self, request, queryset):
        queryset.update(status='approved', approved_by=request.user, approved_at=now())
        self.message_user(request, f"✅ {queryset.count()} loans approved")
    approve_loans.short_description = "Approve Selected Loans"
    
    def disburse_loans(self, request, queryset):
        queryset.update(status='active', disbursement_date=now().date())
        self.message_user(request, f"✅ {queryset.count()} loans disbursed")
    disburse_loans.short_description = "Disburse Selected Loans"
    
# ============================================
# SERVICE ADMIN REGISTRATIONS
# ============================================

from .models import (
    ServiceCategory, Service, ServiceProductRequirement,
    ServiceRequest, ServiceProductUsed, ServiceAppointment,
    ServiceTechnicianAssignment, ServiceFeedback
)

class ServiceProductRequirementInline(admin.TabularInline):
    model = ServiceProductRequirement
    fields = ('product', 'quantity', 'is_optional')
    extra = 1
    raw_id_fields = ('product',)  # ✅ FIXED: Using raw_id_fields
    # OR autocomplete_fields = ('product',) if ProductAdmin has search_fields


class ServiceProductUsedInline(admin.TabularInline):
    model = ServiceProductUsed
    fields = ('product', 'quantity', 'price', 'total')
    readonly_fields = ('total',)
    extra = 1
    raw_id_fields = ('product',)  # ✅ FIXED: Using raw_id_fields
    # OR autocomplete_fields = ('product',) if ProductAdmin has search_fields


@admin.register(ServiceCategory)
class ServiceCategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_active', 'created_at')
    list_filter = ('is_active',)
    search_fields = ('name',)
    list_editable = ('is_active',)


@admin.register(Service)
class ServiceAdmin(admin.ModelAdmin):
    list_display = ('service_code', 'name', 'service_type', 'price', 'pricing_type', 'is_active')
    list_filter = ('service_type', 'pricing_type', 'is_active', 'category')
    search_fields = ('service_code', 'name', 'description')
    list_editable = ('price', 'is_active')
    inlines = [ServiceProductRequirementInline]
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('service_code', 'name', 'service_type', 'category', 'description')
        }),
        ('Pricing', {
            'fields': ('pricing_type', 'price', 'estimated_hours', 'min_charge', 'max_charge')
        }),
        ('Staff & Settings', {
            'fields': ('requires_technician', 'default_technician', 'needs_appointment', 'warranty_months')
        }),
        ('Status', {
            'fields': ('is_active',)
        }),
        ('System', {
            'fields': ('created_by', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    actions = ['activate', 'deactivate']
    
    def activate(self, request, queryset):
        queryset.update(is_active=True)
        self.message_user(request, f"✅ {queryset.count()} services activated")
    activate.short_description = "Activate Selected Services"
    
    def deactivate(self, request, queryset):
        queryset.update(is_active=False)
        self.message_user(request, f"⚠️ {queryset.count()} services deactivated")
    deactivate.short_description = "Deactivate Selected Services"


class ServiceAppointmentInline(admin.TabularInline):
    model = ServiceAppointment
    fields = ('technician', 'appointment_date', 'start_time', 'end_time', 'status')
    extra = 1


@admin.register(ServiceRequest)
class ServiceRequestAdmin(admin.ModelAdmin):
    list_display = ('request_no', 'customer', 'service', 'status', 'priority', 
                    'scheduled_date', 'technician', 'total_amount_display', 'is_overdue_display')
    list_filter = ('status', 'priority', 'scheduled_date', 'service__service_type')
    search_fields = ('request_no', 'customer__name', 'customer__customer_code', 'service__name')
    list_editable = ('status', 'priority')
    inlines = [ServiceProductUsedInline, ServiceAppointmentInline]
    
    fieldsets = (
        ('Request Information', {
            'fields': ('request_no', 'customer', 'service', 'description', 'priority', 'status')
        }),
        ('Scheduling', {
            'fields': ('requested_date', 'scheduled_date', 'appointment_time', 'estimated_hours', 'completed_date')
        }),
        ('Location & Contact', {
            'fields': ('service_address', 'contact_person', 'contact_phone')
        }),
        ('Pricing', {
            'fields': ('quoted_price', 'actual_price', 'discount')
        }),
        ('Technician', {
            'fields': ('technician', 'assigned_at')
        }),
        ('Follow-up & Invoice', {
            'fields': ('follow_up_required', 'follow_up_date', 'invoice')
        }),
        ('Feedback', {
            'fields': ('feedback_rating', 'feedback_comment')
        }),
        ('System', {
            'fields': ('created_by', 'created_at', 'updated_at', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('request_no', 'assigned_at', 'created_at', 'updated_at')
    
    def total_amount_display(self, obj):
        return f"Rs. {obj.total_amount():,.2f}"
    total_amount_display.short_description = "Total Amount"
    
    def is_overdue_display(self, obj):
        if obj.is_overdue():
            days = obj.days_overdue()
            return format_html('<span style="color: red;">⚠️ {} days overdue</span>', days)
        return format_html('<span style="color: green;">✅ On time</span>')
    is_overdue_display.short_description = "Overdue"
    
    actions = ['mark_assigned', 'mark_in_progress', 'mark_completed', 'generate_invoice']
    
    def mark_assigned(self, request, queryset):
        count = queryset.update(status='assigned', assigned_at=now())
        self.message_user(request, f"✅ {count} requests marked as assigned")
    mark_assigned.short_description = "👤 Mark as Assigned"
    
    def mark_in_progress(self, request, queryset):
        count = queryset.update(status='in_progress')
        self.message_user(request, f"⚙️ {count} requests marked as in progress")
    mark_in_progress.short_description = "⚙️ Mark as In Progress"
    
    def mark_completed(self, request, queryset):
        count = queryset.update(status='completed', completed_date=now().date())
        self.message_user(request, f"✅ {count} requests marked as completed")
    mark_completed.short_description = "✅ Mark as Completed"
    
    def generate_invoice(self, request, queryset):
        """Generate invoice for completed services"""
        count = 0
        for service_request in queryset.filter(status='completed', invoice__isnull=True):
            try:
                with transaction.atomic():
                    sale = Sale.objects.create(
                        customer=service_request.customer,
                        bill_no=f"SVC-{service_request.request_no}",
                        sale_date=now(),
                        paid=0,
                        discount_value=service_request.discount,
                        created_by=request.user,
                        warehouse=Warehouse.objects.first()
                    )
                    
                    SaleItem.objects.create(
                        sale=sale,
                        product=None,
                        qty=1,
                        price=service_request.total_amount(),
                        total_amt=service_request.total_amount(),
                        profit=0,
                        description=f"Service: {service_request.service.name} - {service_request.request_no}"
                    )
                    
                    service_request.invoice = sale
                    service_request.save()
                    count += 1
                    
            except Exception as e:
                messages.error(request, f"❌ Error for {service_request.request_no}: {str(e)}")
        
        if count:
            self.message_user(request, f"✅ {count} invoices generated!")
    generate_invoice.short_description = "📄 Generate Invoice"


@admin.register(ServiceAppointment)
class ServiceAppointmentAdmin(admin.ModelAdmin):
    list_display = ('service_request', 'technician', 'appointment_date', 'start_time', 'status')
    list_filter = ('status', 'appointment_date', 'technician')
    search_fields = ('service_request__request_no', 'technician__name')
    list_editable = ('status',)


@admin.register(ServiceTechnicianAssignment)
class ServiceTechnicianAssignmentAdmin(admin.ModelAdmin):
    list_display = ('technician', 'date', 'max_services', 'current_services', 'is_available', 'is_full_display')
    list_filter = ('date', 'is_available')
    search_fields = ('technician__name',)
    list_editable = ('is_available',)
    
    def is_full_display(self, obj):
        return "🔴 Full" if obj.is_full() else "🟢 Available"
    is_full_display.short_description = "Status"


@admin.register(ServiceFeedback)
class ServiceFeedbackAdmin(admin.ModelAdmin):
    list_display = ('service_request', 'rating', 'average_rating_display', 'would_recommend', 'created_at')
    list_filter = ('rating', 'would_recommend')
    search_fields = ('service_request__request_no', 'comment')
    readonly_fields = ('service_request', 'created_at')
    
    def average_rating_display(self, obj):
        return f"{obj.average_rating()}★"
    average_rating_display.short_description = "Avg Rating"
    
# ============================================
# SERVICE INVENTORY ADMIN
# ============================================

from .models import (
    ServiceInventoryCategory, ServiceInventory, ServiceInventoryTransaction,
    ServicePartUsage, ServiceInventoryPurchaseOrder, ServiceInventoryPODetail,
    ServiceInventoryStockAdjustment
)


@admin.register(ServiceInventoryCategory)
class ServiceInventoryCategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_active', 'created_at')
    list_filter = ('is_active',)
    search_fields = ('name', 'description')
    list_editable = ('is_active',)


class ServiceInventoryTransactionInline(admin.TabularInline):
    model = ServiceInventoryTransaction
    fields = ('transaction_type', 'quantity', 'balance_after', 'notes', 'created_at')
    readonly_fields = ('created_at',)
    extra = 0
    can_delete = False


@admin.register(ServiceInventory)
class ServiceInventoryAdmin(admin.ModelAdmin):
    list_display = ('item_code', 'name', 'item_type', 'category', 'current_stock', 
                    'reorder_level', 'unit_cost', 'selling_price', 'is_low_stock_display', 'is_active')
    list_filter = ('item_type', 'category', 'is_active', 'is_consumable')
    search_fields = ('item_code', 'name', 'description', 'barcode')
    list_editable = ('current_stock', 'reorder_level', 'unit_cost', 'selling_price', 'is_active')
    inlines = [ServiceInventoryTransactionInline]
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('item_code', 'name', 'description', 'item_type', 'category', 'product')
        }),
        ('Stock Management', {
            'fields': ('current_stock', 'min_stock', 'max_stock', 'reorder_level', 'reorder_quantity')
        }),
        ('Pricing', {
            'fields': ('unit_cost', 'selling_price', 'markup_percent')
        }),
        ('Location & Supplier', {
            'fields': ('unit', 'warehouse', 'shelf_location', 'preferred_supplier')
        }),
        ('Barcode & Settings', {
            'fields': ('barcode', 'is_active', 'is_consumable', 'is_serialized')
        }),
        ('System', {
            'fields': ('created_by', 'created_at', 'last_updated', 'updated_by'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('item_code', 'created_at', 'last_updated')
    
    def is_low_stock_display(self, obj):
        if obj.is_critical_stock():
            return format_html('<span style="color: red; font-weight: bold;">⚠️ Critical</span>')
        elif obj.is_low_stock():
            return format_html('<span style="color: orange; font-weight: bold;">⚠️ Low</span>')
        return format_html('<span style="color: green;">✅ OK</span>')
    is_low_stock_display.short_description = "Stock Status"
    
    actions = ['mark_active', 'mark_inactive', 'set_reorder_level']
    
    def mark_active(self, request, queryset):
        queryset.update(is_active=True)
        self.message_user(request, f"✅ {queryset.count()} items activated")
    mark_active.short_description = "Activate Selected"
    
    def mark_inactive(self, request, queryset):
        queryset.update(is_active=False)
        self.message_user(request, f"⚠️ {queryset.count()} items deactivated")
    mark_inactive.short_description = "Deactivate Selected"
    
    def set_reorder_level(self, request, queryset):
        # Custom action to set reorder level
        level = request.POST.get('reorder_level')
        if level:
            queryset.update(reorder_level=level)
            self.message_user(request, f"✅ Reorder level set to {level}")
    set_reorder_level.short_description = "Set Reorder Level"


@admin.register(ServiceInventoryTransaction)
class ServiceInventoryTransactionAdmin(admin.ModelAdmin):
    list_display = ('inventory_item', 'transaction_type', 'quantity', 'balance_after', 'created_at', 'created_by')
    list_filter = ('transaction_type', 'created_at')
    search_fields = ('inventory_item__name', 'inventory_item__item_code', 'notes')
    readonly_fields = ('created_at',)
    
    def has_add_permission(self, request):
        return False


@admin.register(ServicePartUsage)
class ServicePartUsageAdmin(admin.ModelAdmin):
    list_display = ('service_request', 'inventory_item', 'quantity', 'total_cost', 'total_charged', 'used_at')
    list_filter = ('used_at', 'charged_to_customer')
    search_fields = ('service_request__request_no', 'inventory_item__name')
    readonly_fields = ('total_cost', 'total_charged', 'used_at')
    
    def has_add_permission(self, request):
        return False


class ServiceInventoryPODetailInline(admin.TabularInline):
    model = ServiceInventoryPODetail
    fields = ('inventory_item', 'quantity', 'unit_price', 'total_price', 'received_quantity')
    readonly_fields = ('total_price',)
    extra = 1


@admin.register(ServiceInventoryPurchaseOrder)
class ServiceInventoryPurchaseOrderAdmin(admin.ModelAdmin):
    list_display = ('po_no', 'supplier', 'order_date', 'status', 'total_amount', 'net_amount')
    list_filter = ('status', 'order_date')
    search_fields = ('po_no', 'supplier__name')
    inlines = [ServiceInventoryPODetailInline]
    
    fieldsets = (
        ('Order Information', {
            'fields': ('po_no', 'supplier', 'status')
        }),
        ('Dates', {
            'fields': ('order_date', 'expected_delivery', 'received_date')
        }),
        ('Amounts', {
            'fields': ('total_amount', 'discount', 'net_amount')
        }),
        ('Notes', {
            'fields': ('notes',)
        }),
    )
    
    readonly_fields = ('po_no', 'order_date', 'total_amount', 'net_amount')
    
    actions = ['mark_received', 'mark_approved', 'mark_ordered']
    
    def mark_received(self, request, queryset):
        for po in queryset.filter(status__in=['ordered', 'partial']):
            po.mark_received()
        self.message_user(request, f"✅ {queryset.count()} POs marked as received")
    mark_received.short_description = "📥 Mark as Received"
    
    def mark_approved(self, request, queryset):
        queryset.update(status='approved')
        self.message_user(request, f"✅ {queryset.count()} POs approved")
    mark_approved.short_description = "✅ Mark as Approved"
    
    def mark_ordered(self, request, queryset):
        queryset.update(status='ordered')
        self.message_user(request, f"📦 {queryset.count()} POs marked as ordered")
    mark_ordered.short_description = "📦 Mark as Ordered"


@admin.register(ServiceInventoryStockAdjustment)
class ServiceInventoryStockAdjustmentAdmin(admin.ModelAdmin):
    list_display = ('adjustment_no', 'inventory_item', 'adjustment_type', 'quantity', 
                    'previous_stock', 'new_stock', 'performed_at', 'performed_by')
    list_filter = ('adjustment_type', 'performed_at')
    search_fields = ('inventory_item__name', 'reason')
    readonly_fields = ('adjustment_no', 'previous_stock', 'new_stock', 'performed_at')
    
    def has_change_permission(self, request, obj=None):
        return False
        

from django.contrib import admin
from django.utils.html import format_html
from django.utils.timezone import now
from django.db.models import Sum

from .models import (
    Department, Project,
    ExpenseCategory, ExpenseBudget, Expense, ExpenseApprovalHistory,
    ExpenseClaim, ExpenseForecast
)


# ============================================
# DEPARTMENT & PROJECT ADMIN
# ============================================

@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'description', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('name', 'code', 'description')
    list_editable = ('is_active',)


@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'department', 'status', 'budget', 'total_expenses_display', 'is_active')
    list_filter = ('status', 'department', 'is_active')
    search_fields = ('name', 'code', 'description')
    list_editable = ('status', 'is_active')
    
    def total_expenses_display(self, obj):
        return f"Rs. {obj.total_expenses():,.2f}"
    total_expenses_display.short_description = "Total Expenses"


# ============================================
# EXPENSE CATEGORY ADMIN
# ============================================

@admin.register(ExpenseCategory)
class ExpenseCategoryAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'category_type', 'parent', 'is_active')
    list_filter = ('category_type', 'is_active')
    search_fields = ('name', 'code', 'description')
    list_editable = ('is_active',)
    readonly_fields = ('code',)


# ============================================
# EXPENSE BUDGET ADMIN
# ============================================

@admin.register(ExpenseBudget)
class ExpenseBudgetAdmin(admin.ModelAdmin):
    list_display = ('name', 'category', 'department', 'project', 'budget_period', 
                    'period_start', 'period_end', 'allocated_amount', 'used_amount', 
                    'remaining_amount', 'status')
    list_filter = ('status', 'budget_period', 'category', 'department')
    search_fields = ('name', 'category__name', 'department__name')
    list_editable = ('status',)
    readonly_fields = ('used_amount', 'remaining_amount')
    
    actions = ['activate_budgets', 'approve_budgets', 'expire_budgets']
    
    def activate_budgets(self, request, queryset):
        count = queryset.update(status='active')
        self.message_user(request, f"✅ {count} budgets activated")
    activate_budgets.short_description = "✅ Activate Selected Budgets"
    
    def approve_budgets(self, request, queryset):
        count = queryset.update(status='approved', approved_by=request.user, approved_at=now())
        self.message_user(request, f"✅ {count} budgets approved")
    approve_budgets.short_description = "✅ Approve Selected Budgets"
    
    def expire_budgets(self, request, queryset):
        count = queryset.update(status='expired')
        self.message_user(request, f"⏰ {count} budgets expired")
    expire_budgets.short_description = "⏰ Expire Selected Budgets"


# ============================================
# EXPENSE APPROVAL HISTORY INLINE
# ============================================

class ExpenseApprovalHistoryInline(admin.TabularInline):
    model = ExpenseApprovalHistory
    fields = ('action', 'performed_by', 'remarks', 'performed_at')
    readonly_fields = ('performed_at',)
    extra = 0
    can_delete = False


# ============================================
# ✅ EXPENSE ADMIN - COMPLETELY FIXED
# ============================================

@admin.register(Expense)
class ExpenseAdmin(admin.ModelAdmin):
    # ✅ NO 'date' FIELD ANYWHERE
    list_display = (
        'expense_no',
        'description',
        'category',
        'amount',
        'expense_date',
        'status_display',
        'payment_method',
        'created_by'
    )
    
    list_filter = (
        'status',
        'category',
        'payment_method',
        'expense_date',
        'department'
    )
    
    search_fields = ('expense_no', 'description', 'reference_no', 'vendor__name')
    
    # ✅ NO list_editable - removed completely
    # list_editable = ()  # Empty
    
    inlines = [ExpenseApprovalHistoryInline]
    readonly_fields = ('expense_no', 'total_amount', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Expense Information', {
            'fields': ('expense_no', 'category', 'description', 'amount', 'expense_date')
        }),
        ('Payment Details', {
            'fields': ('payment_method', 'reference_no', 'vendor', 'vendor_invoice_no')
        }),
        ('Tax', {
            'fields': ('tax_rate', 'tax_amount', 'total_amount')
        }),
        ('Budget & Cost Center', {
            'fields': ('budget', 'department', 'project', 'cost_center')
        }),
        ('Approval', {
            'fields': ('status', 'submitted_by', 'submitted_at', 'approved_by', 
                      'approved_at', 'rejection_reason')
        }),
        ('Payment', {
            'fields': ('paid_by', 'paid_at', 'payment_date')
        }),
        ('Documents', {
            'fields': ('receipt', 'attachment')
        }),
        ('Recurring', {
            'fields': ('is_recurring', 'recurring_frequency', 'recurring_end_date'),
            'classes': ('collapse',)
        }),
        ('System', {
            'fields': ('created_by', 'created_at', 'updated_at', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    def status_display(self, obj):
        return format_html(obj.get_status_badge())
    status_display.short_description = "Status"
    
    actions = ['submit_expenses', 'approve_expenses', 'reject_expenses', 'mark_paid']
    
    def submit_expenses(self, request, queryset):
        count = 0
        for expense in queryset.filter(status='draft'):
            expense.submit(request.user)
            count += 1
        self.message_user(request, f"📤 {count} expenses submitted")
    submit_expenses.short_description = "📤 Submit Selected Expenses"
    
    def approve_expenses(self, request, queryset):
        count = 0
        for expense in queryset.filter(status='submitted'):
            expense.approve(request.user)
            count += 1
        self.message_user(request, f"✅ {count} expenses approved")
    approve_expenses.short_description = "✅ Approve Selected Expenses"
    
    def reject_expenses(self, request, queryset):
        count = 0
        for expense in queryset.filter(status__in=['submitted', 'review']):
            expense.reject(request.user, "Rejected via bulk action")
            count += 1
        self.message_user(request, f"❌ {count} expenses rejected")
    reject_expenses.short_description = "❌ Reject Selected Expenses"
    
    def mark_paid(self, request, queryset):
        count = 0
        for expense in queryset.filter(status='approved'):
            expense.mark_paid(request.user)
            count += 1
        self.message_user(request, f"💰 {count} expenses marked as paid")
    mark_paid.short_description = "💰 Mark as Paid"


# ============================================
# EXPENSE CLAIM ADMIN
# ============================================

@admin.register(ExpenseClaim)
class ExpenseClaimAdmin(admin.ModelAdmin):
    list_display = ('claim_no', 'employee', 'title', 'total_amount', 'claim_date', 'status')
    list_filter = ('status', 'claim_date', 'employee')
    search_fields = ('claim_no', 'employee__name', 'title')
    list_editable = ('status',)
    filter_horizontal = ('expenses',)
    readonly_fields = ('claim_no', 'total_amount')
    
    actions = ['approve_claims', 'reject_claims', 'mark_reimbursed']
    
    def approve_claims(self, request, queryset):
        count = queryset.update(status='approved', approved_by=request.user, approved_at=now())
        self.message_user(request, f"✅ {count} claims approved")
    approve_claims.short_description = "✅ Approve Selected Claims"
    
    def reject_claims(self, request, queryset):
        count = queryset.update(status='rejected', approved_by=request.user, approved_at=now())
        self.message_user(request, f"❌ {count} claims rejected")
    reject_claims.short_description = "❌ Reject Selected Claims"
    
    def mark_reimbursed(self, request, queryset):
        count = queryset.update(status='reimbursed', reimbursed_by=request.user, reimbursed_at=now())
        self.message_user(request, f"💰 {count} claims marked as reimbursed")
    mark_reimbursed.short_description = "💰 Mark as Reimbursed"


# ============================================
# EXPENSE FORECAST ADMIN
# ============================================

@admin.register(ExpenseForecast)
class ExpenseForecastAdmin(admin.ModelAdmin):
    list_display = ('month', 'forecasted_amount', 'actual_amount', 'variance', 'variance_percent')
    list_filter = ('month',)
    search_fields = ('notes',)
    readonly_fields = ('variance', 'variance_percent')

# ============================================
# BUDGET ADMIN REGISTRATION
# ============================================

from .models import Budget, BudgetAllocation, BudgetTransaction, BudgetAlert, BudgetForecast

class BudgetAllocationInline(admin.TabularInline):
    model = BudgetAllocation
    fields = ('category', 'allocated_amount', 'used_amount', 'remaining_amount')
    readonly_fields = ('used_amount', 'remaining_amount')
    extra = 1


class BudgetTransactionInline(admin.TabularInline):
    model = BudgetTransaction
    fields = ('transaction_type', 'amount', 'balance_after', 'description')
    readonly_fields = ('balance_after', 'performed_at')
    extra = 0
    can_delete = False


@admin.register(Budget)
class BudgetAdmin(admin.ModelAdmin):
    list_display = ('budget_no', 'name', 'budget_type', 'department', 'project', 
                    'allocated_amount', 'used_amount', 'remaining_amount', 
                    'utilization_display', 'status_display', 'period_end')
    list_filter = ('budget_type', 'status', 'frequency', 'department')
    search_fields = ('budget_no', 'name', 'description')
    inlines = [BudgetAllocationInline, BudgetTransactionInline]
    
    fieldsets = (
        ('Budget Information', {
            'fields': ('budget_no', 'name', 'budget_type', 'description')
        }),
        ('Amounts', {
            'fields': ('allocated_amount', 'used_amount', 'remaining_amount', 'reserved_amount')
        }),
        ('Period', {
            'fields': ('period_start', 'period_end', 'frequency')
        }),
        ('Links', {
            'fields': ('department', 'project', 'category')
        }),
        ('Status', {
            'fields': ('status', 'alert_threshold', 'critical_threshold', 
                      'approved_by', 'approved_at', 'submitted_by', 'submitted_at')
        }),
        ('Settings', {
            'fields': ('allow_rollover', 'rollover_from')
        }),
        ('System', {
            'fields': ('created_by', 'created_at', 'updated_at', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ('budget_no', 'used_amount', 'remaining_amount', 'created_at', 'updated_at')
    
    def utilization_display(self, obj):
        percent = obj.utilization_percent()
        if percent >= obj.critical_threshold:
            color = 'red'
        elif percent >= obj.alert_threshold:
            color = 'orange'
        else:
            color = 'green'
        return format_html(
            '<span style="color: {}; font-weight: bold;">{:.1f}%</span>',
            color, percent
        )
    utilization_display.short_description = "Utilization"
    
    def status_display(self, obj):
        return format_html(obj.get_status_badge())
    status_display.short_description = "Status"
    
    actions = ['approve_budgets', 'activate_budgets', 'submit_budgets']
    
    def submit_budgets(self, request, queryset):
        for budget in queryset.filter(status='draft'):
            budget.submit(request.user)
        self.message_user(request, f"📤 {queryset.count()} budgets submitted")
    submit_budgets.short_description = "📤 Submit for Approval"
    
    def approve_budgets(self, request, queryset):
        for budget in queryset.filter(status='pending'):
            budget.approve(request.user)
        self.message_user(request, f"✅ {queryset.count()} budgets approved")
    approve_budgets.short_description = "✅ Approve Selected Budgets"
    
    def activate_budgets(self, request, queryset):
        for budget in queryset.filter(status='approved'):
            budget.activate(request.user)
        self.message_user(request, f"✅ {queryset.count()} budgets activated")
    activate_budgets.short_description = "✅ Activate Selected Budgets"


@admin.register(BudgetAlert)
class BudgetAlertAdmin(admin.ModelAdmin):
    list_display = ('budget', 'alert_type', 'message', 'status', 'triggered_at')
    list_filter = ('alert_type', 'status')
    search_fields = ('budget__name', 'message')
    list_editable = ('status',)
    
    actions = ['resolve_alerts']
    
    def resolve_alerts(self, request, queryset):
        for alert in queryset.filter(status__in=['new', 'read']):
            alert.resolve(request.user)
        self.message_user(request, f"✅ {queryset.count()} alerts resolved")
    resolve_alerts.short_description = "✅ Resolve Selected Alerts"


@admin.register(BudgetForecast)
class BudgetForecastAdmin(admin.ModelAdmin):
    list_display = ('budget', 'forecast_date', 'forecasted_amount', 'actual_amount', 
                    'variance', 'variance_percent', 'confidence_level')
    list_filter = ('forecast_date',)
    search_fields = ('budget__name',)